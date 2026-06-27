import streamlit as st
import pandas as pd
import numpy as np
import logging
import base64
from datetime import datetime
from modules.data import fetch_data
from modules.metrics import compute_core_metrics, compute_downside_metrics
from modules.topsis import TOPSIS
from modules.yuan import YuanYuan
from modules.charts import chart_cumulative_returns, chart_drawdown, chart_rolling_sharpe, chart_topsis_heatmap, chart_cmatrix, chart_scores_bar, chart_sensitivity_heatmap
from modules.report import generate_pdf
import streamlit.components.v1 as components
import plotly.graph_objects as go
from plotly.subplots import make_subplots


# ── Standalone ranking helpers ───────────────────────────────────────────────

def compute_naive_ranking(metrics_matrix):
    """Simple equal-weight ranking on 3 metrics using min-max normalisation."""
    naive_metrics = ["Ann. Return (%)", "Sharpe Ratio", "Max Drawdown (%)"]
    # Max Drawdown is stored as a negative number (-10% > -30%), so higher = less negative = better
    NAIVE_BENEFIT = ["Ann. Return (%)", "Sharpe Ratio", "Max Drawdown (%)"]
    NAIVE_COST    = []
    avail = [m for m in naive_metrics if m in metrics_matrix.index]
    if not avail:
        return pd.DataFrame()
    fund_cols = [c for c in metrics_matrix.columns if c != "Benchmark"]
    subset = metrics_matrix.loc[avail, fund_cols].astype(float)
    normalised = pd.DataFrame(index=avail, columns=fund_cols, dtype=float)
    for metric in avail:
        row = subset.loc[metric]
        rng = row.max() - row.min()
        if rng == 0:
            normalised.loc[metric] = 0.5
        elif metric in NAIVE_BENEFIT:
            normalised.loc[metric] = (row - row.min()) / rng
        else:
            normalised.loc[metric] = (row.max() - row) / rng
    scores = normalised.mean()
    scores_sorted = (scores * 100).sort_values(ascending=False)
    return pd.DataFrame({
        "Naive Score (%)": scores_sorted.round(1),
        "Naive Rank": range(1, len(scores_sorted) + 1)
    })


def compute_borda_ranking(metrics_matrix, active_weights):
    """Borda count ranking weighted by metric weights."""
    HIGHER_IS_BETTER = [
        "Ann. Return (%)", "Alpha (ann. %)",
        "Sharpe Ratio", "Sortino Ratio", "Treynor Ratio", "Information Ratio", "R²",
        "Upside Capture (%)", "Calmar Ratio", "ESG Globe Rating",
    ]
    fund_cols = [c for c in metrics_matrix.columns if c != "Benchmark"]
    n = len(fund_cols)
    borda_scores = pd.Series(0.0, index=fund_cols)
    for metric, weight in active_weights.items():
        if metric not in metrics_matrix.index:
            continue
        row = metrics_matrix.loc[metric, fund_cols].astype(float)
        ranked = row.rank(ascending=True, na_option="bottom") - 1 if metric in HIGHER_IS_BETTER else row.rank(ascending=False, na_option="bottom") - 1
        borda_scores += ranked * weight
    borda_sorted = borda_scores.sort_values(ascending=False)
    max_score = borda_sorted.max()
    normalised = (borda_sorted / max_score * 100) if max_score > 0 else borda_sorted
    return pd.DataFrame({
        "Borda Score (%)": normalised.round(1),
        "Borda Rank": range(1, n + 1)
    })


# ── Phase 3 TOPSIS / Rolling helpers ─────────────────────────────────────────

def _inline_topsis(metrics_matrix, fund_names, pillar_weights):
    """Correct in-app TOPSIS (avoids axis misalignment in module)."""
    BENEFITS = [
        "Ann. Return (%)", "Alpha (ann. %)", "Sharpe Ratio", "Sortino Ratio",
        "Treynor Ratio", "Information Ratio", "R²", "Upside Capture (%)",
        "Calmar Ratio", "ESG Globe Rating", "Max Drawdown (%)",
    ]
    PILLAR_METRICS = {
        'Returns': ['Ann. Return (%)', 'Alpha (ann. %)', 'Upside Capture (%)', 'Calmar Ratio'],
        'Risk-Adj': ['Sharpe Ratio', 'Sortino Ratio', 'Treynor Ratio', 'Information Ratio', 'R²'],
        'Risk/DD': ['Ann. Volatility (%)', 'Beta', 'Tracking Error (%)', 'Max Drawdown (%)',
                    'Max DD Duration (mths)', 'Downside Capture (%)'],
        'Costs': ['OCF'],
        'ESG': ['ESG Globe Rating', 'Carbon Risk Score'],
    }
    metric_weights = {}
    for pillar, metrics in PILLAR_METRICS.items():
        w = pillar_weights.get(pillar, 0)
        if w > 0:
            per_m = w / len(metrics)
            for m in metrics:
                metric_weights[m] = per_m
    fund_cols = [c for c in fund_names if c in metrics_matrix.columns]
    act_m = [m for m in metric_weights if m in metrics_matrix.index]
    if not act_m or not fund_cols:
        return pd.DataFrame()
    mm = metrics_matrix.loc[act_m, fund_cols].astype(float)
    row_l2 = np.sqrt((mm ** 2).sum(axis=1)).replace(0, 1)
    norm = mm.div(row_l2, axis=0)
    amw_s = pd.Series({m: metric_weights[m] for m in act_m})
    wmat = norm.mul(amw_s / amw_s.sum(), axis=0)
    Ap = pd.Series(index=act_m, dtype=float)
    Am = pd.Series(index=act_m, dtype=float)
    for m in act_m:
        r = wmat.loc[m]
        if m in BENEFITS:
            Ap[m] = r.max(); Am[m] = r.min()
        else:
            Ap[m] = r.min(); Am[m] = r.max()
    Dp = np.sqrt(wmat.sub(Ap, axis=0).pow(2).sum(axis=0))
    Dm = np.sqrt(wmat.sub(Am, axis=0).pow(2).sum(axis=0))
    _denom = Dp + Dm
    sc = (Dm / _denom).where(_denom != 0, 0.5)
    sc_s = sc.sort_values(ascending=False)
    return pd.DataFrame({'TOPSIS Score (%)': (sc_s * 100).round(1), 'Rank': range(1, len(sc_s) + 1)})


def _metrics_from_returns(win_fund, win_bench, fund_names, tickers, costs, esg_globe, carbon_risk, rf_annual):
    """Build a metrics_matrix DataFrame from windowed return Series."""
    core_m, down_m = {}, {}
    for j, name in enumerate(fund_names):
        ticker = tickers[j]
        if ticker not in win_fund.columns:
            continue
        try:
            core_m[name] = compute_core_metrics(win_fund[ticker], win_bench, rf_annual)
            down_m[name] = compute_downside_metrics(win_fund[ticker], win_bench)
        except Exception as _e:
            logging.warning("Metric computation skipped for fund '%s': %s", name, _e)
            continue
    if not core_m:
        return pd.DataFrame()
    core_df = pd.DataFrame(core_m).T
    down_df = pd.DataFrame(down_m).T
    core_df['OCF'] = [costs.get(n, np.nan) for n in core_df.index]
    core_df['ESG Globe Rating'] = [esg_globe.get(n, np.nan) for n in core_df.index]
    core_df['Carbon Risk Score'] = [carbon_risk.get(n, np.nan) for n in core_df.index]
    return pd.concat([core_df, down_df], axis=1).T


def compute_rolling_rankings(fund_returns, benchmark_returns, fund_names, tickers,
                              costs, esg_globe, carbon_risk, rf_annual,
                              pillar_weights, window=36, step=6):
    """TOPSIS over rolling windows. Returns {date_str: DataFrame(TOPSIS Score (%), Rank)}."""
    dates = fund_returns.index
    results = {}
    i = window - 1
    while i < len(dates):
        win_fund = fund_returns.iloc[i - window + 1:i + 1]
        win_bench = benchmark_returns.iloc[i - window + 1:i + 1]
        win_matrix = _metrics_from_returns(win_fund, win_bench, fund_names, tickers,
                                           costs, esg_globe, carbon_risk, rf_annual)
        if not win_matrix.empty:
            ranking = _inline_topsis(win_matrix, fund_names, pillar_weights)
            if not ranking.empty:
                results[str(dates[i].date())] = ranking
        i += step
    return results


def run_period_topsis(fund_returns, benchmark_returns, fund_names, tickers,
                      costs, esg_globe, carbon_risk, rf_annual, pillar_weights,
                      start_date, end_date):
    """Run TOPSIS on a specific date range."""
    mask = (fund_returns.index >= pd.Timestamp(start_date)) & (fund_returns.index <= pd.Timestamp(end_date))
    win_fund = fund_returns.loc[mask]
    win_bench = benchmark_returns.loc[mask]
    if len(win_fund) < 6:
        return pd.DataFrame()
    win_matrix = _metrics_from_returns(win_fund, win_bench, fund_names, tickers,
                                       costs, esg_globe, carbon_risk, rf_annual)
    if win_matrix.empty:
        return pd.DataFrame()
    return _inline_topsis(win_matrix, fund_names, pillar_weights)


def run_topsis_on_filtered_returns(fund_returns, benchmark_returns, fund_names, tickers,
                                   costs, esg_globe, carbon_risk, rf_annual, pillar_weights,
                                   date_ranges, exclude=False):
    """Run TOPSIS on filtered returns (include or exclude date ranges)."""
    if exclude:
        mask = pd.Series(True, index=fund_returns.index)
        for start, end in date_ranges:
            mask &= ~((fund_returns.index >= pd.Timestamp(start)) & (fund_returns.index <= pd.Timestamp(end)))
    else:
        mask = pd.Series(False, index=fund_returns.index)
        for start, end in date_ranges:
            mask |= ((fund_returns.index >= pd.Timestamp(start)) & (fund_returns.index <= pd.Timestamp(end)))
    win_fund = fund_returns.loc[mask]
    win_bench = benchmark_returns.loc[mask]
    if len(win_fund) < 6:
        return pd.DataFrame()
    win_matrix = _metrics_from_returns(win_fund, win_bench, fund_names, tickers,
                                       costs, esg_globe, carbon_risk, rf_annual)
    if win_matrix.empty:
        return pd.DataFrame()
    return _inline_topsis(win_matrix, fund_names, pillar_weights)


def compute_bootstrap_ci(fund_returns, benchmark_returns, pillar_weights,
                          costs_dict, esg_globe_dict, carbon_risk_dict,
                          n_bootstrap=200, ci=90):
    """
    Resamples monthly returns with replacement n_bootstrap times.
    Recomputes TOPSIS ranking for each resample.
    Returns a dict mapping fund name -> (mean_score, lower_ci, upper_ci, mean_rank).
    fund_returns must have fund names as columns (not tickers).
    """
    fund_names = list(fund_returns.columns)
    n_months   = len(fund_returns)
    score_records = {f: [] for f in fund_names}
    rank_records  = {f: [] for f in fund_names}
    n_failed = 0

    for _ in range(n_bootstrap):
        idx = np.random.randint(0, n_months, size=n_months)
        boot_fund = fund_returns.iloc[idx].reset_index(drop=True)
        boot_bm   = benchmark_returns.iloc[idx].reset_index(drop=True)
        try:
            core_r = {}
            down_r = {}
            for fund in fund_names:
                core_r[fund] = compute_core_metrics(boot_fund[fund], boot_bm)
                down_r[fund] = compute_downside_metrics(boot_fund[fund], boot_bm)
            core_mm = pd.DataFrame(core_r)
            down_mm = pd.DataFrame(down_r)
            matrix = pd.DataFrame({
                fund: {
                    "Ann. Return (%)":        (np.exp(boot_fund[fund].mean() * 12) - 1) * 100,
                    "Alpha (ann. %)":         core_mm.loc["Alpha (ann. %)", fund],
                    "Sharpe Ratio":           core_mm.loc["Sharpe Ratio", fund],
                    "Sortino Ratio":          core_mm.loc["Sortino Ratio", fund],
                    "Treynor Ratio":          core_mm.loc["Treynor Ratio", fund],
                    "Information Ratio":      core_mm.loc["Information Ratio", fund],
                    "R²":                     core_mm.loc["R²", fund],
                    "Ann. Volatility (%)":    core_mm.loc["Ann. Volatility (%)", fund],
                    "Beta":                   core_mm.loc["Beta", fund],
                    "Tracking Error (%)":     core_mm.loc["Tracking Error (%)", fund],
                    "Max Drawdown (%)":       down_mm.loc["Max Drawdown (%)", fund],
                    "Max DD Duration (mths)": down_mm.loc["Max DD Duration (mths)", fund],
                    "Downside Capture (%)":   down_mm.loc["Downside Capture (%)", fund],
                    "Upside Capture (%)":     down_mm.loc["Upside Capture (%)", fund],
                    "Calmar Ratio":           down_mm.loc["Calmar Ratio", fund],
                    "OCF":                    costs_dict.get(fund, 0.99),
                    "ESG Globe Rating":       esg_globe_dict.get(fund, 3),
                    "Carbon Risk Score":      carbon_risk_dict.get(fund, 7.0),
                }
                for fund in fund_names
            })
            ranking = _inline_topsis(matrix, fund_names, pillar_weights)
            for fund in fund_names:
                if fund in ranking.index:
                    score_records[fund].append(float(ranking.loc[fund, "TOPSIS Score (%)"]))
                    rank_records[fund].append(int(ranking.loc[fund, "Rank"]))
        except Exception as _e:
            n_failed += 1
            logging.warning("Bootstrap iteration failed: %s", _e)
            continue

    if n_failed > 0:
        st.warning(
            f"Bootstrap: {n_failed} of {n_bootstrap} resamples failed and were skipped. "
            f"Results are based on {n_bootstrap - n_failed} iterations."
        )

    alpha_p = (100 - ci) / 2
    results = {}
    for fund in fund_names:
        scores = score_records[fund]
        ranks  = rank_records[fund]
        if scores:
            results[fund] = {
                "mean_score": np.mean(scores),
                "lower_ci":   np.percentile(scores, alpha_p),
                "upper_ci":   np.percentile(scores, 100 - alpha_p),
                "mean_rank":  np.mean(ranks),
                "rank_std":   np.std(ranks),
            }
    return results


# ── Excel export ─────────────────────────────────────────────────────────────

def generate_excel(fund_names, sample_period, core_metrics_df, downside_df,
                   topsis_ranking, yuan_ranking, metrics_matrix,
                   naive_ranking=None, borda_ranking=None):
    """
    Generates a formatted .xlsx file with multiple sheets.
    Returns bytes.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    TEAL       = "154D57"
    TEAL_LIGHT = "EAF2F3"
    WARM       = "F5F0EB"
    BORDER_COL = "E8DDD3"
    WHITE      = "FFFFFF"
    GREEN_BG   = "D4EDDA"
    RED_BG     = "FAE0E0"

    header_font    = Font(name="Calibri", bold=True, color=WHITE, size=10)
    header_fill    = PatternFill("solid", fgColor=TEAL)
    subheader_font = Font(name="Calibri", bold=True, color=TEAL, size=10)
    subheader_fill = PatternFill("solid", fgColor=TEAL_LIGHT)
    body_font      = Font(name="Calibri", size=10)
    title_font     = Font(name="Calibri", bold=True, size=13, color=TEAL)
    warm_fill      = PatternFill("solid", fgColor=WARM)
    green_fill     = PatternFill("solid", fgColor=GREEN_BG)
    red_fill       = PatternFill("solid", fgColor=RED_BG)
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left", vertical="center")
    right_align    = Alignment(horizontal="right", vertical="center")

    thin   = Side(style="thin", color=BORDER_COL)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num, n_cols):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border

    def style_data_row(ws, row_num, n_cols, zebra=False):
        fill = warm_fill if zebra else PatternFill("solid", fgColor=WHITE)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = body_font
            cell.fill      = fill
            cell.alignment = right_align if col > 1 else left
            cell.border    = border

    def auto_width(ws, min_w=10, max_w=35):
        for col in ws.columns:
            length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

    def write_df_to_sheet(ws, df, title, start_row=1):
        ws.cell(row=start_row, column=1, value=title).font = title_font
        start_row += 1
        headers = ["Metric"] + list(df.columns)
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col_idx, value=h)
        style_header_row(ws, start_row, len(headers))
        start_row += 1
        for row_idx, (idx, row) in enumerate(df.iterrows()):
            ws.cell(row=start_row, column=1, value=str(idx))
            for col_idx, val in enumerate(row, 2):
                try:
                    ws.cell(row=start_row, column=col_idx, value=round(float(val), 4))
                except (ValueError, TypeError):
                    ws.cell(row=start_row, column=col_idx, value=str(val))
            style_data_row(ws, start_row, len(headers), zebra=(row_idx % 2 == 1))
            start_row += 1
        return start_row

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.cell(row=1, column=1, value="Fund Analysis Engine  -  Summary Report").font = Font(
        name="Calibri", bold=True, size=16, color=TEAL)
    ws1.cell(row=2, column=1, value=f"Sample Period: {sample_period}").font = body_font
    ws1.cell(row=3, column=1, value=f"Funds: {', '.join(fund_names)}").font = body_font
    ws1.cell(row=4, column=1, value=f"Generated: {pd.Timestamp.now().strftime('%d %B %Y')}").font = body_font
    ws1.row_dimensions[1].height = 24

    row = 6
    ws1.cell(row=row, column=1, value="TOPSIS Rankings").font = subheader_font
    ws1.cell(row=row, column=1).fill = subheader_fill
    row += 1

    topsis_funds = topsis_ranking.loc[topsis_ranking.index.isin(set(fund_names))]
    topsis_display = topsis_funds.sort_values("Rank") if "Rank" in topsis_funds.columns else topsis_funds

    for col_idx, h in enumerate(["Fund", "TOPSIS Score", "Rank"], 1):
        ws1.cell(row=row, column=col_idx, value=h)
    style_header_row(ws1, row, 3)
    row += 1

    for i, (idx, r) in enumerate(topsis_display.iterrows()):
        score_col = "Score" if "Score" in r.index else r.index[0]
        rank_val  = int(r.get("Rank", i + 1))
        score_val = round(float(r.get(score_col, 0)), 4)
        ws1.cell(row=row, column=1, value=str(idx))
        ws1.cell(row=row, column=2, value=score_val)
        ws1.cell(row=row, column=3, value=rank_val)
        for col in range(1, 4):
            ws1.cell(row=row, column=col).font   = body_font
            ws1.cell(row=row, column=col).border = border
            ws1.cell(row=row, column=col).fill   = green_fill if rank_val == 1 else (
                red_fill if rank_val == len(fund_names) else
                PatternFill("solid", fgColor=WHITE if i % 2 == 0 else WARM))
        row += 1

    row += 1
    ws1.cell(row=row, column=1, value="Yuan & Yuan Rankings").font = subheader_font
    ws1.cell(row=row, column=1).fill = subheader_fill
    row += 1

    yuan_funds   = yuan_ranking.loc[yuan_ranking.index.isin(set(fund_names))]
    yuan_display = yuan_funds.sort_values("Rank") if "Rank" in yuan_funds.columns else yuan_funds

    for col_idx, h in enumerate(["Fund", "Yuan Score", "Rank"], 1):
        ws1.cell(row=row, column=col_idx, value=h)
    style_header_row(ws1, row, 3)
    row += 1

    for i, (idx, r) in enumerate(yuan_display.iterrows()):
        score_col = "Score" if "Score" in r.index else r.index[0]
        rank_val  = int(r.get("Rank", i + 1))
        ws1.cell(row=row, column=1, value=str(idx))
        ws1.cell(row=row, column=2, value=round(float(r.get(score_col, 0)), 6))
        ws1.cell(row=row, column=3, value=rank_val)
        for col in range(1, 4):
            ws1.cell(row=row, column=col).font   = body_font
            ws1.cell(row=row, column=col).border = border
            ws1.cell(row=row, column=col).fill   = green_fill if rank_val == 1 else (
                red_fill if rank_val == len(fund_names) else
                PatternFill("solid", fgColor=WHITE if i % 2 == 0 else WARM))
        row += 1

    auto_width(ws1)

    # Sheet 2: Core Metrics
    ws2 = wb.create_sheet("Core Metrics")
    write_df_to_sheet(ws2, core_metrics_df, "Core Risk-Adjusted Metrics")
    auto_width(ws2)

    # Sheet 3: Downside Metrics
    ws3 = wb.create_sheet("Downside Metrics")
    write_df_to_sheet(ws3, downside_df, "Downside Risk Metrics")
    auto_width(ws3)

    # Sheet 4: Full Metrics Matrix
    ws4 = wb.create_sheet("Metrics Matrix")
    write_df_to_sheet(ws4, metrics_matrix, "Full 19-Metric Evaluation Matrix")
    auto_width(ws4)

    # Sheet 5: All Rankings
    ws5 = wb.create_sheet("All Rankings")
    r5 = 1
    if naive_ranking is not None:
        r5 = write_df_to_sheet(ws5, naive_ranking, "Naive Ranking", start_row=r5)
        r5 += 2
    if borda_ranking is not None:
        write_df_to_sheet(ws5, borda_ranking, "Borda Count Ranking", start_row=r5)
    auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Phase 4 UI helpers ────────────────────────────────────────────────────────

def hex_to_rgba(hex_colour, alpha=0.3):
    """Convert a hex colour string to rgba() format for Plotly."""
    hex_colour = hex_colour.lstrip('#')
    if len(hex_colour) == 6:
        r, g, b = int(hex_colour[0:2], 16), int(hex_colour[2:4], 16), int(hex_colour[4:6], 16)
    elif len(hex_colour) == 3:
        r = int(hex_colour[0] * 2, 16)
        g = int(hex_colour[1] * 2, 16)
        b = int(hex_colour[2] * 2, 16)
    else:
        return f"rgba(100,100,100,{alpha})"
    return f"rgba({r},{g},{b},{alpha})"


def auto_commentary(template, **kwargs):
    """Renders a styled auto-generated commentary block below a chart."""
    try:
        text = template.format(**kwargs)
        st.markdown(
            '<div style="background:#F5F3F0;border-left:3px solid #C8A96E;border-radius:0 8px 8px 0;'
            'padding:0.75rem 1rem;margin-top:0.5rem;margin-bottom:1rem;">'
            '<p style="margin:0;font-size:0.82rem;color:#5A5A5A;line-height:1.6;">'
            '💬 <em>' + text + '</em>'
            '</p></div>',
            unsafe_allow_html=True
        )
    except Exception as _e:
        logging.debug("auto_commentary render skipped (non-critical): %s", _e)


def section_header(icon, subtitle, title, anchor=None):
    anchor_html = f'<span id="{anchor}"></span>' if anchor else ''
    st.markdown(
        f'{anchor_html}<div style="margin-bottom:1.5rem;">'
        f'<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;'
        f'color:#9B8C7E;margin-bottom:0.35rem;">{icon} {subtitle}</p>'
        f'<h1 style="font-size:2rem;font-weight:700;color:#0A0A0A;margin:0;letter-spacing:-0.03em;">{title}</h1>'
        f'</div>',
        unsafe_allow_html=True
    )


def chart_card(title, description, fig):
    st.markdown(
        f'<div style="margin-bottom:0.4rem;">'
        f'<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;'
        f'color:#9B8C7E;margin-bottom:0.2rem;">{title}</p>'
        f'<p style="font-size:0.78rem;color:#7A6F65;margin:0 0 0.5rem 0;line-height:1.5;">{description}</p>'
        f'</div>',
        unsafe_allow_html=True
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


# ── Phase 4 chart functions ───────────────────────────────────────────────────

def _assign_textpositions(pts):
    """Assign per-point Plotly textposition strings to reduce label overlap.
    pts: dict of name -> (x, y). Returns dict of name -> textposition string.
    """
    if not pts:
        return {}
    names = list(pts.keys())
    x_span = (max(v[0] for v in pts.values()) - min(v[0] for v in pts.values())) or 1.0
    sorted_names = sorted(names, key=lambda nm: pts[nm][0])
    count = len(sorted_names)
    result = {}
    for rank, name in enumerate(sorted_names):
        x, y = pts[name]
        if rank == 0:
            h = "right"
        elif rank == count - 1:
            h = "left"
        else:
            h = "center"
        close = [
            pts[nm][1] for nm in names
            if nm != name and abs(pts[nm][0] - x) / x_span < 0.25
        ]
        v = "bottom" if (close and y <= sum(close) / len(close)) else "top"
        result[name] = f"{v} {h}"
    return result


def chart_risk_return_scatter(core_metrics_df, colours, plotly_layout):
    """Scatter: x = Ann. Volatility (%), y = Ann. Return (%), size = Sharpe Ratio."""
    fig = go.Figure()
    funds = [c for c in core_metrics_df.columns if c != "Benchmark"]

    # Collect all coordinates upfront for axis range and label placement
    all_pts = {}
    for fund in funds:
        all_pts[fund] = (
            float(core_metrics_df.loc["Ann. Volatility (%)", fund]),
            float(core_metrics_df.loc["Ann. Return (%)", fund]),
        )
    if "Benchmark" in core_metrics_df.columns:
        all_pts["Benchmark"] = (
            float(core_metrics_df.loc["Ann. Volatility (%)", "Benchmark"]),
            float(core_metrics_df.loc["Ann. Return (%)", "Benchmark"]),
        )
    text_pos = _assign_textpositions(all_pts)

    # Explicit axis ranges with padding so edge labels are never clipped
    all_x = [v[0] for v in all_pts.values()]
    all_y = [v[1] for v in all_pts.values()]
    x_span = (max(all_x) - min(all_x)) or 1.0
    y_span = (max(all_y) - min(all_y)) or 1.0
    x_range = [min(all_x) - 0.25 * x_span, max(all_x) + 0.40 * x_span]
    y_range = [min(all_y) - 0.20 * y_span, max(all_y) + 0.45 * y_span]

    for fund in funds:
        x, y = all_pts[fund]
        sharpe = float(core_metrics_df.loc["Sharpe Ratio", fund])
        size = max(sharpe * 60, 20)
        colour = colours.get(fund, "#9B9B9B")
        fig.add_trace(go.Scatter(
            x=[x], y=[y],
            mode="markers+text",
            name=fund,
            text=[fund],
            textposition=text_pos.get(fund, "top center"),
            textfont=dict(size=11, color="#1A1A1A"),
            marker=dict(size=size, color=colour, opacity=0.85, line=dict(width=2, color="#FFFFFF")),
            hovertemplate=(
                f"<b>{fund}</b><br>"
                f"Volatility: %{{x:.2f}}%<br>"
                f"Return: %{{y:.2f}}%<br>"
                f"Sharpe: {sharpe:.3f}<extra></extra>"
            )
        ))
    if "Benchmark" in core_metrics_df.columns:
        bx, by = all_pts["Benchmark"]
        fig.add_trace(go.Scatter(
            x=[bx], y=[by],
            mode="markers+text",
            name="Benchmark",
            text=["Benchmark"],
            textposition=text_pos.get("Benchmark", "top center"),
            textfont=dict(size=10, color="#9B9B9B"),
            marker=dict(size=18, color="#9B9B9B", symbol="diamond", line=dict(width=2, color="#FFFFFF")),
            hovertemplate="<b>Benchmark</b><br>Volatility: %{x:.2f}%<br>Return: %{y:.2f}%<extra></extra>"
        ))
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Risk-Return Profile"),
        xaxis=dict(title="Annualised Volatility (%)", gridcolor="#F0EFEC", range=x_range),
        yaxis=dict(title="Annualised Return (%)", gridcolor="#F0EFEC", range=y_range),
        showlegend=False,
        height=480,
    )
    return fig


def chart_factor_attribution(core_metrics_df, benchmark_returns, fund_returns, colours, plotly_layout):
    """Stacked bar: beta return vs alpha contribution per fund."""
    funds = [c for c in core_metrics_df.columns if c != "Benchmark"]
    bm_ann_return = float((np.exp(benchmark_returns.mean() * 12) - 1) * 100)
    alpha_returns, beta_returns, fund_labels = [], [], []
    for fund in funds:
        alpha = float(core_metrics_df.loc["Alpha (ann. %)", fund])
        beta = float(core_metrics_df.loc["Beta", fund])
        b_ret = beta * bm_ann_return
        alpha_returns.append(round(alpha, 2))
        beta_returns.append(round(b_ret, 2))
        fund_labels.append(fund)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Beta Return (Market)",
        x=fund_labels,
        y=beta_returns,
        marker_color="#94A3B8",
        text=[f"{v:.1f}%" for v in beta_returns],
        textposition="inside",
        textfont=dict(color="#FFFFFF", size=11),
        hovertemplate="<b>%{x}</b><br>Beta Return: %{y:.2f}%<extra></extra>"
    ))
    fig.add_trace(go.Bar(
        name="Alpha (Manager Skill)",
        x=fund_labels,
        y=alpha_returns,
        marker_color=[colours.get(f, "#0D5C63") for f in fund_labels],
        text=[f"{v:.1f}%" for v in alpha_returns],
        textposition="inside",
        textfont=dict(color="#FFFFFF", size=11),
        hovertemplate="<b>%{x}</b><br>Alpha: %{y:.2f}%<extra></extra>"
    ))
    fig.add_hline(y=0, line_dash="dot", line_color="#9B9B9B", opacity=0.6)
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Return Attribution: Alpha vs Beta"),
        barmode="stack",
        xaxis=dict(title="Fund"),
        yaxis=dict(title="Annualised Return (%)", gridcolor="#F0EFEC"),
        height=420,
    )
    return fig


def chart_drawdown_recovery(fund_returns, benchmark_returns, colours, plotly_layout):
    """Annotated drawdown chart with max-drawdown labels per series."""
    fig = go.Figure()
    all_series = dict(fund_returns)
    all_series["Benchmark"] = benchmark_returns
    # Stagger annotation offsets so labels spread out and reduce overlap
    _ax = [35, -45, 55, -30, 45]
    _ay = [-22, -38, -12, -50, -28]
    for idx, (name, returns) in enumerate(all_series.items()):
        colour = colours.get(name, "#9B9B9B")
        wealth = np.exp(returns.cumsum())
        peak = wealth.cummax()
        dd = (wealth - peak) / peak * 100
        dash = "dot" if name == "Benchmark" else "solid"
        fill_colour = hex_to_rgba(colour, alpha=0.08)
        fig.add_trace(go.Scatter(
            x=dd.index, y=dd.values,
            name=name,
            mode="lines",
            line=dict(color=colour, width=2, dash=dash),
            fill="tozeroy" if name != "Benchmark" else None,
            fillcolor=fill_colour,
            opacity=0.9,
            hovertemplate=f"<b>{name}</b><br>Date: %{{x}}<br>Drawdown: %{{y:.2f}}%<extra></extra>"
        ))
        min_idx = dd.idxmin()
        min_val = dd.min()
        fig.add_annotation(
            x=min_idx, y=min_val,
            text=f"{name}: {min_val:.1f}%",
            showarrow=True,
            arrowhead=2,
            arrowcolor=colour,
            arrowsize=0.8,
            ax=_ax[idx % len(_ax)], ay=_ay[idx % len(_ay)],
            font=dict(size=9, color=colour),
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor=colour,
            borderwidth=1,
            borderpad=3,
        )
    fig.add_hline(y=0, line_dash="dash", line_color="#9B9B9B", opacity=0.4)
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Drawdown Recovery Timeline"),
        xaxis=dict(title="Date", gridcolor="#F0EFEC"),
        yaxis=dict(title="Drawdown (%)", gridcolor="#F0EFEC"),
        height=560,
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.12,
            xanchor="center",
            x=0.5,
        ),
        margin=dict(b=75),
    )
    return fig


def chart_monthly_heatmap(fund_returns, colours, plotly_layout):
    """Calendar heatmap of monthly returns, one subplot per fund."""
    funds = list(fund_returns.columns)
    n = len(funds)
    fig = make_subplots(
        rows=n, cols=1,
        subplot_titles=funds,
        vertical_spacing=0.05 / max(n - 1, 1),
    )
    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, fund in enumerate(funds):
        series = fund_returns[fund].copy()
        series.index = pd.to_datetime(series.index)
        years = sorted(series.index.year.unique())
        matrix, text_vals, hover_vals = [], [], []
        for year in years:
            row, trow, hrow = [], [], []
            for month in range(1, 13):
                mask = (series.index.year == year) & (series.index.month == month)
                if mask.sum() > 0:
                    val = float(series[mask].values[0]) * 100
                    row.append(round(val, 2))
                    trow.append(f"{val:.1f}%")
                    hrow.append(f"{month_labels[month - 1]} {year}: {val:+.2f}%")
                else:
                    row.append(None)
                    trow.append("")
                    hrow.append("No data")
            matrix.append(row)
            text_vals.append(trow)
            hover_vals.append(hrow)
        fig.add_trace(
            go.Heatmap(
                z=matrix,
                x=month_labels,
                y=[str(y) for y in years],
                text=text_vals,
                texttemplate="%{text}",
                textfont=dict(size=9),
                customdata=hover_vals,
                hovertemplate="%{customdata}<extra></extra>",
                hoverongaps=True,
                colorscale=[
                    [0.0, "#DC2626"],
                    [0.4, "#FCA5A5"],
                    [0.5, "#F5F5F5"],
                    [0.6, "#86EFAC"],
                    [1.0, "#059669"],
                ],
                zmid=0,
                showscale=(i == 0),
                colorbar=dict(
                    title=dict(text="Return %", font=dict(size=10)),
                    len=0.75,
                    y=0.5,
                    yanchor="middle",
                    thickness=12,
                    tickfont=dict(size=9),
                    x=1.02,
                    xanchor="left",
                ),
                name=fund,
            ),
            row=i + 1, col=1
        )
    # Compact tick fonts across all axes
    fig.update_xaxes(tickfont=dict(size=9))
    fig.update_yaxes(tickfont=dict(size=9))
    # Smaller subplot panel titles
    fig.update_annotations(font_size=11)
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Monthly Returns Heatmap"),
        height=160 * n,
        margin=dict(r=90),
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#F3F3F3",
    )
    return fig


def chart_return_distribution(fund_returns, colours, plotly_layout):
    """Violin + box overlay of monthly return distributions per fund."""
    fig = go.Figure()
    for fund in fund_returns.columns:
        colour = colours.get(fund, "#9B9B9B")
        returns_pct = fund_returns[fund] * 100
        fill_colour = hex_to_rgba(colour, alpha=0.3)
        fig.add_trace(go.Violin(
            x=[fund] * len(returns_pct),
            y=returns_pct,
            name=fund,
            box_visible=True,
            meanline_visible=True,
            fillcolor=fill_colour,
            line_color=colour,
            opacity=0.8,
            hovertemplate=f"<b>{fund}</b><br>Return: %{{y:.2f}}%<extra></extra>",
            showlegend=False,
        ))
    fig.add_hline(y=0, line_dash="dot", line_color="#9B9B9B", opacity=0.5)
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Monthly Return Distribution"),
        xaxis=dict(title="Fund"),
        yaxis=dict(title="Monthly Return (%)", gridcolor="#F0EFEC"),
        height=480,
        violinmode="overlay",
    )
    return fig


def chart_correlation_heatmap(log_returns, plotly_layout):
    """Annotated pairwise correlation matrix for all series.
    Two-trace approach: main heatmap (off-diagonal, focused scale) +
    diagonal overlay (fixed neutral colour, shows 1.000 text).
    """
    corr_full = log_returns.corr()
    logging.debug("Correlation matrix (unrounded):\n%s", corr_full.to_string())
    corr = corr_full.round(3)
    labels = corr.columns.tolist()
    n = len(labels)

    # Writable NumPy arrays -- to_numpy(copy=True) avoids read-only buffer issues
    corr_values = corr_full.to_numpy(dtype=float, copy=True)

    # Trace 1: off-diagonal only (diagonal = NaN, filled by overlay)
    main_z = corr_values.copy()
    np.fill_diagonal(main_z, np.nan)

    # Trace 2: diagonal only (all other cells = NaN)
    diag_z = np.full_like(corr_values, np.nan, dtype=float)
    np.fill_diagonal(diag_z, 1.0)

    # Text matrices: no overlap between traces
    off_diag_text = [
        ["" if r == c else f"{corr.iloc[r, c]:.3f}" for c in range(n)]
        for r in range(n)
    ]
    diag_text = [
        ["1.000" if r == c else "" for c in range(n)]
        for r in range(n)
    ]

    # Trace 1: main heatmap -- off-diagonal correlations with focused colour scale
    fig = go.Figure(data=go.Heatmap(
        z=main_z,
        x=labels,
        y=labels,
        text=off_diag_text,
        texttemplate="%{text}",
        textfont=dict(size=11, color="#1A1A1A"),
        colorscale=[
            [0.0, "#FCA5A5"],
            [0.5, "#F5F5F5"],
            [1.0, "#059669"],
        ],
        zmin=0.80, zmax=1.00,
        colorbar=dict(
            title=dict(text="Corr.", font=dict(color="#9B9B9B", size=10)),
            tickvals=[0.80, 0.85, 0.90, 0.95, 1.00],
            tickformat=".2f",
            tickfont=dict(size=9, color="#9B9B9B"),
            x=1.02,
            xanchor="left",
            thickness=12,
            len=0.8,
        ),
        hoverongaps=False,
    ))

    # Trace 2: diagonal overlay -- fixed warm-neutral colour, "1.000" label
    fig.add_trace(go.Heatmap(
        z=diag_z,
        x=labels,
        y=labels,
        text=diag_text,
        texttemplate="%{text}",
        textfont=dict(size=11, color="#6B6B6B"),
        colorscale=[[0.0, "#F6F2EC"], [1.0, "#F6F2EC"]],
        zmin=0.9, zmax=1.1,
        showscale=False,
        hoverongaps=False,
        hovertemplate="<b>%{x}</b><br>Self-correlation: 1.000<extra></extra>",
    ))

    fig.add_annotation(
        text="Colour scale 0.80-1.00. Diagonal shows self-correlation (1.000).",
        xref="paper", yref="paper",
        x=0.0, y=-0.30,
        showarrow=False,
        font=dict(size=9, color="#9B9B9B"),
        xanchor="left",
    )
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Return Correlation Matrix"),
        xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
        yaxis=dict(autorange="reversed", tickfont=dict(size=10)),
        height=480,
        margin=dict(b=120, r=90),
        plot_bgcolor="#FFFFFF",
        showlegend=False,
    )
    return fig


def chart_rolling_correlation(fund_returns, benchmark_returns, colours, plotly_layout, window=12):
    """Rolling 12-month correlation of each fund with the benchmark."""
    fig = go.Figure()
    for fund in fund_returns.columns:
        rolling_corr = fund_returns[fund].rolling(window).corr(benchmark_returns)
        colour = colours.get(fund, "#9B9B9B")
        fig.add_trace(go.Scatter(
            x=rolling_corr.index,
            y=rolling_corr.values,
            name=fund,
            mode="lines",
            line=dict(color=colour, width=2.5),
            hovertemplate=f"<b>{fund}</b><br>Date: %{{x}}<br>Correlation: %{{y:.3f}}<extra></extra>"
        ))
    fig.add_hline(y=1.0, line_dash="dot", line_color="#9B9B9B", opacity=0.3)
    layout = {k: v for k, v in plotly_layout.items() if k != "title"}
    fig.update_layout(**layout)
    fig.update_layout(
        title=dict(text="Rolling 12-Month Correlation with Benchmark"),
        xaxis=dict(title="Date", gridcolor="#F0EFEC"),
        yaxis=dict(title="Correlation", range=[0, 1.05], gridcolor="#F0EFEC"),
        height=420,
    )
    return fig


# Page config
st.set_page_config(
    page_title="Fund Analysis Engine",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500;9..40,600;9..40,700&display=swap');

/* Scoped font — does not touch Streamlit internals or Material icon elements */
html, body, .stApp, .stMarkdown, p, h1, h2, h3, h4, h5, h6,
.stButton, .stSelectbox, .stTextInput, .stNumberInput,
.stRadio, .stSlider, .stTabs, .stExpander, label, .element-container {
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif !important;
}

/* Restore Material Icons so dataframe column-menu icons render as symbols not text */
.material-icons,
.material-icons-outlined,
.material-symbols-outlined,
.material-symbols-rounded,
span[class*="material-icons"],
span[class*="material-symbols"] {
    font-family: 'Material Symbols Rounded', 'Material Symbols Outlined', 'Material Icons' !important;
    font-weight: normal !important;
    font-style: normal !important;
    font-size: 20px !important;
    line-height: 1 !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}

.stApp { background: #FFFFFF !important; color: #0A0A0A; }

section[data-testid="stSidebar"] {
    background: #FEFAF7 !important;
    border-right: 1px solid #E8DDD3 !important;
}
section[data-testid="stSidebar"] * { color: #0A0A0A !important; }
.sidebar-alert, .sidebar-alert * { color: #C0392B !important; }

.main .block-container { padding: 0.75rem 4rem 0 4rem; max-width: 1400px; }

h1 { font-size: 1.75rem !important; font-weight: 700 !important; letter-spacing: -0.02em !important; color: #0A0A0A !important; margin-bottom: 0.25rem !important; line-height: 1.2 !important; }
h2 { font-size: 1.25rem !important; font-weight: 600 !important; letter-spacing: -0.015em !important; color: #0A0A0A !important; margin-top: 2rem !important; line-height: 1.2 !important; }
h3 { font-size: 1rem !important; font-weight: 600 !important; color: #0A0A0A !important; letter-spacing: -0.01em !important; line-height: 1.3 !important; }
p, .stMarkdown p { color: #7A6F65; font-size: 0.875rem; line-height: 1.5; }

div[data-testid="stMetric"] {
    background: #FFFFFF;
    border: 1px solid #E8DDD3;
    border-radius: 16px;
    padding: 1.5rem !important;
    transition: transform 0.15s ease-out, box-shadow 0.18s ease-out !important;
}
div[data-testid="stMetric"]:hover { box-shadow: 0 8px 28px rgba(21,77,87,0.14) !important; transform: translateY(-3px) !important; }
div[data-testid="stMetric"] label { color: #9B8C7E !important; font-size: 0.6875rem !important; font-weight: 700 !important; text-transform: uppercase !important; letter-spacing: 0.12em !important; }
div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color: #0A0A0A !important; font-size: 1.75rem !important; font-weight: 700 !important; letter-spacing: -0.02em !important; font-variant-numeric: tabular-nums !important; }
div[data-testid="stDataFrame"] td { font-variant-numeric: tabular-nums; }

.stButton > button { background: #154D57 !important; color: #FFFFFF !important; border: none !important; border-radius: 980px !important; padding: 0.65rem 1.8rem !important; font-size: 0.875rem !important; font-weight: 600 !important; transition: all 0.2s ease !important; letter-spacing: 0.01em !important; }
.stButton > button:hover { background: #0F3940 !important; transform: scale(1.02) !important; box-shadow: 0 4px 15px rgba(21,77,87,0.3) !important; }
.stButton > button p, .stButton > button span, .stButton > button div { color: #FFFFFF !important; }
.stDownloadButton > button p, .stDownloadButton > button span, .stDownloadButton > button div { color: #FFFFFF !important; }

.stTextInput > div > div > input,
.stNumberInput > div > div > input {
    background: #FFFFFF !important;
    border: none !important;
    border-bottom: 1.5px solid #D4C3B0 !important;
    border-radius: 0 !important;
    color: #0A0A0A !important;
    padding: 0.45rem 0.5rem !important;
    font-size: 0.875rem !important;
    box-shadow: none !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus {
    border-bottom-color: #D4C3B0 !important;
    box-shadow: none !important;
    outline: none !important;
}
div[data-testid="stTextInputRootElement"],
div[data-testid="stNumberInputRootElement"] {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}
div[data-testid="stTextInputRootElement"] > div,
div[data-testid="stNumberInputRootElement"] > div {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}

div[data-testid="stDataFrame"] { border: 1px solid #E8DDD3 !important; border-radius: 16px !important; overflow: hidden !important; background: #FFFFFF !important; }
div[data-testid="stDataFrame"] > div { background: #FFFFFF !important; }
div[data-testid="stDataFrame"] [data-testid="stDataFrameResizable"] { background: #FFFFFF !important; }
iframe { background: #FFFFFF !important; }
div[data-testid="stDataFrame"] th { background: #FEFAF7 !important; color: #9B8C7E !important; font-size: 0.6875rem !important; font-weight: 700 !important; text-transform: uppercase !important; letter-spacing: 0.12em !important; }
div[data-testid="stDataFrame"] td { color: #0A0A0A !important; font-size: 0.875rem !important; border-bottom: 1px solid #F5EEE7 !important; }
div[data-testid="stDataFrame"] tr:hover td { background: #F5EEE7 !important; }

details { background: #FFFFFF !important; border: 1px solid #E8DDD3 !important; border-radius: 12px !important; }
details summary { color: #7A6F65 !important; font-size: 0.875rem !important; font-weight: 500 !important; padding: 0.75rem 1rem !important; }

hr { border: none !important; border-top: 1px solid #E8DDD3 !important; margin: 2rem 0 !important; }

::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #F5EEE7; }
::-webkit-scrollbar-thumb { background: #B7A08B; border-radius: 3px; }

#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
header { visibility: hidden; }

/* ===== CLEAN SIDEBAR ===== */

section[data-testid="stSidebar"] {
    background: #FEFAF7 !important;
    border-right: 1px solid #E8DDD3 !important;
    min-width: 260px !important;
    max-width: 260px !important;
}

section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    padding-top: 1rem !important;
    padding-bottom: 1rem !important;
    overflow-y: auto !important;
    overflow-x: hidden !important;
}

section[data-testid="stSidebar"] .stVerticalBlock,
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"],
section[data-testid="stSidebar"] .element-container {
    gap: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
}

/* Hide collapse button */
button[data-testid="stSidebarNavCollapseButton"],
button[data-testid="stBaseButton-headerNoPadding"],
section[data-testid="stSidebar"] button[kind="header"],
[data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"] {
    display: none !important;
}

/* Hide radio widget label */
section[data-testid="stSidebar"] .stRadio > label,
section[data-testid="stSidebar"] .stRadio [data-testid="stWidgetLabel"] {
    display: none !important;
}

/* Radio group */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] {
    display: flex !important;
    flex-direction: column !important;
    gap: 0.12rem !important;
    padding: 0 0.35rem !important;
}

/* Hide radio circles */
section[data-testid="stSidebar"] .stRadio input[type="radio"],
section[data-testid="stSidebar"] .stRadio div[data-baseweb="radio"],
section[data-testid="stSidebar"] .stRadio svg {
    display: none !important;
}

/* Nav item */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
    display: flex !important;
    align-items: center !important;
    width: 100% !important;
    min-height: 2rem !important;
    padding: 0.28rem 0.65rem !important;
    margin: 0 !important;
    border-radius: 8px !important;
    border-left: 2px solid transparent !important;
    background: transparent !important;
    color: #0A0A0A !important;
    font-size: 0.86rem !important;
    font-weight: 400 !important;
    line-height: 1.25 !important;
    cursor: pointer !important;
    position: relative !important;
    transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease !important;
}

/* Selected nav item */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] label:has(input[type="radio"]:checked) {
    background: rgba(21, 77, 87, 0.09) !important;
    color: #154D57 !important;
    font-weight: 600 !important;
    border-left-color: #154D57 !important;
}

/* Hover */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] label:hover {
    background: rgba(21, 77, 87, 0.055) !important;
    color: #154D57 !important;
}

/* Text inside nav items */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] label p,
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] label span,
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] label div {
    color: inherit !important;
    font-size: inherit !important;
    font-weight: inherit !important;
    line-height: inherit !important;
    margin: 0 !important;
}

/* Section spacing */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(3),
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(7),
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(9),
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(12) {
    margin-top: 0.65rem !important;
}

/* Section labels via ::before */
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(3)::before { content: "ANALYSIS"; }
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(7)::before { content: "RANKINGS"; }
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(9)::before  { content: "TOOLS"; }
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(12)::before { content: "DATA"; }

section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(3)::before,
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(7)::before,
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(9)::before,
section[data-testid="stSidebar"] .stRadio [role="radiogroup"] > label:nth-child(12)::before {
    position: absolute !important;
    transform: translateY(-1.05rem) !important;
    font-size: 0.58rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.13em !important;
    color: #B7A08B !important;
    pointer-events: none !important;
}

/* ===== FINAL SIDEBAR SPACING FIX ===== */

section[data-testid="stSidebar"] > div,
section[data-testid="stSidebar"] [data-testid="stSidebarContent"],
section[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] {
    padding-top: 0.25rem !important;
    margin-top: 0 !important;
}

.sidebar-title {
    padding: 0.1rem 0.25rem 0.55rem 0.25rem !important;
    margin: 0 0 0.75rem 0 !important;
    border-bottom: 1px solid #E8DDD3 !important;
    font-size: 1.05rem !important;
    font-weight: 800 !important;
    line-height: 1.1 !important;
    letter-spacing: -0.03em !important;
    white-space: nowrap !important;
    color: #0A0A0A !important;
}

section[data-testid="stSidebar"] .stRadio {
    margin-top: 0 !important;
    padding-top: 0 !important;
}

section[data-testid="stSidebar"] .stRadio [role="radiogroup"] {
    padding-top: 0.15rem !important;
}

/* Force-hide radio circles */
section[data-testid="stSidebar"] .stRadio input[type="radio"],
section[data-testid="stSidebar"] .stRadio div[data-baseweb="radio"],
section[data-testid="stSidebar"] .stRadio label > div:first-child,
section[data-testid="stSidebar"] .stRadio svg,
section[data-testid="stSidebar"] .stRadio circle {
    display: none !important;
    opacity: 0 !important;
    width: 0 !important;
    height: 0 !important;
}

.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    background: #F5EEE7;
    border-radius: 10px;
    padding: 4px;
    border: 1px solid #E8DDD3;
    width: fit-content;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    padding: 0.4rem 1.2rem;
    font-size: 0.875rem;
    font-weight: 500;
    color: #9B8C7E;
    background: transparent;
    border: none;
}
.stTabs [aria-selected="true"] {
    background: #154D57 !important;
    color: #FFFFFF !important;
    font-weight: 600 !important;
    box-shadow: 0 1px 4px rgba(21,77,87,0.2) !important;
}
.stTabs [aria-selected="true"] p,
.stTabs [aria-selected="true"] span,
.stTabs [aria-selected="true"] div,
.stTabs [aria-selected="true"] * {
    color: #FFFFFF !important;
}

.card-panel {
    background: #FFFFFF;
    border-radius: 18px;
    border: 1px solid #E8DDD3;
    padding: 1.1rem;
    transition: transform 0.18s ease, box-shadow 0.18s ease;
}
.card-panel:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(21,77,87,0.10);
}

:focus-visible {
    outline: 2px dashed #154D57 !important;
    outline-offset: 2px !important;
}

.stButton > button:disabled {
    opacity: 0.5 !important;
    cursor: not-allowed !important;
}

.infobox {
    background: #EAF2F3;
    border: 1px solid #A3C9CE;
    border-radius: 12px;
    padding: 1rem;
    color: #154D57;
}

.card-container {
    border-radius: 18px;
    border: 1px solid #E8DDD3;
    background: #FFFFFF;
    padding: 1rem;
    margin-bottom: 1.2rem;
}

div[data-testid="stSlider"] > div > div > div > div[role="slider"] {
    background: #154D57 !important;
    border-color: #154D57 !important;
}
div[data-testid="stSlider"] > div > div > div > div:first-child {
    background: linear-gradient(90deg, #154D57, #1A6B77) !important;
}
.stSlider > div > div > div > div {
    color: #154D57 !important;
}

div[data-testid="stProgress"] > div {
    background: #E8DDD3 !important;
    border-radius: 4px !important;
    height: 4px !important;
}
div[data-testid="stProgress"] > div > div {
    background: #154D57 !important;
    border-radius: 4px !important;
}

/* ===== MOTION SYSTEM ===== */

/* Keyframes */
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(14px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes fadeIn {
    from { opacity: 0; }
    to   { opacity: 1; }
}
@keyframes shimmer {
    0%   { background-position: -400px 0; }
    100% { background-position:  400px 0; }
}
@keyframes indeterminate {
    0%   { transform: translateX(-120%) scaleX(0.4); }
    60%  { transform: translateX(50%)   scaleX(0.6); }
    100% { transform: translateX(220%)  scaleX(0.4); }
}

/* Skeleton shimmer card */
.skeleton-card {
    background: linear-gradient(90deg, #EEEBE6 25%, #FAFAFA 50%, #EEEBE6 75%);
    background-size: 400px 100%;
    animation: shimmer 1.5s linear infinite;
    border-radius: 6px;
    display: block;
}

/* Indeterminate progress bar */
.ind-track {
    height: 3px;
    background: #E8DDD3;
    border-radius: 2px;
    overflow: hidden;
    margin: 0.5rem 0 1.25rem 0;
}
.ind-bar {
    height: 100%;
    width: 35%;
    background: linear-gradient(90deg, #154D57, #1A6B77);
    border-radius: 2px;
    animation: indeterminate 1.3s cubic-bezier(0.4, 0, 0.6, 1) infinite;
    transform-origin: left center;
}

/* Results reveal */
.results-reveal {
    animation: fadeInUp 0.45s cubic-bezier(0.16, 1, 0.3, 1) both;
}

/* Tab panel fade on switch */
div[role="tabpanel"] {
    animation: fadeIn 0.18s ease-out;
}

/* Chart containers  -  card styling + subtle fade */
div[data-testid="stPlotlyChart"] {
    animation: fadeIn 0.25s ease-out;
    background: #FFFFFF !important;
    border: 1px solid #E8DDD3 !important;
    border-radius: 16px !important;
    padding: 1rem !important;
    margin-bottom: 1rem !important;
}

/* Collapse zero-height iframes from components.html(height=0) */
div[data-testid="stCustomComponentV1"][style*="height: 0"],
div[data-testid="stCustomComponentV1"] iframe[height="0"] {
    height: 0 !important;
    min-height: 0 !important;
    overflow: hidden !important;
    margin: 0 !important;
    padding: 0 !important;
    display: block !important;
}
/* Collapse the parent element-container that wraps the zero-height component */
.element-container:has(> div[data-testid="stCustomComponentV1"]) {
    margin: 0 !important;
    padding: 0 !important;
    min-height: 0 !important;
    line-height: 0 !important;
}

/* Remove excess top and bottom padding from main content area */
section[data-testid="stMain"] > div:first-child {
    padding-top: 0 !important;
    padding-bottom: 0 !important;
}
.block-container {
    padding-bottom: 0 !important;
}

/* Button press tactile feedback */
.stButton > button:active {
    transform: scale(0.97) !important;
    box-shadow: none !important;
    transition-duration: 0.08s !important;
}

/* Input focus underline grow via pseudo-element */
div[data-testid="stTextInputRootElement"] > div,
div[data-testid="stNumberInputRootElement"] > div {
    position: relative !important;
}
div[data-testid="stTextInputRootElement"] > div::after,
div[data-testid="stNumberInputRootElement"] > div::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: #154D57;
    transform: scaleX(0);
    transform-origin: left center;
    transition: transform 0.22s cubic-bezier(0.4, 0, 0.2, 1);
    z-index: 2;
    pointer-events: none;
    border-radius: 1px;
}
div[data-testid="stTextInputRootElement"] > div:focus-within::after,
div[data-testid="stNumberInputRootElement"] > div:focus-within::after {
    transform: scaleX(1);
}

/* Feature card hover  -  warm tint (works inside overflow:hidden grid) */
.feature-card {
    transition: background 0.15s ease-out !important;
    cursor: default;
}
.feature-card:hover {
    background: #F3EDE5 !important;
}

/* Scroll-reveal sections */
.reveal-section {
    transition: opacity 0.5s cubic-bezier(0.16, 1, 0.3, 1),
                transform 0.5s cubic-bezier(0.16, 1, 0.3, 1);
}
.reveal-section.pre-reveal {
    opacity: 0;
    transform: translateY(18px);
}
.reveal-section.revealed {
    opacity: 1 !important;
    transform: translateY(0) !important;
}

/* Sidebar nav label subtle slide-in on hover */
section[data-testid="stSidebar"] .stRadio label:hover {
    padding-left: 1.25rem !important;
}

/* Accessibility: honour reduced-motion preference */
@media (prefers-reduced-motion: reduce) {
    *, *::before, *::after {
        animation-duration: 0.01ms !important;
        animation-iteration-count: 1 !important;
        transition-duration: 0.01ms !important;
    }
    .skeleton-card { animation: none !important; background: #F0EDE8 !important; }
    .ind-bar       { animation: none !important; width: 65% !important; }
}

/* Preset buttons  -  outlined style */
div[data-testid="stHorizontalBlock"] .stButton > button {
    background: #EAF2F3 !important;
    color: #154D57 !important;
    border: 1.5px solid #154D57 !important;
    border-radius: 980px !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
    padding: 0.45rem 0.75rem !important;
    transition: background 0.15s ease, color 0.15s ease !important;
}
div[data-testid="stHorizontalBlock"] .stButton > button:hover {
    background: #154D57 !important;
    color: #FFFFFF !important;
}
div[data-testid="stHorizontalBlock"] .stButton > button p,
div[data-testid="stHorizontalBlock"] .stButton > button span,
div[data-testid="stHorizontalBlock"] .stButton > button div {
    color: #154D57 !important;
}
div[data-testid="stHorizontalBlock"] .stButton > button:hover p,
div[data-testid="stHorizontalBlock"] .stButton > button:hover span,
div[data-testid="stHorizontalBlock"] .stButton > button:hover div {
    color: #FFFFFF !important;
}
.export-dl-btn {
    display: block;
    background-color: #154D57;
    color: #FFFFFF !important;
    text-align: center;
    padding: 0.55rem 1rem;
    border-radius: 8px;
    font-weight: 600;
    font-size: 0.875rem;
    text-decoration: none !important;
    margin-top: 0.75rem;
    transition: background-color 0.15s ease;
}
.export-dl-btn:hover {
    background-color: #1A6070;
    color: #FFFFFF !important;
    text-decoration: none !important;
}
</style>
""", unsafe_allow_html=True)

# Colours
COLOURS = {
    "BlackRock EM": "#60A5FA",
    "Fidelity ESG EM": "#F87171",
    "JPMorgan EM": "#4ADE80",
    "Schroder EM": "#FBBF24",
    "Benchmark": "#94A3B8",
}
# Fallback palette for custom / extra funds
_EXTRA_PALETTE = ["#A78BFA", "#34D399", "#FB923C", "#E879F9", "#38BDF8", "#FCD34D"]

def _get_colour(name, idx=0):
    """Return a colour for any fund  -  falls back to the extra palette for unknown names."""
    if name in COLOURS:
        return COLOURS[name]
    return _EXTRA_PALETTE[idx % len(_EXTRA_PALETTE)]

# Plotly layout
PLOTLY_LAYOUT = dict(
    template="plotly_white",
    paper_bgcolor="#FFFFFF",
    plot_bgcolor="#F5F5F7",
    font=dict(family="Inter, -apple-system, sans-serif", color="#000000", size=12),
    title=dict(font=dict(size=16, color="#000000", family="Inter"), pad=dict(b=20)),
    xaxis=dict(gridcolor="#E8DDD3", linecolor="#E8DDD3", tickfont=dict(color="#7A6F65")),
    yaxis=dict(gridcolor="#E8DDD3", linecolor="#E8DDD3", tickfont=dict(color="#7A6F65")),
    legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#7A6F65", size=11)),
    margin=dict(l=20, r=20, t=60, b=20),
    hoverlabel=dict(bgcolor="#FFFFFF", bordercolor="#E8DDD3", font=dict(color="#000000")),
    transition=dict(duration=400, easing="cubic-in-out"),
)

# Initialize session state
defaults = {
    "analysis_run": False,
    "fund_names": ["BlackRock EM", "Fidelity ESG EM", "JPMorgan EM", "Schroder EM"],
    "tickers": ["IEMG", "EMXF", "JPEM", "GEM"],
    "benchmark_ticker": "EEM",
    "benchmark_name": "Benchmark",
    "start_date": "2020-01-01",
    "end_date": "2025-10-31",
    "rf_annual": 0.05,
    "log_returns": None,
    "fund_returns": None,
    "benchmark_returns": None,
    "core_metrics_df": None,
    "downside_df": None,
    "metrics_matrix": None,
    "topsis_obj": None,
    "topsis_ranking": None,
    "yuan_obj": None,
    "yuan_ranking": None,
    "naive_ranking": None,
    "borda_ranking": None,
    "rolling_rankings": None,
    "period_a_ranking": None,
    "period_b_ranking": None,
    "period_a_label": "Period A",
    "period_b_label": "Period B",
    "stress_ranking": None,
    "active_stress_test": None,
    "stress_description": "",
    "_apply_alloc": False,
    "_active_preset": None,
    "mc_results": None,
    "mc_best": None,
    "mc_win_counts": None,
    "pillar_weights": None,
    "saved_scenarios": {},
    "bootstrap_results": None,
    "extra_funds": [],
    "costs": {"BlackRock EM": 0.99, "Fidelity ESG EM": 1.07, "JPMorgan EM": 0.90, "Schroder EM": 0.98},
    "esg_globe": {"BlackRock EM": 4, "Fidelity ESG EM": 5, "JPMorgan EM": 3, "Schroder EM": 4},
    "carbon_risk": {"BlackRock EM": 7.85, "Fidelity ESG EM": 5.72, "JPMorgan EM": 8.72, "Schroder EM": 7.62},
}

for key, value in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value

# Sidebar
st.sidebar.markdown("""
<div class="sidebar-title">Fund Analysis Engine</div>
""", unsafe_allow_html=True)
pages = ["Home", "Setup", "Performance", "TOPSIS", "Yuan & Yuan", "📋 Executive Summary", "📋 Model Comparison", "📅 Rolling Rankings", "🎨 Visualisations", "📦 Portfolio", "Sensitivity & Report", "🗃️ Raw Data", "⚖️ Comparison"]
if "_nav_page_idx" not in st.session_state:
    st.session_state._nav_page_idx = 0
if "_force_nav" not in st.session_state:
    st.session_state._force_nav = False
if st.session_state._force_nav:
    # Set nav_radio BEFORE the widget renders  -  allowed per Streamlit 1.38+ pre-instantiation rules
    st.session_state["nav_radio"] = pages[st.session_state._nav_page_idx]
    st.session_state._force_nav = False
page = st.sidebar.radio("Navigation", pages, key="nav_radio")
st.session_state._nav_page_idx = pages.index(page)
st.session_state.current_page = page
st.sidebar.markdown("<div style='height: 0.4rem'></div>", unsafe_allow_html=True)
if st.session_state.analysis_run:
    st.sidebar.markdown("""
    <div style="border-top: 1px solid #E8DDD3; padding-top: 0.6rem; margin-top: 0.25rem;">
        <div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.35rem;">
            <div style="width:6px; height:6px; border-radius:50%; background:#154D57; flex-shrink:0;"></div>
            <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#154D57;">Analysis Ready</div>
        </div>
        <div style="font-size:0.75rem; color:#9B8C7E; line-height:1.5; padding-left:1rem;">All modules unlocked</div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.sidebar.markdown("""
    <div style="border-top: 1px solid #E8DDD3; padding-top: 0.6rem; margin-top: 0.25rem;">
        <div class="sidebar-alert" style="display:flex; align-items:center; gap:0.45rem; margin-bottom:0.3rem;">
            <span style="font-size:0.85rem; line-height:1;">&#9733;</span>
            <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase;">Setup Required</div>
        </div>
        <div style="font-size:0.75rem; color:#9B8C7E; line-height:1.5; padding-left:1.3rem;">Run analysis to begin</div>
    </div>
    """, unsafe_allow_html=True)






# Page routing


if page == "Home":
    st.markdown("""
<div style="max-width:860px;margin:1rem auto 0;padding:0 2rem;">
    <div style="display:flex;justify-content:flex-end;margin-bottom:0.9rem;">
        <div style="display:inline-flex;align-items:center;gap:.5rem;
                    background:rgba(21,77,87,.07);border:1px solid rgba(21,77,87,.18);
                    border-radius:980px;padding:.25rem 1rem;">
            <span style="width:6px;height:6px;border-radius:50%;background:#154D57;display:inline-block;"></span>
            <span style="font-size:.7rem;font-weight:600;color:#154D57;letter-spacing:.1em;text-transform:uppercase;">
                MSc Finance &middot; University of Nottingham &middot; BUSI4519
            </span>
        </div>
    </div>
    <h1 style="font-size:clamp(2.4rem,4.5vw,3.8rem);font-weight:800;color:#0A0A0A;
               letter-spacing:-.04em;line-height:1.08;margin:0 0 0.75rem;">
        Quantitative Fund Analysis<br>
        <span style="background:linear-gradient(135deg,#4ECDC4,#154D57);
                     -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                     background-clip:text;">Built from First Principles.</span>
    </h1>
    <p style="font-size:1.15rem;color:#5A5A5A;line-height:1.65;margin:0 0 1.4rem;max-width:680px;">
        A rigorous multi-criteria evaluation framework comparing emerging market equity funds
        across <strong>19 metrics</strong> using TOPSIS and Yuan &amp; Yuan (2023) eigenvector
        ranking  -  built in Python and deployed as an interactive tool.
    </p>
</div>
""", unsafe_allow_html=True)

    _, col_btn, _ = st.columns([2, 3, 2])
    with col_btn:
        if st.button("Launch Analysis Engine →", type="primary", use_container_width=True):
            st.session_state._nav_page_idx = pages.index("Setup")
            st.session_state._force_nav = True
            st.rerun()

    st.markdown("<div style='max-width:860px;margin:1.5rem auto 0;padding:0 2rem;'>", unsafe_allow_html=True)

    st.markdown("""
<div style="border-top:1px solid #E8DDD3;padding-top:1.25rem;margin-top:0;">
    <div style="font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.12em;
                color:#154D57;margin-bottom:1rem;">What this tool does</div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:.85rem;">
        <div style="background:#FAF7F3;border:1px solid #E8DDD3;border-radius:14px;padding:1.2rem 1.3rem;">
            <div style="font-size:.95rem;font-weight:700;color:#0A0A0A;margin-bottom:.35rem;">📊 19-Metric Analysis</div>
            <div style="font-size:.875rem;color:#5A5A5A;line-height:1.55;">Returns, risk-adjusted ratios, drawdown, cost &amp; ESG  -  all from monthly log returns.</div>
        </div>
        <div style="background:#FAF7F3;border:1px solid #E8DDD3;border-radius:14px;padding:1.2rem 1.3rem;">
            <div style="font-size:.95rem;font-weight:700;color:#0A0A0A;margin-bottom:.35rem;">🎯 TOPSIS Ranking</div>
            <div style="font-size:.875rem;color:#5A5A5A;line-height:1.55;">Full 5-step Hwang &amp; Yoon (1981) algorithm with configurable pillar weights.</div>
        </div>
        <div style="background:#FAF7F3;border:1px solid #E8DDD3;border-radius:14px;padding:1.2rem 1.3rem;">
            <div style="font-size:.95rem;font-weight:700;color:#0A0A0A;margin-bottom:.35rem;">⚡ Yuan &amp; Yuan Ranking</div>
            <div style="font-size:.875rem;color:#5A5A5A;line-height:1.55;">Eigenvector-based pairwise competition matrix converging via power iteration.</div>
        </div>
        <div style="background:#FAF7F3;border:1px solid #E8DDD3;border-radius:14px;padding:1.2rem 1.3rem;">
            <div style="font-size:.95rem;font-weight:700;color:#0A0A0A;margin-bottom:.35rem;">📈 Portfolio &amp; Monte Carlo</div>
            <div style="font-size:.875rem;color:#5A5A5A;line-height:1.55;">Efficient frontier, min-variance &amp; max-Sharpe portfolios with 10,000-path simulation.</div>
        </div>
        <div style="background:#FAF7F3;border:1px solid #E8DDD3;border-radius:14px;padding:1.2rem 1.3rem;">
            <div style="font-size:.95rem;font-weight:700;color:#0A0A0A;margin-bottom:.35rem;">📅 Rolling Rankings</div>
            <div style="font-size:.875rem;color:#5A5A5A;line-height:1.55;">Track fund rank stability over 12–36 month rolling windows.</div>
        </div>
        <div style="background:#FAF7F3;border:1px solid #E8DDD3;border-radius:14px;padding:1.2rem 1.3rem;">
            <div style="font-size:.95rem;font-weight:700;color:#0A0A0A;margin-bottom:.35rem;">📄 PDF &amp; Excel Export</div>
            <div style="font-size:.875rem;color:#5A5A5A;line-height:1.55;">Formatted multi-page reports with per-cell highlights and full methodology notes.</div>
        </div>
    </div>
</div>

<div style="margin-top:1.5rem;padding:1.2rem 1.4rem;
            background:rgba(21,77,87,.04);border-left:3px solid #154D57;border-radius:0 10px 10px 0;">
    <div style="font-size:.78rem;font-weight:700;color:#154D57;text-transform:uppercase;
                letter-spacing:.08em;margin-bottom:.45rem;">How to get started</div>
    <ol style="font-size:.9rem;color:#5A5A5A;line-height:1.8;margin:0;padding-left:1.2rem;">
        <li>Go to <strong>Setup</strong>  -  enter fund tickers, benchmark, date range and risk-free rate.</li>
        <li>Click <strong>Run Analysis</strong>  -  metrics are computed from live Yahoo Finance data.</li>
        <li>Explore each page: Performance, TOPSIS, Yuan &amp; Yuan, Visualisations, Portfolio.</li>
        <li>Download your <strong>PDF or Excel report</strong> from the Sensitivity &amp; Report page.</li>
    </ol>
</div>

<div style="margin-top:1.5rem;font-size:.78rem;color:#9B9B9B;border-top:1px solid #E8DDD3;padding-top:1rem;">
    Default funds: EIMI, VFEM, HMEF, LGEM &nbsp;&middot;&nbsp; Benchmark: SPY &nbsp;&middot;&nbsp;
    Hwang &amp; Yoon (1981) &nbsp;&middot;&nbsp; Yuan &amp; Yuan (2023)
</div>
</div>
""", unsafe_allow_html=True)


elif page == "📋 Executive Summary":
    if not st.session_state.get("analysis_run"):
        st.markdown("""
        <div style="text-align:center;padding:4rem 2rem;background:#F5F3F0;
                    border-radius:20px;border:2px dashed #D4D2CF;">
            <div style="font-size:3rem;margin-bottom:1rem;">📋</div>
            <div style="font-size:1.1rem;font-weight:600;color:#1A1A1A;margin-bottom:0.5rem;">
                No analysis yet
            </div>
            <div style="color:#9B9B9B;">Run the analysis first from the Setup page.</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        fund_names  = st.session_state.fund_names
        core_df     = st.session_state.core_metrics_df
        downside_df = st.session_state.downside_df

        # Auto-compute rankings if the user navigated here without visiting TOPSIS/Yuan pages first
        if st.session_state.topsis_ranking is None or st.session_state.yuan_ranking is None:
            _mm = st.session_state.metrics_matrix
            if _mm is not None:
                try:
                    _pw = st.session_state.pillar_weights or {
                        'Returns': 40, 'Risk-Adj': 25, 'Risk/DD': 20, 'Costs': 10, 'ESG': 5
                    }
                    _benefits = [
                        'Ann. Return (%)', 'Alpha (ann. %)', 'Sharpe Ratio', 'Sortino Ratio',
                        'Treynor Ratio', 'Information Ratio', 'R²', 'Upside Capture (%)',
                        'Calmar Ratio', 'ESG Globe Rating', 'Max Drawdown (%)',
                    ]
                    _costs = [
                        'Ann. Volatility (%)', 'Beta', 'Tracking Error (%)',
                        'Max DD Duration (mths)', 'Downside Capture (%)', 'OCF', 'Carbon Risk Score',
                    ]
                    if st.session_state.topsis_ranking is None:
                        _tr = _inline_topsis(_mm, fund_names, _pw)
                        if not _tr.empty:
                            # Rename and rescale to match format stored by the TOPSIS page (Score 0-1, Rank)
                            _tr = _tr.rename(columns={'TOPSIS Score (%)': 'Score'})
                            _tr['Score'] = (_tr['Score'] / 100.0).round(6)
                            st.session_state.topsis_ranking = _tr
                            st.session_state.topsis_obj = TOPSIS(_mm, _benefits, _costs)
                    if st.session_state.yuan_ranking is None:
                        _yuan_obj = YuanYuan(_mm, _benefits, _costs)
                        _yuan_ranking, _c_matrix, _n = _yuan_obj.run(_pw)
                        st.session_state.yuan_ranking = _yuan_ranking
                        st.session_state.yuan_obj = _yuan_obj
                except Exception as _e:
                    st.warning(
                        f"Model rankings could not be auto-computed: {_e}. "
                        "Visit the TOPSIS and Yuan & Yuan pages to set weights and compute rankings."
                    )

        topsis_r    = st.session_state.topsis_ranking
        yuan_r      = st.session_state.yuan_ranking

        # Hero banner
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#154D57 0%,#1A6B77 100%);
                    border-radius:24px;padding:2.5rem;margin-bottom:2rem;">
            <div style="font-size:0.72rem;font-weight:700;text-transform:uppercase;
                        letter-spacing:0.12em;color:rgba(255,255,255,0.6);margin-bottom:0.75rem;">
                FUND ANALYSIS ENGINE · EXECUTIVE SUMMARY
            </div>
            <div style="font-size:2rem;font-weight:700;color:#FFFFFF;
                        letter-spacing:-0.03em;margin-bottom:0.5rem;">
                {len(fund_names)} Funds Analysed · {len(fund_names)} Rankings Computed
            </div>
            <div style="font-size:0.95rem;color:rgba(255,255,255,0.75);">
                {st.session_state.get('start_date', '2020-01-01')} to {st.session_state.get('end_date', '2025-10-31')}
                · Risk-free rate {st.session_state.get('rf_rate', 5.0):.1f}%
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Key finding cards
        st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">KEY FINDINGS</p>', unsafe_allow_html=True)

        try:
            fund_only_topsis   = topsis_r.loc[topsis_r.index.isin(set(fund_names))]
            top_topsis         = fund_only_topsis.sort_values("Rank").index[0]
            top_topsis_score   = float(fund_only_topsis.loc[top_topsis, "Score"]) if "Score" in fund_only_topsis.columns else 0.0
        except Exception as e:
            st.error(f"Executive summary TOPSIS ranking error: {e}")
            top_topsis       = "N/A"
            top_topsis_score = 0.0

        try:
            fund_only_yuan = yuan_r.loc[yuan_r.index.isin(set(fund_names))]
            top_yuan       = fund_only_yuan.sort_values("Rank").index[0]
        except Exception as e:
            st.error(f"Executive summary Yuan ranking error: {e}")
            top_yuan = "N/A"

        try:
            comparison_funds = [f for f in fund_names if f in core_df.index]

            best_sharpe_fund = core_df.loc[comparison_funds, "Sharpe Ratio"].astype(float).idxmax()
            best_sharpe_val  = float(core_df.loc[best_sharpe_fund, "Sharpe Ratio"])

            best_alpha_fund  = core_df.loc[comparison_funds, "Alpha (ann. %)"].astype(float).idxmax()
            best_alpha_val   = float(core_df.loc[best_alpha_fund, "Alpha (ann. %)"])

            downside_funds   = [f for f in fund_names if f in downside_df.index]
            best_dd_fund     = downside_df.loc[downside_funds, "Max Drawdown (%)"].astype(float).idxmax()
            best_dd_val      = float(downside_df.loc[best_dd_fund, "Max Drawdown (%)"])

        except Exception as e:
            st.error(f"Executive summary metric card error: {e}")
            best_sharpe_fund = best_dd_fund = best_alpha_fund = "N/A"
            best_sharpe_val = best_dd_val = best_alpha_val = np.nan

        finding_cards = [
            ("#154D57", "🏆 Top Ranked (TOPSIS)",        top_topsis,       f"Closeness coefficient {top_topsis_score:.3f}"),
            ("#4ECDC4", "⚡ Top Ranked (Yuan & Yuan)",    top_yuan,         "Highest eigenvector score"),
            ("#C8A96E", "📈 Best Risk-Adjusted Return",   best_sharpe_fund, f"Sharpe ratio {best_sharpe_val:.3f}"),
            ("#8B5CF6", "🛡️ Best Capital Preservation",  best_dd_fund,     f"Max drawdown {best_dd_val:.1f}%"),
            ("#059669", "⚡ Best Alpha Generator",        best_alpha_fund,  f"Jensen's alpha {best_alpha_val:.2f}% p.a."),
        ]

        cols = st.columns(len(finding_cards))
        for col, (accent, label, fund, sub) in zip(cols, finding_cards):
            with col:
                st.markdown(f"""
                <div style="background:#FFFFFF;border-radius:16px;padding:1.1rem 1.25rem;
                            box-shadow:0 2px 10px rgba(0,0,0,0.06);
                            border-top:3px solid {accent};">
                    <div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;
                                letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.4rem;">{label}</div>
                    <div style="font-size:1rem;font-weight:700;color:#1A1A1A;
                                margin-bottom:0.3rem;">{fund}</div>
                    <div style="font-size:0.72rem;color:{accent};font-weight:500;">{sub}</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)

        # All funds at a glance
        st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">ALL FUNDS AT A GLANCE</p>', unsafe_allow_html=True)

        try:
            summary_df = pd.DataFrame({
                fund: {
                    "Ann. Return (%)":      round(float(core_df.loc[fund, "Ann. Return (%)"]), 2),
                    "Sharpe Ratio":         round(float(core_df.loc[fund, "Sharpe Ratio"]), 3),
                    "Max Drawdown (%)":     round(float(downside_df.loc[fund, "Max Drawdown (%)"]), 2),
                    "Downside Capture (%)": round(float(downside_df.loc[fund, "Downside Capture (%)"]), 2),
                    "Alpha (ann. %)":       round(float(core_df.loc[fund, "Alpha (ann. %)"]), 3),
                }
                for fund in fund_names
            }).T
            st.dataframe(summary_df, use_container_width=True)
        except Exception as _e:
            st.warning(f"Could not render summary table: {_e}")

        st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)

        # Consensus ranking
        st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">CONSENSUS RANKING</p>', unsafe_allow_html=True)

        try:
            fund_topsis = topsis_r.loc[topsis_r.index.isin(set(fund_names))].sort_values("Rank")
            fund_yuan   = yuan_r.loc[yuan_r.index.isin(set(fund_names))].sort_values("Rank")
            medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣"]

            rank_cols = st.columns(2)
            with rank_cols[0]:
                st.markdown('<p style="font-size:0.78rem;font-weight:600;color:#154D57;margin-bottom:0.5rem;">TOPSIS</p>', unsafe_allow_html=True)
                for i, (fund, row) in enumerate(fund_topsis.iterrows()):
                    score_col = "Score" if "Score" in row.index else row.index[0]
                    score     = float(row.get(score_col, 0))
                    f_colour  = COLOURS.get(fund, "#154D57")
                    bar_w     = int(score * 100) if score <= 1 else int(score)
                    st.markdown(f"""
                    <div style="background:#FFFFFF;border-radius:12px;padding:0.85rem 1rem;
                                margin-bottom:0.4rem;border-left:4px solid {f_colour};
                                box-shadow:0 1px 4px rgba(0,0,0,0.04);">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.3rem;">
                            <span style="font-weight:600;font-size:0.88rem;">{medals[i]} {fund}</span>
                            <span style="font-weight:700;font-size:0.88rem;">{score:.3f}</span>
                        </div>
                        <div style="background:#F5F3F0;border-radius:3px;height:3px;">
                            <div style="background:{f_colour};height:3px;border-radius:3px;width:{bar_w}%;"></div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

            with rank_cols[1]:
                st.markdown('<p style="font-size:0.78rem;font-weight:600;color:#4ECDC4;margin-bottom:0.5rem;">YUAN & YUAN</p>', unsafe_allow_html=True)
                for i, (fund, row) in enumerate(fund_yuan.iterrows()):
                    score_col = "Score" if "Score" in row.index else row.index[0]
                    score     = float(row.get(score_col, 0))
                    f_colour  = COLOURS.get(fund, "#4ECDC4")
                    bar_w     = int(score * 100) if score <= 1 else int(score)
                    st.markdown(f"""
                    <div style="background:#FFFFFF;border-radius:12px;padding:0.85rem 1rem;
                                margin-bottom:0.4rem;border-left:4px solid {f_colour};
                                box-shadow:0 1px 4px rgba(0,0,0,0.04);">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.3rem;">
                            <span style="font-weight:600;font-size:0.88rem;">{medals[i]} {fund}</span>
                            <span style="font-weight:700;font-size:0.88rem;">{score:.4f}</span>
                        </div>
                        <div style="background:#F5F3F0;border-radius:3px;height:3px;">
                            <div style="background:{f_colour};height:3px;border-radius:3px;width:{bar_w}%;"></div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        except Exception as _e:
            st.warning(f"Could not render consensus ranking bars: {_e}")

elif page == "Setup":
    st.markdown("# Fund Setup")
    st.markdown('<p style="font-size:0.875rem; color:#9B8C7E; margin-top:-0.25rem; margin-bottom:2rem; max-width:640px; line-height:1.5;">Define your fund universe, benchmark, and analysis parameters. Data is fetched live from Yahoo Finance using monthly log returns.</p>', unsafe_allow_html=True)

    # ── FUND UNIVERSE ──────────────────────────────────────────────────────
    st.markdown("""
    <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:1.5rem;">
        <div style="width:3px; height:18px; background:#154D57; border-radius:2px; flex-shrink:0;"></div>
        <span style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E;">Fund Universe</span>
    </div>
    """, unsafe_allow_html=True)

    col_lbl, col_sld = st.columns([1.4, 4])
    with col_lbl:
        st.markdown('<div style="padding-top:0.5rem; font-size:0.875rem; color:#7A6F65; font-weight:500;">Number of funds</div>', unsafe_allow_html=True)
    with col_sld:
        num_funds = st.slider("nf", 2, 6, len(st.session_state.fund_names), label_visibility="collapsed", key="num_funds_slider")

    if num_funds != len(st.session_state.fund_names):
        if num_funds > len(st.session_state.fund_names):
            st.session_state.fund_names.extend([f"Fund {i+1}" for i in range(len(st.session_state.fund_names), num_funds)])
            st.session_state.tickers.extend([f"TICK{i+1}" for i in range(len(st.session_state.tickers), num_funds)])
        else:
            st.session_state.fund_names = st.session_state.fund_names[:num_funds]
            st.session_state.tickers = st.session_state.tickers[:num_funds]

    # Column headers
    st.markdown("""
    <div style="display:grid; grid-template-columns:48px 160px 1fr; gap:0.75rem;
                padding:0.6rem 0; border-bottom:1.5px solid #154D57; margin-bottom:0.15rem; margin-top:1rem;">
        <div></div>
        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#154D57;">Ticker</div>
        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#154D57;">Fund Name</div>
    </div>
    """, unsafe_allow_html=True)

    for i in range(num_funds):
        col_num, col_ticker, col_name = st.columns([0.38, 1.1, 2.8])
        with col_num:
            st.markdown(f'<div style="padding-top:0.5rem; font-size:1rem; font-weight:700; color:#D4C3B0; text-align:center; letter-spacing:-0.02em; font-variant-numeric:tabular-nums;">{i+1:02d}</div>', unsafe_allow_html=True)
        with col_ticker:
            st.session_state.tickers[i] = st.text_input(f"Ticker {i}", value=st.session_state.tickers[i], key=f"ticker_{i}", label_visibility="collapsed", placeholder="e.g. IEMG")
        with col_name:
            st.session_state.fund_names[i] = st.text_input(f"Name {i}", value=st.session_state.fund_names[i], key=f"name_{i}", label_visibility="collapsed", placeholder="Fund name")

    # Benchmark row
    st.markdown("""
    <div style="display:flex; align-items:center; gap:0.75rem; margin:2rem 0 0.15rem 0;">
        <div style="width:3px; height:18px; background:#B7A08B; border-radius:2px; flex-shrink:0;"></div>
        <span style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E;">Benchmark</span>
    </div>
    <div style="display:grid; grid-template-columns:48px 160px 1fr; gap:0.75rem;
                padding:0.6rem 0; border-bottom:1.5px solid #B7A08B; margin-bottom:0.15rem;">
        <div></div>
        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#B7A08B;">Ticker</div>
        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#B7A08B;">Name</div>
    </div>
    """, unsafe_allow_html=True)
    col_num, col_ticker, col_name = st.columns([0.38, 1.1, 2.8])
    with col_num:
        st.markdown('<div style="padding-top:0.5rem; font-size:0.6875rem; font-weight:700; color:#B7A08B; text-align:center; letter-spacing:0.1em;">BM</div>', unsafe_allow_html=True)
    with col_ticker:
        benchmark_ticker = st.text_input("Benchmark Ticker", value=st.session_state.benchmark_ticker, key="benchmark_ticker", label_visibility="collapsed", placeholder="e.g. EEM")
    with col_name:
        benchmark_name = st.text_input("Benchmark Name", value=st.session_state.benchmark_name, key="benchmark_name", label_visibility="collapsed", placeholder="Benchmark name")
    st.markdown('<p style="font-size:0.75rem; color:#9B8C7E; margin-top:0.5rem; padding-left:0;">Use EEM for MSCI EM, SPY for S&P 500, or any Yahoo Finance ticker.</p>', unsafe_allow_html=True)

    st.markdown('<hr style="border:none; border-top:1px solid #E8DDD3; margin:2.5rem 0;">', unsafe_allow_html=True)

    # ── CUSTOM FUNDS ──────────────────────────────────────────────────────
    st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
    st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.5rem;">ADD CUSTOM FUNDS</p>', unsafe_allow_html=True)
    st.markdown('<p style="font-size:0.82rem;color:#9B9B9B;margin-bottom:0.75rem;">Add extra funds by Yahoo Finance ticker. These will be included in all rankings and charts.</p>', unsafe_allow_html=True)

    _cf_cols = st.columns([1, 2, 1])
    with _cf_cols[0]:
        _custom_ticker = st.text_input(
            "Extra ticker",
            placeholder="e.g. VWO",
            key="custom_ticker_input",
            label_visibility="collapsed"
        )
    with _cf_cols[1]:
        _custom_name = st.text_input(
            "Extra fund name",
            placeholder="e.g. Vanguard FTSE EM",
            key="custom_name_input",
            label_visibility="collapsed"
        )
    with _cf_cols[2]:
        if st.button("➕ Add Fund", key="add_custom_fund_btn", use_container_width=True):
            if _custom_ticker.strip() and _custom_name.strip():
                _extras = st.session_state.get("extra_funds", [])
                _ticker_clean = _custom_ticker.strip().upper()
                _name_clean   = _custom_name.strip()
                if _ticker_clean not in [e["ticker"] for e in _extras]:
                    _extras.append({"ticker": _ticker_clean, "name": _name_clean})
                    st.session_state.extra_funds = _extras
                    st.success(f"Added {_name_clean} ({_ticker_clean}). Click Run Analysis to include it.")
                else:
                    st.warning(f"{_ticker_clean} is already in the list.")
            else:
                st.warning("Enter both a ticker and a fund name.")

    _extra_funds = st.session_state.get("extra_funds", [])
    if _extra_funds:
        st.markdown('<p style="font-size:0.72rem;color:#9B9B9B;margin:0.5rem 0 0.25rem 0;">Extra funds queued for next Run Analysis:</p>', unsafe_allow_html=True)
        for _ef in _extra_funds:
            _ef_c1, _ef_c2 = st.columns([5, 1])
            with _ef_c1:
                st.markdown(
                    f'<div style="background:#FFFFFF;border-radius:10px;padding:0.6rem 1rem;'
                    f'margin-bottom:0.4rem;border-left:3px solid #154D57;font-size:0.82rem;">'
                    f'<strong>{_ef["name"]}</strong> <span style="color:#9B9B9B;">({_ef["ticker"]})</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            with _ef_c2:
                if st.button("✕", key=f"remove_extra_{_ef['ticker']}", use_container_width=True):
                    st.session_state.extra_funds = [
                        e for e in _extra_funds if e["ticker"] != _ef["ticker"]
                    ]
                    st.rerun()

    st.markdown("""
    <div style="background:#F5F3F0;border-radius:10px;padding:0.75rem 1rem;margin-top:0.5rem;">
        <p style="margin:0;font-size:0.75rem;color:#9B9B9B;line-height:1.5;">
            💡 Use any Yahoo Finance ticker, for example VWO (Vanguard EM), EEMS (iShares EM Small Cap),
            or EMXF (Fidelity ESG EM). Cost and ESG data for custom funds will need to be entered manually
            in the Cost &amp; ESG section.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<hr style="border:none; border-top:1px solid #E8DDD3; margin:2.5rem 0;">', unsafe_allow_html=True)

    # ── PARAMETERS ────────────────────────────────────────────────────────
    st.markdown("""
    <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:1.5rem;">
        <div style="width:3px; height:18px; background:#154D57; border-radius:2px; flex-shrink:0;"></div>
        <span style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E;">Parameters</span>
    </div>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        start_date = st.date_input("Start Date", value=pd.to_datetime(st.session_state.start_date))
    with col2:
        end_date = st.date_input("End Date", value=pd.to_datetime(st.session_state.end_date))
    with col3:
        rf_annual = st.number_input("Risk-Free Rate (%)", value=st.session_state.rf_annual * 100, step=0.1) / 100
        st.markdown('<p style="font-size:0.75rem; color:#9B8C7E; margin-top:0.3rem;">Annual rate for Sharpe calculations.</p>', unsafe_allow_html=True)

    st.markdown('<hr style="border:none; border-top:1px solid #E8DDD3; margin:2.5rem 0;">', unsafe_allow_html=True)

    # ── COST & ESG DATA ───────────────────────────────────────────────────
    with st.expander("Cost & ESG Data: Manual Entry"):
        st.markdown('<p style="font-size:0.875rem; color:#9B8C7E; margin-bottom:1rem;">OCF, ESG Globe Rating, and Carbon Risk Score are not available from Yahoo Finance and must be entered manually.</p>', unsafe_allow_html=True)
        st.markdown("""
        <div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:0; padding:0.5rem 0; border-bottom:1px solid #E8DDD3; margin-bottom:0.25rem;">
            <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E;">Fund</div>
            <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E;">OCF (%)</div>
            <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E;">ESG Globe / Carbon Risk</div>
        </div>
        """, unsafe_allow_html=True)
        for i in range(num_funds):
            col1, col2, col3, col4 = st.columns([1.2, 1, 1, 1])
            with col1:
                st.markdown(f'<div style="padding-top:0.5rem; font-size:0.875rem; font-weight:500; color:#000000;">{st.session_state.fund_names[i]}</div>', unsafe_allow_html=True)
            with col2:
                st.session_state.costs[st.session_state.fund_names[i]] = st.number_input(f"OCF", value=st.session_state.costs.get(st.session_state.fund_names[i], 0.99), step=0.01, key=f"ocf_{i}", label_visibility="collapsed")
            with col3:
                st.session_state.esg_globe[st.session_state.fund_names[i]] = st.number_input(f"ESG Globe", value=st.session_state.esg_globe.get(st.session_state.fund_names[i], 4), min_value=1, max_value=5, key=f"esg_{i}", label_visibility="collapsed")
            with col4:
                st.session_state.carbon_risk[st.session_state.fund_names[i]] = st.number_input(f"Carbon Risk", value=st.session_state.carbon_risk.get(st.session_state.fund_names[i], 7.85), step=0.01, key=f"carbon_{i}", label_visibility="collapsed")
        # Extra custom funds also need cost/ESG entries
        for _cef in st.session_state.get("extra_funds", []):
            _cef_name = _cef["name"]
            _cef_tk   = _cef["ticker"]
            _cef_c1, _cef_c2, _cef_c3, _cef_c4 = st.columns([1.2, 1, 1, 1])
            with _cef_c1:
                st.markdown(f'<div style="padding-top:0.5rem; font-size:0.875rem; font-weight:500; color:#154D57;">{_cef_name} <span style="font-size:0.7rem;color:#9B9B9B;">({_cef_tk})</span></div>', unsafe_allow_html=True)
            with _cef_c2:
                st.session_state.costs[_cef_name] = st.number_input("OCF", value=st.session_state.costs.get(_cef_name, 0.99), step=0.01, key=f"ocf_extra_{_cef_tk}", label_visibility="collapsed")
            with _cef_c3:
                st.session_state.esg_globe[_cef_name] = st.number_input("ESG Globe", value=st.session_state.esg_globe.get(_cef_name, 4), min_value=1, max_value=5, key=f"esg_extra_{_cef_tk}", label_visibility="collapsed")
            with _cef_c4:
                st.session_state.carbon_risk[_cef_name] = st.number_input("Carbon Risk", value=st.session_state.carbon_risk.get(_cef_name, 7.85), step=0.01, key=f"carbon_extra_{_cef_tk}", label_visibility="collapsed")

    st.markdown('<hr style="border:none; border-top:1px solid #E8DDD3; margin:2.5rem 0;">', unsafe_allow_html=True)

    # ── RUN ANALYSIS ──────────────────────────────────────────────────────
    col1, col2, col3 = st.columns([1,1,1])
    with col2:
        run_clicked = st.button("Run Analysis", key="run_button")

    # Full-width placeholders rendered OUTSIDE the column so they span the page
    _progress_slot = st.empty()
    _status_slot = st.empty()
    _results_slot = st.empty()

    if run_clicked:
        # Skeleton loading screen  -  appears instantly while data is fetched
        _results_slot.markdown("""
        <div style="animation: fadeIn 0.2s ease-out;">
            <div style="background:#EAF2F3; border:1px solid #A3C9CE; border-radius:10px; padding:0.75rem 1.25rem; margin:1rem 0; display:flex; align-items:center; gap:0.75rem; opacity:0.55;">
                <div style="width:6px; height:6px; border-radius:50%; background:#A3C9CE; flex-shrink:0;"></div>
                <span class="skeleton-card" style="height:13px; width:220px;"></span>
            </div>
            <div style="display:grid; grid-template-columns:repeat(5,1fr); gap:1px; background:#E8DDD3; border-radius:12px; overflow:hidden; border:1px solid #E8DDD3; margin:1rem 0;">
                <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                    <span class="skeleton-card" style="height:9px; width:65%; margin-bottom:0.9rem;"></span>
                    <span class="skeleton-card" style="height:26px; width:38%;"></span>
                </div>
                <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                    <span class="skeleton-card" style="height:9px; width:72%; margin-bottom:0.9rem;"></span>
                    <span class="skeleton-card" style="height:26px; width:44%;"></span>
                </div>
                <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                    <span class="skeleton-card" style="height:9px; width:68%; margin-bottom:0.9rem;"></span>
                    <span class="skeleton-card" style="height:26px; width:40%;"></span>
                </div>
                <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                    <span class="skeleton-card" style="height:9px; width:75%; margin-bottom:0.9rem;"></span>
                    <span class="skeleton-card" style="height:18px; width:58%;"></span>
                </div>
                <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                    <span class="skeleton-card" style="height:9px; width:70%; margin-bottom:0.9rem;"></span>
                    <span class="skeleton-card" style="height:16px; width:52%;"></span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        _progress_slot.markdown('<div class="ind-track"><div class="ind-bar"></div></div>', unsafe_allow_html=True)
        try:
            # Extend with any custom extra funds before fetching data
            _run_extra = st.session_state.get("extra_funds", [])
            for _ref in _run_extra:
                if _ref["ticker"] not in st.session_state.tickers:
                    st.session_state.tickers.append(_ref["ticker"])
                    st.session_state.fund_names.append(_ref["name"])
            log_returns = fetch_data(st.session_state.tickers, benchmark_ticker, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
            # Compute metrics
            core_metrics = {}
            downside_metrics = {}
            for i, name in enumerate(st.session_state.fund_names):
                core = compute_core_metrics(log_returns[st.session_state.tickers[i]], log_returns[benchmark_ticker], rf_annual)
                core_metrics[name] = core
                downside = compute_downside_metrics(log_returns[st.session_state.tickers[i]], log_returns[benchmark_ticker])
                downside_metrics[name] = downside
            # Benchmark
            core_bench = compute_core_metrics(log_returns[benchmark_ticker], log_returns[benchmark_ticker], rf_annual)
            core_metrics[benchmark_name] = core_bench
            downside_bench = compute_downside_metrics(log_returns[benchmark_ticker], log_returns[benchmark_ticker])
            downside_metrics[benchmark_name] = downside_bench
            # DataFrames
            core_df = pd.DataFrame(core_metrics).T
            downside_df = pd.DataFrame(downside_metrics).T
            # Add manual
            core_df['OCF'] = [st.session_state.costs.get(name, np.nan) for name in core_df.index]
            core_df['ESG Globe Rating'] = [st.session_state.esg_globe.get(name, np.nan) for name in core_df.index]
            core_df['Carbon Risk Score'] = [st.session_state.carbon_risk.get(name, np.nan) for name in core_df.index]
            metrics_matrix = pd.concat([core_df, downside_df], axis=1).T
            # Store
            st.session_state.log_returns = log_returns
            st.session_state.fund_returns = log_returns[st.session_state.tickers]
            st.session_state.benchmark_returns = log_returns[benchmark_ticker]
            st.session_state.core_metrics_df = core_df
            st.session_state.downside_df = downside_df
            st.session_state.metrics_matrix = metrics_matrix
            st.session_state.analysis_run = True
            _progress_slot.empty()
            _status_slot.empty()
            best_fund = core_df.drop(benchmark_name)['Sharpe Ratio'].idxmax()
            n_funds = len(st.session_state.fund_names)
            n_obs = len(log_returns)
            _results_slot.markdown(f"""
            <div class="results-reveal">
                <div style="background:#EAF2F3; border:1px solid #A3C9CE; border-radius:10px; padding:0.75rem 1.25rem; margin:1rem 0; display:flex; align-items:center; gap:0.75rem;">
                    <div style="width:6px; height:6px; border-radius:50%; background:#154D57; flex-shrink:0;"></div>
                    <span style="color:#154D57; font-size:0.875rem; font-weight:500;">Analysis complete. Use the sidebar to explore results.</span>
                </div>
                <div style="display:grid; grid-template-columns:repeat(5,1fr); gap:1px; background:#E8DDD3; border-radius:12px; overflow:hidden; border:1px solid #E8DDD3; margin:1rem 0;">
                    <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E; margin-bottom:0.75rem; white-space:nowrap;">Total Funds</div>
                        <div id="kpi-funds" data-counter="{n_funds}" data-fmt="int" style="font-size:1.75rem; font-weight:700; color:#0A0A0A; letter-spacing:-0.03em; line-height:1; font-variant-numeric:tabular-nums;">{n_funds}</div>
                    </div>
                    <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E; margin-bottom:0.75rem; white-space:nowrap;">Sample Period</div>
                        <div style="font-size:1.75rem; font-weight:700; color:#0A0A0A; letter-spacing:-0.03em; line-height:1; font-variant-numeric:tabular-nums;">
                            <span id="kpi-period" data-counter="{n_obs}" data-fmt="int">{n_obs}</span><span style="font-size:0.75rem; font-weight:500; color:#9B8C7E; margin-left:0.3rem;">mo</span>
                        </div>
                    </div>
                    <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E; margin-bottom:0.75rem; white-space:nowrap;">Observations</div>
                        <div id="kpi-obs" data-counter="{n_obs}" data-fmt="int" style="font-size:1.75rem; font-weight:700; color:#0A0A0A; letter-spacing:-0.03em; line-height:1; font-variant-numeric:tabular-nums;">{n_obs}</div>
                    </div>
                    <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E; margin-bottom:0.75rem; white-space:nowrap;">Best Sharpe</div>
                        <div style="font-size:1rem; font-weight:700; color:#154D57; letter-spacing:-0.015em; line-height:1.3;">{best_fund}</div>
                    </div>
                    <div style="background:#FFFFFF; padding:1.25rem 1.5rem;">
                        <div style="font-size:0.6875rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#9B8C7E; margin-bottom:0.75rem; white-space:nowrap;">Date Analysed</div>
                        <div style="font-size:0.875rem; font-weight:600; color:#0A0A0A; letter-spacing:-0.01em; line-height:1.3;">{datetime.now().strftime('%d %b %Y')}</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            # KPI counter animation  -  counts up from 0 to final value on results reveal
            components.html("""
            <script>
            (function() {
                function easeOutExpo(t) { return t >= 1 ? 1 : 1 - Math.pow(2, -10 * t); }
                function animateCounter(el) {
                    var end = parseFloat(el.getAttribute('data-counter'));
                    var fmt = el.getAttribute('data-fmt') || 'int';
                    var dur = Math.max(500, Math.min(900, end * 30));
                    var startTime = performance.now();
                    function tick(now) {
                        var p = Math.min((now - startTime) / dur, 1);
                        var val = easeOutExpo(p) * end;
                        el.textContent = fmt === 'int' ? Math.round(val) : val.toFixed(2);
                        if (p < 1) { requestAnimationFrame(tick); }
                        else { el.textContent = fmt === 'int' ? Math.round(end) : end.toFixed(2); }
                    }
                    requestAnimationFrame(tick);
                }
                // Wait for React to mount the markdown output
                setTimeout(function() {
                    var els = window.parent.document.querySelectorAll('[data-counter]');
                    els.forEach(function(el) { animateCounter(el); });
                }, 180);
            })();
            </script>
            """, height=0)
        except Exception as e:
            _progress_slot.empty()
            _status_slot.empty()
            _results_slot.error(f"Could not download data. Please check the ticker symbols. Error: {str(e)}")
    # Footer
    st.markdown("""
    <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
        <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
            Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
        </p>
        <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
            Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
        </p>
    </div>
    """, unsafe_allow_html=True)
elif page == "Performance":
    if not st.session_state.analysis_run:
        st.markdown("# Performance Overview")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Monthly log returns, risk-adjusted metrics, and drawdown analysis across the full sample period.
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center; padding: 4rem 2rem; background: #F5EEE7; border-radius: 20px; border: 2px dashed #D4C3B0;">
            <div style="margin-bottom: 1rem;"><svg width="48" height="48" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><rect x="3" y="13" width="4" height="8" rx="1" fill="#B7A08B"/><rect x="10" y="8" width="4" height="13" rx="1" fill="#B7A08B"/><rect x="17" y="4" width="4" height="17" rx="1" fill="#B7A08B"/><line x1="2" y1="21" x2="22" y2="21" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/></svg></div>
            <div style="font-size: 1rem; font-weight: 600; color: #0A0A0A; margin-bottom: 0.5rem;">No data yet</div>
            <div style="color: #7A6F65; margin-bottom: 1.5rem;">Head to Setup and click Run Analysis to begin.</div>
        </div>
        """, unsafe_allow_html=True)
        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("# Performance Overview")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Monthly log returns, risk-adjusted metrics, and drawdown analysis across the full sample period.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2, tab3 = st.tabs(["Metrics", "Charts", "Details"])
        
        with tab1:
            st.markdown("## Core Metrics Table")
            BEST_STYLE  = "background-color: #EAF2F3; color: #154D57"
            WORST_STYLE = "background-color: #F8F3EE; color: #7A5642"

            def highlight_metrics_table(df, directions, benchmark_name="Benchmark"):
                styles = pd.DataFrame("", index=df.index, columns=df.columns)
                comparison_rows = [idx for idx in df.index if idx != benchmark_name]
                for metric, direction in directions.items():
                    if metric not in df.columns:
                        continue
                    values = pd.to_numeric(df.loc[comparison_rows, metric], errors="coerce").dropna()
                    if values.empty:
                        continue
                    if direction == "higher":
                        best_val, worst_val = values.max(), values.min()
                    else:
                        best_val, worst_val = values.min(), values.max()
                    styles.loc[values[values == best_val].index, metric]  = BEST_STYLE
                    styles.loc[values[values == worst_val].index, metric] = WORST_STYLE
                return styles

            CORE_DIRECTIONS = {
                "Ann. Return (%)":      "higher",
                "Ann. Volatility (%)":  "lower",
                "Alpha (ann. %)":       "higher",
                "Beta":                 "lower",
                "R²":                   "higher",
                "Sharpe Ratio":         "higher",
                "Treynor Ratio":        "higher",
                "Sortino Ratio":        "higher",
                "Information Ratio":    "higher",
                "Tracking Error (%)":   "lower",
            }
            DOWNSIDE_DIRECTIONS = {
                "Upside Capture (%)":     "higher",
                "Downside Capture (%)":   "lower",
                "Max Drawdown (%)":       "higher",
                "Max DD Duration (mths)": "lower",
                "Calmar Ratio":           "higher",
            }

            core_df_r = st.session_state.core_metrics_df.round(3)
            styled_core = core_df_r.style.apply(
                lambda df: highlight_metrics_table(df, CORE_DIRECTIONS, st.session_state.benchmark_name),
                axis=None
            )
            st.dataframe(styled_core, use_container_width=True)

            st.markdown("## Downside Metrics Table")
            downside_df_r = st.session_state.downside_df.round(3)
            styled_downside = downside_df_r.style.apply(
                lambda df: highlight_metrics_table(df, DOWNSIDE_DIRECTIONS, st.session_state.benchmark_name),
                axis=None
            )
            st.dataframe(styled_downside, use_container_width=True)

            with st.expander("📖 Metric Glossary"):
                GLOSSARY = {
                    "Ann. Return (%)":        ("Annualised Return", "The geometric mean of monthly log returns scaled to one year. Computed as (exp(r̄ × 12) − 1) × 100 where r̄ is the mean monthly log return."),
                    "Ann. Volatility (%)":    ("Annualised Volatility", "The standard deviation of monthly returns scaled by √12. Higher values mean returns are more variable over time."),
                    "Alpha (ann. %)":         ("Jensen's Alpha", "The fund's excess return above what CAPM theory predicts given its market exposure. Computed from OLS regression of fund excess returns on benchmark excess returns. Positive alpha suggests manager skill."),
                    "Beta":                   ("Market Beta", "The sensitivity of fund returns to benchmark movements. A beta of 0.8 means the fund typically moves 0.8% for every 1% benchmark move. Lower beta = less market sensitivity."),
                    "Sharpe Ratio":           ("Sharpe Ratio", "Excess return (above the risk-free rate) per unit of total volatility. The most widely used risk-adjusted performance measure. Higher is better."),
                    "Sortino Ratio":          ("Sortino Ratio", "Like the Sharpe ratio but only penalises downside volatility (negative returns). More relevant when return distributions are asymmetric or negatively skewed."),
                    "Treynor Ratio":          ("Treynor Ratio", "Excess return per unit of market risk (beta). Uses systematic risk rather than total risk, which makes it more relevant when evaluating well-diversified portfolios."),
                    "Information Ratio":      ("Information Ratio", "Active return (fund minus benchmark) divided by tracking error. Measures the consistency of outperformance. An IR above 0.5 is generally considered strong."),
                    "R²":                     ("R-Squared", "The proportion of fund return variance explained by benchmark movements. A high R² (near 1.0) indicates the fund tracks the benchmark closely."),
                    "Tracking Error (%)":     ("Tracking Error", "The standard deviation of the difference between fund and benchmark monthly returns. Low tracking error means the fund moves closely with the benchmark."),
                    "Max Drawdown (%)":       ("Maximum Drawdown", "The largest peak-to-trough decline in the cumulative wealth index. Computed using W_t = exp(∑r_t). A key measure of tail risk and loss potential."),
                    "Max DD Duration (mths)": ("Max DD Duration", "The number of months from the peak before the maximum drawdown to the point where the fund fully recovered to that peak level."),
                    "Downside Capture (%)":   ("Downside Capture Ratio", "The fund's average return in months when the benchmark fell, expressed as a percentage of the benchmark's average return in those months. Below 100% means the fund lost less than the benchmark in down markets."),
                    "Upside Capture (%)":     ("Upside Capture Ratio", "The fund's average return in months when the benchmark rose, as a percentage of the benchmark's gain. Above 100% means the fund outperformed in up markets."),
                    "Calmar Ratio":           ("Calmar Ratio", "Annualised return divided by the absolute value of maximum drawdown. Balances return against the worst historical loss. Higher is better."),
                    "OCF":                    ("Ongoing Charges Figure", "The total annual cost of owning the fund as a percentage of NAV. Includes management fees and operating expenses but excludes transaction costs."),
                    "ESG Globe Rating":       ("ESG Globe Rating", "Morningstar's sustainability rating on a scale of 1-5 globes. Based on the fund's Sustainalytics portfolio ESG risk score. 5 globes = highest sustainability."),
                    "Carbon Risk Score":      ("Carbon Risk Score", "Sustainalytics' measure of a fund's exposure to carbon transition risk. Lower scores indicate the portfolio is less exposed to risks from the transition to a low-carbon economy."),
                }
                for metric_key, (metric_name, definition) in GLOSSARY.items():
                    st.markdown(f"""
                    <div style="display:flex;gap:1rem;padding:0.75rem 0;
                                border-bottom:1px solid #F0EFEC;">
                        <div style="min-width:160px;max-width:160px;">
                            <span style="font-size:0.78rem;font-weight:700;color:#154D57;">{metric_key}</span><br>
                            <span style="font-size:0.7rem;color:#9B9B9B;">{metric_name}</span>
                        </div>
                        <div style="font-size:0.8rem;color:#5A5A5A;line-height:1.55;flex:1;">
                            {definition}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

            st.markdown("## Key Metric Cards")
            sharpe_series = st.session_state.core_metrics_df['Sharpe Ratio']
            best_sharpe_fund = sharpe_series.drop(st.session_state.benchmark_name).idxmax()
            best_sharpe_val = sharpe_series[best_sharpe_fund]
            downside_cap_series = st.session_state.downside_df['Downside Capture (%)']
            best_downside_cap_fund = downside_cap_series.drop(st.session_state.benchmark_name).idxmin()
            best_downside_cap_val = downside_cap_series[best_downside_cap_fund]
            max_dd_series = st.session_state.downside_df['Max Drawdown (%)']
            lowest_dd_fund = max_dd_series.drop(st.session_state.benchmark_name).idxmin()
            lowest_dd_val = max_dd_series[lowest_dd_fund]
            alpha_series = st.session_state.core_metrics_df['Alpha (ann. %)']
            best_alpha_fund = alpha_series.drop(st.session_state.benchmark_name).idxmax()
            best_alpha_val = alpha_series[best_alpha_fund]
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #154D57 0%, #0F3940 100%); color: white; padding: 1.5rem; border-radius: 16px; margin-bottom: 1rem; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);">
                    <div style="font-size: 0.6875rem; font-weight: 700; letter-spacing: 0.12em; opacity: 0.75; margin-bottom: 0.5rem;">BEST SHARPE RATIO</div>
                    <div style="font-size: 2rem; font-weight: 700; margin-bottom: 0.25rem;">{best_sharpe_val:.2f}</div>
                    <div style="font-size: 0.875rem; font-weight: 500; opacity: 0.85;">{best_sharpe_fund}</div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #B7A08B 0%, #8B7668 100%); color: white; padding: 1.5rem; border-radius: 16px; margin-bottom: 1rem; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);">
                    <div style="font-size: 0.6875rem; font-weight: 700; letter-spacing: 0.12em; opacity: 0.75; margin-bottom: 0.5rem;">LOWEST MAX DRAWDOWN</div>
                    <div style="font-size: 2rem; font-weight: 700; margin-bottom: 0.25rem;">{lowest_dd_val:.2f}%</div>
                    <div style="font-size: 0.875rem; font-weight: 500; opacity: 0.85;">{lowest_dd_fund}</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #1A6B77 0%, #154D57 100%); color: white; padding: 1.5rem; border-radius: 16px; margin-bottom: 1rem; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);">
                    <div style="font-size: 0.6875rem; font-weight: 700; letter-spacing: 0.12em; opacity: 0.75; margin-bottom: 0.5rem;">BEST DOWNSIDE CAPTURE</div>
                    <div style="font-size: 2rem; font-weight: 700; margin-bottom: 0.25rem;">{best_downside_cap_val:.2f}%</div>
                    <div style="font-size: 0.875rem; font-weight: 500; opacity: 0.85;">{best_downside_cap_fund}</div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #C9B8A4 0%, #B7A08B 100%); color: white; padding: 1.5rem; border-radius: 16px; margin-bottom: 1rem; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);">
                    <div style="font-size: 0.6875rem; font-weight: 700; letter-spacing: 0.12em; opacity: 0.75; margin-bottom: 0.5rem;">BEST ALPHA</div>
                    <div style="font-size: 2rem; font-weight: 700; margin-bottom: 0.25rem;">{best_alpha_val:.2f}%</div>
                    <div style="font-size: 0.875rem; font-weight: 500; opacity: 0.85;">{best_alpha_fund}</div>
                </div>
                """, unsafe_allow_html=True)
        
        with tab2:
            st.markdown("## Charts")
            st.markdown('<div style="font-size:1rem; font-weight:600; color:#0A0A0A; margin-bottom:0.25rem; letter-spacing:-0.01em;">Cumulative Returns</div><div style="font-size:0.875rem; color:#7A6F65; margin-bottom:0.5rem;">Total wealth index growth from January 2020 to October 2025, showing relative performance over the full period.</div>', unsafe_allow_html=True)
            fig1 = chart_cumulative_returns(st.session_state.log_returns, COLOURS, PLOTLY_LAYOUT)
            st.plotly_chart(fig1, use_container_width=True, config={"displayModeBar": False})
            try:
                _fn_local = st.session_state.fund_names
                _mm = st.session_state.metrics_matrix
                _best_ret_fund = _mm.loc["Ann. Return (%)", _fn_local].astype(float).idxmax()
                _best_ret_val = float(_mm.loc["Ann. Return (%)", _best_ret_fund])
                _worst_ret_fund = _mm.loc["Ann. Return (%)", _fn_local].astype(float).idxmin()
                _bm_ret = float(_mm.loc["Ann. Return (%)", "Benchmark"])
                _funds_beating_bm = [f for f in _fn_local if float(_mm.loc["Ann. Return (%)", f]) > _bm_ret]
                auto_commentary(
                    "{best} leads the peer group with an annualised return of {ret:.1f}%, "
                    "while {worst} trails with the lowest cumulative return over the period. "
                    "{n_beat} of {total} funds outperformed the benchmark return of {bm:.1f}% per annum.",
                    best=_best_ret_fund, ret=_best_ret_val,
                    worst=_worst_ret_fund, n_beat=len(_funds_beating_bm),
                    total=len(_fn_local), bm=_bm_ret
                )
            except Exception as _e:
                logging.debug("Returns chart commentary skipped (non-critical): %s", _e)

            st.markdown('<div style="font-size:1rem; font-weight:600; color:#0A0A0A; margin-bottom:0.25rem; letter-spacing:-0.01em;">Drawdown Analysis</div><div style="font-size:0.875rem; color:#7A6F65; margin-bottom:0.5rem;">Peak-to-trough declines in portfolio value, highlighting maximum loss periods and recovery patterns.</div>', unsafe_allow_html=True)
            fig2 = chart_drawdown(st.session_state.fund_returns, st.session_state.benchmark_returns, COLOURS, PLOTLY_LAYOUT)
            st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar": False})
            try:
                _fn_local = st.session_state.fund_names
                _mm = st.session_state.metrics_matrix
                _best_dd_fund = _mm.loc["Max Drawdown (%)", _fn_local].astype(float).idxmax()
                _best_dd_val = float(_mm.loc["Max Drawdown (%)", _best_dd_fund])
                _worst_dd_fund = _mm.loc["Max Drawdown (%)", _fn_local].astype(float).idxmin()
                _worst_dd_val = float(_mm.loc["Max Drawdown (%)", _worst_dd_fund])
                _best_dc_fund = _mm.loc["Downside Capture (%)", _fn_local].astype(float).idxmin()
                _best_dc_val = float(_mm.loc["Downside Capture (%)", _best_dc_fund])
                auto_commentary(
                    "{best} demonstrated the strongest capital preservation with a maximum drawdown of just {dd:.1f}%, "
                    "compared to {worst} which fell {wdd:.1f}% from peak to trough. "
                    "{dc_fund} showed the best downside protection, capturing only {dc:.1f}% of benchmark losses in negative months.",
                    best=_best_dd_fund, dd=_best_dd_val,
                    worst=_worst_dd_fund, wdd=_worst_dd_val,
                    dc_fund=_best_dc_fund, dc=_best_dc_val
                )
            except Exception as _e:
                logging.debug("Drawdown chart commentary skipped (non-critical): %s", _e)

            st.markdown('<div style="font-size:1rem; font-weight:600; color:#0A0A0A; margin-bottom:0.25rem; letter-spacing:-0.01em;">Rolling Sharpe Ratio</div><div style="font-size:0.875rem; color:#7A6F65; margin-bottom:0.5rem;">12-month rolling risk-adjusted return measure, showing consistency of performance over time.</div>', unsafe_allow_html=True)
            fig3 = chart_rolling_sharpe(st.session_state.fund_returns, st.session_state.rf_annual / 12, COLOURS, PLOTLY_LAYOUT)
            st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar": False})
            try:
                _fn_local = st.session_state.fund_names
                _mm = st.session_state.metrics_matrix
                _best_sr_fund = _mm.loc["Sharpe Ratio", _fn_local].astype(float).idxmax()
                _best_sr_val = float(_mm.loc["Sharpe Ratio", _best_sr_fund])
                _funds_pos_sr = [f for f in _fn_local if float(_mm.loc["Sharpe Ratio", f]) > 0]
                auto_commentary(
                    "{best} achieves the highest risk-adjusted return with a Sharpe ratio of {s:.3f}, "
                    "meaning it generates {s:.2f} units of excess return per unit of volatility. "
                    "{n_pos} of {total} funds maintain a positive Sharpe ratio over the full sample period.",
                    best=_best_sr_fund, s=_best_sr_val,
                    n_pos=len(_funds_pos_sr), total=len(_fn_local)
                )
            except Exception as _e:
                logging.debug("Sharpe chart commentary skipped (non-critical): %s", _e)

        with tab3:
            with st.expander("ℹ️ Metric definitions"):
                definitions = {
                    "Ann. Return (%)": "Effective annualised return: (exp(mean × 12) − 1) × 100",
                    "Sharpe Ratio": "Annualised excess return per unit of volatility",
                    "Sortino Ratio": "Like Sharpe but only penalises downside volatility",
                    "Max Drawdown (%)": "Largest peak-to-trough decline in the wealth index",
                    "Calmar Ratio": "Annualised return divided by absolute max drawdown",
                    "Downside Capture": "Fund return in down months as % of benchmark return",
                    "Alpha": "Excess return above CAPM expectation (annualised)",
                    "Beta": "Sensitivity of fund returns to benchmark movements",
                    "Information Ratio": "Active return per unit of tracking error",
                }
                def_df = pd.DataFrame(list(definitions.items()), columns=["Metric", "Definition"])
                st.dataframe(def_df.round(3) if def_df.select_dtypes(include=[np.number]).any().any() else def_df, use_container_width=True)
        
        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)

elif page == "TOPSIS":
    if not st.session_state.analysis_run:
        st.markdown("# TOPSIS Ranking")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Technique for Order Preference by Similarity to Ideal Solution (Hwang & Yoon, 1981) with interactive pillar weights.
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center; padding: 4rem 2rem; background: #F5EEE7; border-radius: 20px; border: 2px dashed #D4C3B0;">
            <div style="margin-bottom: 1rem;"><svg width="48" height="48" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><circle cx="12" cy="12" r="9" stroke="#B7A08B" stroke-width="1.5"/><circle cx="12" cy="12" r="5" stroke="#B7A08B" stroke-width="1.5"/><circle cx="12" cy="12" r="1.5" fill="#B7A08B"/><line x1="12" y1="2" x2="12" y2="4" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/><line x1="12" y1="20" x2="12" y2="22" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/><line x1="2" y1="12" x2="4" y2="12" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/><line x1="20" y1="12" x2="22" y2="12" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/></svg></div>
            <div style="font-size: 1rem; font-weight: 600; color: #0A0A0A; margin-bottom: 0.5rem;">No data yet</div>
            <div style="color: #7A6F65; margin-bottom: 1.5rem;">Head to Setup and click Run Analysis to begin.</div>
        </div>
        """, unsafe_allow_html=True)
        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("# TOPSIS Ranking")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Technique for Order Preference by Similarity to Ideal Solution (Hwang & Yoon, 1981) with interactive pillar weights.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2 = st.tabs(["Rankings", "Methodology"])
        
        with tab1:
            st.info("TOPSIS ranks alternatives by measuring their geometric distance from an ideal solution and anti-ideal solution in a multi-dimensional space. The score S_i = D⁻ / (D⁺ + D⁻) represents closeness to the ideal.")
            
            # Weights
            st.markdown("## Pillar Weight Sliders")
            PRESETS = {
                "⚡ Return Seeker":       {"returns": 60, "risk_adj": 20, "risk_dd": 10, "costs": 5,  "esg": 5},
                "🛡️ Downside Protection": {"returns": 20, "risk_adj": 25, "risk_dd": 45, "costs": 5,  "esg": 5},
                "🌱 ESG Investor":        {"returns": 25, "risk_adj": 20, "risk_dd": 15, "costs": 10, "esg": 30},
                "💰 Low Cost":            {"returns": 30, "risk_adj": 25, "risk_dd": 20, "costs": 20, "esg": 5},
            }
            st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.5rem;">QUICK PRESETS</p>', unsafe_allow_html=True)
            _pc1, _pc2, _pc3, _pc4 = st.columns(4)
            for _col, (_pname, _pw) in zip([_pc1, _pc2, _pc3, _pc4], PRESETS.items()):
                with _col:
                    if st.button(_pname, key=f"preset_{_pname}", use_container_width=True):
                        st.session_state["w_returns"]  = _pw["returns"]
                        st.session_state["w_risk_adj"] = _pw["risk_adj"]
                        st.session_state["w_risk_dd"]  = _pw["risk_dd"]
                        st.session_state["w_costs"]    = _pw["costs"]
                        st.session_state["w_esg"]      = _pw["esg"]
                        st.rerun()
            st.markdown('<p style="font-size:0.75rem;color:#9B9B9B;margin-top:0.25rem;">Select a preset or adjust sliders manually. Weights must sum to 100%.</p>', unsafe_allow_html=True)
            st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin:1rem 0 0.5rem 0;">EXCLUDE PILLARS</p>', unsafe_allow_html=True)
            _ep1, _ep2, _ep3, _ep4, _ep5 = st.columns(5)
            _excluded_pillars = set()
            for _pn, _ec in [("Returns", _ep1), ("Risk-Adj", _ep2), ("Risk/DD", _ep3), ("Costs", _ep4), ("ESG", _ep5)]:
                with _ec:
                    if st.checkbox(f"Excl. {_pn}", key=f"excl_{_pn}", value=False):
                        _excluded_pillars.add(_pn)
            if _excluded_pillars:
                st.markdown(f'<p style="font-size:0.78rem;color:#B7A08B;margin-top:0.25rem;">⚠️ Excluding: {", ".join(_excluded_pillars)}. Weights redistributed across remaining pillars.</p>', unsafe_allow_html=True)

            # ── Custom Scenario Builder ───────────────────────────────────────
            # Apply pending slider weights BEFORE sliders instantiate (avoids widget key conflict)
            if st.session_state.get("_apply_slider_weights"):
                for _sk in ["w_returns", "w_risk_adj", "w_risk_dd", "w_costs", "w_esg"]:
                    _spv = st.session_state.get(f"_pending_alloc_{_sk}")
                    if _spv is not None:
                        st.session_state[_sk] = _spv
                        del st.session_state[f"_pending_alloc_{_sk}"]
                del st.session_state["_apply_slider_weights"]

            st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin:1rem 0 0.5rem 0;">CUSTOM SCENARIOS</p>', unsafe_allow_html=True)
            _saved = st.session_state.get("saved_scenarios", {})
            _scen_col1, _scen_col2 = st.columns([3, 1])
            with _scen_col1:
                _scenario_name = st.text_input(
                    "Scenario name",
                    placeholder="e.g. My Balanced Approach",
                    key="scenario_name_input",
                    label_visibility="collapsed"
                )
            with _scen_col2:
                if st.button("💾 Save Current Weights", key="save_scenario_btn", use_container_width=True):
                    if _scenario_name.strip():
                        st.session_state.saved_scenarios[_scenario_name.strip()] = {
                            "returns":  st.session_state.get("w_returns", 40),
                            "risk_adj": st.session_state.get("w_risk_adj", 25),
                            "risk_dd":  st.session_state.get("w_risk_dd", 20),
                            "costs":    st.session_state.get("w_costs", 10),
                            "esg":      st.session_state.get("w_esg", 5),
                        }
                        st.success(f"Saved: {_scenario_name.strip()}")
                    else:
                        st.warning("Enter a name for the scenario first.")

            if _saved:
                st.markdown('<p style="font-size:0.72rem;color:#9B9B9B;margin:0.5rem 0 0.25rem 0;">Saved scenarios (click to load):</p>', unsafe_allow_html=True)
                _saved_cols = st.columns(min(len(_saved), 4))
                for _scol, (_sname, _sweights) in zip(_saved_cols, list(_saved.items())[:4]):
                    with _scol:
                        if st.button(f"📂 {_sname}", key=f"load_scenario_{_sname}", use_container_width=True):
                            st.session_state["_pending_alloc_w_returns"]  = _sweights["returns"]
                            st.session_state["_pending_alloc_w_risk_adj"] = _sweights["risk_adj"]
                            st.session_state["_pending_alloc_w_risk_dd"]  = _sweights["risk_dd"]
                            st.session_state["_pending_alloc_w_costs"]    = _sweights["costs"]
                            st.session_state["_pending_alloc_w_esg"]      = _sweights["esg"]
                            st.session_state["_apply_slider_weights"]     = True
                            st.rerun()

                with st.expander("🗑️ Manage saved scenarios"):
                    for _sname in list(_saved.keys()):
                        _del_c1, _del_c2 = st.columns([4, 1])
                        with _del_c1:
                            _sw = _saved[_sname]
                            st.markdown(
                                f'<span style="font-size:0.82rem;color:#5A5A5A;">'
                                f'<strong>{_sname}</strong>: '
                                f'Returns {_sw["returns"]}% / Risk-Adj {_sw["risk_adj"]}% / '
                                f'Risk/DD {_sw["risk_dd"]}% / Costs {_sw["costs"]}% / ESG {_sw["esg"]}%'
                                f'</span>',
                                unsafe_allow_html=True
                            )
                        with _del_c2:
                            if st.button("Delete", key=f"del_scenario_{_sname}", use_container_width=True):
                                del st.session_state.saved_scenarios[_sname]
                                st.rerun()

            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                returns_w = st.slider("Returns", 0, 100, st.session_state.get("w_returns", 40), key="w_returns")
            with col2:
                risk_adj_w = st.slider("Risk-Adj", 0, 100, st.session_state.get("w_risk_adj", 25), key="w_risk_adj")
            with col3:
                risk_dd_w = st.slider("Risk/DD", 0, 100, st.session_state.get("w_risk_dd", 20), key="w_risk_dd")
            with col4:
                costs_w = st.slider("Costs", 0, 100, st.session_state.get("w_costs", 10), key="w_costs")
            with col5:
                esg_w = st.slider("ESG", 0, 100, st.session_state.get("w_esg", 5), key="w_esg")
            pillar_weights = {'Returns': returns_w, 'Risk-Adj': risk_adj_w, 'Risk/DD': risk_dd_w, 'Costs': costs_w, 'ESG': esg_w}
            total_w = sum(pillar_weights.values())

            with st.expander("⚙️ Advanced: Include / Exclude Metrics"):
                st.markdown('<p style="font-size:0.78rem;color:#7A6F65;margin-bottom:1rem;line-height:1.5;">Toggle individual metrics on or off. Weights redistribute proportionally within each pillar. Pillars marked as excluded above are greyed out here.</p>', unsafe_allow_html=True)
                _TOGGLE_PILLARS = {
                    "Returns":  (["Ann. Return (%)", "Alpha (ann. %)"],                                                "#154D57"),
                    "Risk-Adj": (["Sharpe Ratio", "Sortino Ratio", "Treynor Ratio", "Information Ratio", "R²"],        "#1A6B77"),
                    "Risk/DD":  (["Ann. Volatility (%)", "Beta", "Tracking Error (%)", "Max Drawdown (%)",
                                  "Max DD Duration (mths)", "Downside Capture (%)", "Upside Capture (%)", "Calmar Ratio"], "#B7A08B"),
                    "Costs":    (["OCF"],                                                                               "#8B7668"),
                    "ESG":      (["ESG Globe Rating", "Carbon Risk Score"],                                            "#A3C9CE"),
                }
                for _tpillar, (_tmetrics, _tpcol) in _TOGGLE_PILLARS.items():
                    _tis_excl = _tpillar in _excluded_pillars
                    _talpha   = "0.38" if _tis_excl else "1"
                    _tnote    = " <span style='font-size:0.62rem;color:#B7A08B;font-weight:500;'>(pillar excluded)</span>" if _tis_excl else ""
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:0.5rem;margin:1.1rem 0 0.4rem 0;opacity:{_talpha};">'
                        f'<div style="width:7px;height:7px;border-radius:50%;background:{_tpcol};flex-shrink:0;"></div>'
                        f'<span style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#3A3A3A;">{_tpillar}</span>'
                        f'{_tnote}</div>',
                        unsafe_allow_html=True
                    )
                    _MAX_COLS = 4
                    for _tr in range(0, len(_tmetrics), _MAX_COLS):
                        _trow = _tmetrics[_tr:_tr + _MAX_COLS]
                        _tmcols = st.columns(4)
                        for _tj, _tm in enumerate(_trow):
                            with _tmcols[_tj]:
                                st.checkbox(_tm, value=True, key=f"mt_{_tm}", disabled=_tis_excl)

            if total_w == 100:
                st.session_state.pillar_weights = pillar_weights
                benefits = ['Ann. Return (%)', 'Alpha (ann. %)', 'R²', 'Sharpe Ratio', 'Treynor Ratio', 'Sortino Ratio', 'Information Ratio', 'Upside Capture (%)', 'Calmar Ratio', 'ESG Globe Rating', 'Max Drawdown (%)']
                costs = ['Ann. Volatility (%)', 'Beta', 'Tracking Error (%)', 'Max DD Duration (mths)', 'Downside Capture (%)', 'OCF', 'Carbon Risk Score']
                # Build metric-level weights respecting toggles and pillar exclusions
                _TOPSIS_PILLAR_GROUPS = {
                    "Returns":  ["Ann. Return (%)", "Alpha (ann. %)"],
                    "Risk-Adj": ["Sharpe Ratio", "Sortino Ratio", "Treynor Ratio", "Information Ratio", "R²"],
                    "Risk/DD":  ["Ann. Volatility (%)", "Beta", "Tracking Error (%)", "Max Drawdown (%)", "Max DD Duration (mths)", "Downside Capture (%)", "Upside Capture (%)", "Calmar Ratio"],
                    "Costs":    ["OCF"],
                    "ESG":      ["ESG Globe Rating", "Carbon Risk Score"],
                }
                _active_metric_weights = {}
                for _pg_pillar, _pg_metrics in _TOPSIS_PILLAR_GROUPS.items():
                    if _pg_pillar in _excluded_pillars:
                        continue
                    _active = [m for m in _pg_metrics if st.session_state.get(f"mt_{m}", True)]
                    if not _active:
                        continue
                    _pillar_w = pillar_weights.get(_pg_pillar, 0)
                    _per = _pillar_w / len(_active)
                    for _am in _active:
                        _active_metric_weights[_am] = _per
                if not _active_metric_weights:
                    st.warning("All metrics excluded. Please enable at least one metric.")
                    st.stop()
                _topsis_pw = dict(pillar_weights)
                for _excl_p in _excluded_pillars:
                    _topsis_pw[_excl_p] = 0
                _remaining_total = sum(_topsis_pw.values())
                if _remaining_total > 0:
                    _scale = 100.0 / _remaining_total
                    _topsis_pw = {k: v * _scale for k, v in _topsis_pw.items()}
                topsis_obj = TOPSIS(st.session_state.metrics_matrix, benefits, costs)
                topsis_obj.run(_topsis_pw)  # populate topsis_obj internals for methodology tab
                st.session_state.topsis_obj = topsis_obj
                # Correct TOPSIS: module's distances() misaligns axes  -  compute directly
                _mm = st.session_state.metrics_matrix.astype(float)
                _act_m = [m for m in _active_metric_weights if m in _mm.index]
                _mm_act = _mm.loc[_act_m]
                _row_l2 = np.sqrt((_mm_act ** 2).sum(axis=1)).replace(0, 1)
                _norm = _mm_act.div(_row_l2, axis=0)
                _amw_s = pd.Series({m: _active_metric_weights[m] for m in _act_m})
                _wmat = _norm.mul(_amw_s / _amw_s.sum(), axis=0)
                _Ap = pd.Series(index=_act_m, dtype=float)
                _Am = pd.Series(index=_act_m, dtype=float)
                for _m in _act_m:
                    _r = _wmat.loc[_m]
                    if _m in benefits:
                        _Ap[_m] = _r.max(); _Am[_m] = _r.min()
                    else:
                        _Ap[_m] = _r.min(); _Am[_m] = _r.max()
                _Dp = np.sqrt(_wmat.sub(_Ap, axis=0).pow(2).sum(axis=0))
                _Dm = np.sqrt(_wmat.sub(_Am, axis=0).pow(2).sum(axis=0))
                _sc = _Dm / (_Dp + _Dm)
                _sc_s = _sc.sort_values(ascending=False)
                topsis_ranking = pd.DataFrame({'Score': _sc_s, 'Rank': range(1, len(_sc_s) + 1)})
                st.session_state.topsis_ranking = topsis_ranking
                # Compute naive and borda rankings (store for Model Comparison)
                st.session_state.naive_ranking = compute_naive_ranking(st.session_state.metrics_matrix)
                st.session_state.borda_ranking = compute_borda_ranking(st.session_state.metrics_matrix, _active_metric_weights)
                colours = {name: _get_colour(name, i) for i, name in enumerate(st.session_state.fund_names)}

                # Visual bar
                returns = pillar_weights['Returns']
                risk_adj = pillar_weights['Risk-Adj']
                risk_dd = pillar_weights['Risk/DD']
                costs = pillar_weights['Costs']
                esg = pillar_weights['ESG']
                bar_html = f"""
                <div style="display:flex; height:8px; border-radius:4px; overflow:hidden; margin:1rem 0;">
                    <div style="width:{returns}%; background:#154D57;"></div>
                    <div style="width:{risk_adj}%; background:#1A6B77;"></div>
                    <div style="width:{risk_dd}%; background:#B7A08B;"></div>
                    <div style="width:{costs}%; background:#8B7668;"></div>
                    <div style="width:{esg}%; background:#A3C9CE;"></div>
                </div>
                <div style="display:flex; gap:1rem; font-size:0.75rem; color:#7A6F65; flex-wrap:wrap;">
                    <span><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#154D57;vertical-align:middle;margin-right:3px;"></span>Returns {returns}%</span><span><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#1A6B77;vertical-align:middle;margin-right:3px;"></span>Risk-Adj {risk_adj}%</span>
                    <span><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#B7A08B;vertical-align:middle;margin-right:3px;"></span>Risk/DD {risk_dd}%</span><span><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#8B7668;vertical-align:middle;margin-right:3px;"></span>Costs {costs}%</span><span><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#A3C9CE;vertical-align:middle;margin-right:3px;"></span>ESG {esg}%</span>
                </div>
                """
                st.markdown(bar_html, unsafe_allow_html=True)
                
                # Chart
                _fund_set = set(st.session_state.fund_names) | {'Benchmark'}
                topsis_ranking_display = topsis_ranking.loc[topsis_ranking.index.isin(_fund_set)].copy()
                topsis_ranking_display['Rank'] = range(1, len(topsis_ranking_display) + 1)
                fig_bar = chart_scores_bar(topsis_ranking_display, colours, PLOTLY_LAYOUT, "TOPSIS Final Scores")
                st.plotly_chart(fig_bar, use_container_width=True)

                # ── Feature 1 & 2: Naive + Borda ranking expanders ────────────────
                _MEDALS = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣"]

                def _render_rank_card(i, fund, score, rank, col, topsis_r=None, label=""):
                    """Render a single medal + progress bar card via st.markdown (inline HTML  -  no indent issues)."""
                    medal = _MEDALS[i] if i < len(_MEDALS) else f"#{rank}"
                    if topsis_r is not None and fund in topsis_r.index:
                        t_rank = int(topsis_r.loc[fund, "Rank"])
                        if t_rank == rank:
                            note = f'<div style="font-size:0.63rem;font-weight:600;color:#059669;margin-top:0.3rem;letter-spacing:0.02em;">&#10003; Agrees with TOPSIS</div>'
                        else:
                            note = f'<div style="font-size:0.63rem;font-weight:600;color:#D97706;margin-top:0.3rem;">&#8597; {label} #{rank} &nbsp;&middot;&nbsp; TOPSIS #{t_rank}</div>'
                    else:
                        note = ""
                    bar_w = min(score, 100)
                    st.markdown(
                        f'<div style="background:#FFFFFF;border-radius:14px;padding:1.1rem 1.4rem;'
                        f'box-shadow:0 1px 6px rgba(0,0,0,0.06);border-left:4px solid {col};margin-bottom:0.65rem;">'
                        f'<div style="display:flex;align-items:center;gap:0.9rem;">'
                        f'<div style="font-size:1.45rem;line-height:1;flex-shrink:0;">{medal}</div>'
                        f'<div style="flex:1;min-width:0;">'
                        f'<div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#ADADAD;margin-bottom:0.15rem;">Rank {rank}</div>'
                        f'<div style="font-size:0.97rem;font-weight:700;color:#111111;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{fund}</div>'
                        f'{note}'
                        f'</div>'
                        f'<div style="font-size:1.15rem;font-weight:700;color:{col};letter-spacing:-0.01em;flex-shrink:0;">{score:.1f}%</div>'
                        f'</div>'
                        f'<div style="margin-top:0.7rem;height:5px;background:#F0EDE8;border-radius:3px;overflow:hidden;">'
                        f'<div style="height:100%;width:{bar_w:.1f}%;background:{col};border-radius:3px;"></div>'
                        f'</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

                with st.expander("📊 Compare with Baseline Naive Ranking"):
                    st.markdown('<p style="font-size:0.8rem;color:#7A6F65;margin-bottom:1.25rem;line-height:1.55;">A simplified benchmark using only Annualised Return, Sharpe Ratio, and Max Drawdown with equal weights. No cost, ESG, or additional risk metrics are included. Useful as a sanity check against the full TOPSIS result.</p>', unsafe_allow_html=True)
                    _naive_r = st.session_state.naive_ranking
                    if _naive_r is not None and not _naive_r.empty:
                        for _ni, (_nfund, _nrow) in enumerate(_naive_r.iterrows()):
                            _render_rank_card(
                                _ni, _nfund, float(_nrow["Naive Score (%)"]), int(_nrow["Naive Rank"]),
                                colours.get(_nfund, "#154D57"),
                                topsis_r=topsis_ranking_display, label="Naive"
                            )
                    else:
                        st.info("Run analysis to compute naive ranking.")

                with st.expander("🗳️ Borda Count Ranking"):
                    st.markdown('<p style="font-size:0.8rem;color:#7A6F65;margin-bottom:1.25rem;line-height:1.55;">Each fund receives positional scores across all active metrics, weighted by pillar weights. A fund ranked 1st on a metric scores N−1 points; last scores 0. Scores are summed and normalised to 100.</p>', unsafe_allow_html=True)
                    _borda_r = st.session_state.borda_ranking
                    if _borda_r is not None and not _borda_r.empty:
                        for _bi, (_bfund, _brow) in enumerate(_borda_r.iterrows()):
                            _render_rank_card(
                                _bi, _bfund, float(_brow["Borda Score (%)"]), int(_brow["Borda Rank"]),
                                colours.get(_bfund, "#154D57")
                            )
                    else:
                        st.info("Run analysis to compute Borda ranking.")

                # ── Bootstrap Confidence Intervals ───────────────────────────
                st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
                st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.5rem;">RANKING ROBUSTNESS</p>', unsafe_allow_html=True)
                st.markdown('<p style="font-size:0.85rem;color:#5A5A5A;margin-bottom:1rem;">Bootstrap resampling tests whether the ranking holds under different samples of the same data. 200 resamples with replacement, 90% confidence intervals on TOPSIS scores.</p>', unsafe_allow_html=True)

                if st.button("🔁 Run Bootstrap Confidence Intervals (200 resamples)", key="run_bootstrap", use_container_width=True):
                    with st.spinner("Resampling 200 times. This takes about 10 to 15 seconds..."):
                        try:
                            # Rename fund_returns columns from tickers to fund names for bootstrap
                            _boot_tickers = st.session_state.tickers
                            _boot_fnames = st.session_state.fund_names
                            _boot_fr = st.session_state.fund_returns.copy()
                            _boot_fr = _boot_fr[[t for t in _boot_tickers if t in _boot_fr.columns]]
                            _boot_fr = _boot_fr.rename(columns=dict(zip(_boot_tickers, _boot_fnames)))
                            _boot_results = compute_bootstrap_ci(
                                _boot_fr,
                                st.session_state.benchmark_returns,
                                st.session_state.pillar_weights,
                                st.session_state.get("costs", {}),
                                st.session_state.get("esg_globe", {}),
                                st.session_state.get("carbon_risk", {}),
                            )
                            st.session_state.bootstrap_results = _boot_results
                        except Exception as _be:
                            st.error(f"Bootstrap error: {_be}")

                if st.session_state.get("bootstrap_results"):
                    _boot = st.session_state.bootstrap_results
                    _boot_fn = st.session_state.fund_names
                    _sorted_funds = sorted(
                        [f for f in _boot_fn if f in _boot],
                        key=lambda f: _boot[f]["mean_score"],
                        reverse=True
                    )
                    _fig_ci = go.Figure()
                    for _bf in _sorted_funds:
                        _br = _boot[_bf]
                        _bc = COLOURS.get(_bf, "#154D57")
                        _fig_ci.add_trace(go.Bar(
                            name=_bf,
                            x=[_bf],
                            y=[_br["mean_score"]],
                            error_y=dict(
                                type="data",
                                symmetric=False,
                                array=[_br["upper_ci"] - _br["mean_score"]],
                                arrayminus=[_br["mean_score"] - _br["lower_ci"]],
                                color=_bc,
                                thickness=2,
                                width=8,
                            ),
                            marker_color=_bc,
                            marker_opacity=0.85,
                            text=f"{_br['mean_score']:.1f}%",
                            textposition="outside",
                            hovertemplate=(
                                f"<b>{_bf}</b><br>"
                                f"Mean Score: {_br['mean_score']:.1f}%<br>"
                                f"90% CI: [{_br['lower_ci']:.1f}%, {_br['upper_ci']:.1f}%]<br>"
                                f"Mean Rank: {_br['mean_rank']:.1f} (±{_br['rank_std']:.2f})<extra></extra>"
                            )
                        ))
                    _ci_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "title"}
                    _fig_ci.update_layout(**_ci_layout)
                    _fig_ci.update_layout(
                        xaxis=dict(title="Fund", gridcolor="#F0EFEC"),
                        yaxis=dict(title="TOPSIS Score (%), Mean +/- 90% CI", gridcolor="#F0EFEC"),
                        height=420,
                        showlegend=False,
                    )
                    chart_card(
                        "Bootstrap Confidence Intervals",
                        "Error bars show the 90% confidence interval on each fund's TOPSIS score across 200 resamples. Overlapping intervals suggest rankings are not statistically distinguishable.",
                        _fig_ci
                    )

                    st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin:1rem 0 0.5rem 0;">RANK STABILITY</p>', unsafe_allow_html=True)
                    _stab_cols = st.columns(len(_sorted_funds))
                    for _scol, _sf in zip(_stab_cols, _sorted_funds):
                        _sr = _boot[_sf]
                        _rstd = _sr["rank_std"]
                        _is_stable = _rstd < 0.5
                        _sbg    = "#D1FAE5" if _is_stable else "#FEF3C7"
                        _sbdr   = "#059669" if _is_stable else "#D97706"
                        _slabel = "Very Stable" if _rstd < 0.3 else ("Stable" if _rstd < 0.6 else "Variable")
                        with _scol:
                            st.markdown(f"""
                            <div style="background:{_sbg};border-radius:14px;padding:1rem;
                                        border-left:4px solid {_sbdr};text-align:center;">
                                <div style="font-size:0.65rem;font-weight:700;text-transform:uppercase;
                                            letter-spacing:0.08em;color:#9B9B9B;margin-bottom:0.3rem;">
                                    {_sf}
                                </div>
                                <div style="font-size:1.2rem;font-weight:700;color:#1A1A1A;">
                                    Rank {_sr['mean_rank']:.1f}
                                </div>
                                <div style="font-size:0.72rem;color:#5A5A5A;margin-top:0.2rem;">
                                    σ = {_rstd:.2f} · {_slabel}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)

                    _top_fund    = _sorted_funds[0]
                    _second_fund = _sorted_funds[1] if len(_sorted_funds) > 1 else None
                    if _second_fund:
                        _top_lower    = _boot[_top_fund]["lower_ci"]
                        _second_upper = _boot[_second_fund]["upper_ci"]
                        if _top_lower < _second_upper:
                            st.markdown(f"""
                            <div style="background:#FEF3C7;border-left:4px solid #D97706;border-radius:0 8px 8px 0;
                                        padding:1rem 1.25rem;margin-top:1rem;">
                                <strong style="color:#92400E;">⚠️ Confidence intervals overlap between {_top_fund} and {_second_fund}</strong>
                                <p style="margin:0.25rem 0 0;color:#92400E;font-size:0.82rem;">
                                    The ranking between these two funds is not statistically robust under bootstrap resampling.
                                    Consider this when interpreting the results.
                                </p>
                            </div>
                            """, unsafe_allow_html=True)
                        else:
                            st.markdown(f"""
                            <div style="background:#D1FAE5;border-left:4px solid #059669;border-radius:0 8px 8px 0;
                                        padding:1rem 1.25rem;margin-top:1rem;">
                                <strong style="color:#065F46;">✓ Rankings are statistically robust</strong>
                                <p style="margin:0.25rem 0 0;color:#065F46;font-size:0.82rem;">
                                    Confidence intervals do not overlap between {_top_fund} and {_second_fund}.
                                    The top ranking is stable under bootstrap resampling.
                                </p>
                            </div>
                            """, unsafe_allow_html=True)

            else:
                st.warning(f"Weights sum to {total_w}%. Please adjust to exactly 100%.")
        
        with tab2:
            if st.session_state.topsis_obj is not None:
                normalised_df = st.session_state.topsis_obj.normalise()
                fig_heat = chart_topsis_heatmap(normalised_df, PLOTLY_LAYOUT)
                st.plotly_chart(fig_heat, use_container_width=True)
                with st.expander("Show distances and ranking table"):
                    weighted_df = st.session_state.topsis_obj.weight(st.session_state.pillar_weights)
                    A_plus, A_minus = st.session_state.topsis_obj.ideal_solutions(weighted_df)
                    D_plus, D_minus = st.session_state.topsis_obj.distances(weighted_df, A_plus, A_minus)
                    distances_df = pd.DataFrame({'D+': D_plus, 'D-': D_minus})
                    st.dataframe(distances_df.round(3))
                    st.dataframe(st.session_state.topsis_ranking.round(3))
            else:
                st.info("Run the analysis with valid weights to see intermediate steps.")

            with st.expander("📚 TOPSIS Methodology Walkthrough"):
                st.markdown("""
                <p style="font-size:0.85rem;color:#5A5A5A;margin-bottom:1rem;">
                    This walkthrough shows exactly how TOPSIS produced the ranking above,
                    using the actual numbers from this analysis run.
                </p>
                """, unsafe_allow_html=True)
                steps = [
                    ("Step 1: Decision Matrix",
                     "The raw metrics matrix contains {n_metrics} criteria across {n_funds} funds. "
                     "Each row is one metric, each column is one fund.",
                     "metrics_matrix"),
                    ("Step 2: Min-Max Normalisation",
                     "Each metric is scaled to [0, 1] so that higher values always mean better performance. "
                     "For cost and risk metrics where lower is better, values are inverted during normalisation.",
                     None),
                    ("Step 3: Weighted Normalised Matrix",
                     "Each normalised value is multiplied by its metric weight. "
                     "The five pillar weights (Returns {r}%, Risk-Adj {ra}%, Risk/DD {rdd}%, Costs {c}%, ESG {e}%) "
                     "determine how much each group of metrics influences the final ranking.",
                     None),
                    ("Step 4: Ideal and Anti-Ideal Solutions",
                     "The positive ideal solution (PIS) takes the best value in each column. "
                     "The negative ideal solution (NIS) takes the worst. "
                     "Each fund is then measured by its distance from both extremes.",
                     None),
                    ("Step 5: Closeness Coefficient",
                     "The final score S_i = D⁻ / (D⁺ + D⁻) where D⁺ is distance to PIS and D⁻ is distance to NIS. "
                     "A score closer to 1.0 means the fund is close to the ideal solution across all criteria.",
                     None),
                ]
                try:
                    mm = st.session_state.metrics_matrix
                    pw = st.session_state.pillar_weights
                    for i, (title, description, data_key) in enumerate(steps):
                        st.markdown(f"""
                        <div style="background:#FFFFFF;border-radius:12px;padding:1rem 1.25rem;
                                    margin-bottom:0.75rem;border-left:3px solid #154D57;
                                    box-shadow:0 1px 4px rgba(0,0,0,0.04);">
                            <div style="font-size:0.78rem;font-weight:700;color:#154D57;
                                        margin-bottom:0.35rem;">{title}</div>
                            <div style="font-size:0.82rem;color:#5A5A5A;line-height:1.55;">
                                {description.format(
                                    n_metrics=len(mm),
                                    n_funds=len(st.session_state.fund_names),
                                    r=pw.get("Returns", 40),
                                    ra=pw.get("Risk-Adj", 25),
                                    rdd=pw.get("Risk/DD", 20),
                                    c=pw.get("Costs", 10),
                                    e=pw.get("ESG", 5)
                                )}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                except Exception as _e:
                    logging.debug("TOPSIS methodology cards skipped (non-critical): %s", _e)

        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)

elif page == "Yuan & Yuan":
    if not st.session_state.analysis_run:
        st.markdown("# Yuan & Yuan Ranking")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Pairwise competition eigenvector method (Yuan & Yuan, 2023) for stable ranking convergence.
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center; padding: 4rem 2rem; background: #F5EEE7; border-radius: 20px; border: 2px dashed #D4C3B0;">
            <div style="margin-bottom: 1rem;"><svg width="48" height="48" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><circle cx="5" cy="7" r="2" stroke="#B7A08B" stroke-width="1.5"/><circle cx="19" cy="7" r="2" stroke="#B7A08B" stroke-width="1.5"/><circle cx="12" cy="19" r="2" stroke="#B7A08B" stroke-width="1.5"/><line x1="6.7" y1="8.5" x2="10.5" y2="17.2" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/><line x1="17.3" y1="8.5" x2="13.5" y2="17.2" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/><line x1="7" y1="7" x2="17" y2="7" stroke="#B7A08B" stroke-width="1.5" stroke-linecap="round"/></svg></div>
            <div style="font-size: 1rem; font-weight: 600; color: #0A0A0A; margin-bottom: 0.5rem;">No data yet</div>
            <div style="color: #7A6F65; margin-bottom: 1.5rem;">Head to Setup and click Run Analysis to begin.</div>
        </div>
        """, unsafe_allow_html=True)
        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("# Yuan & Yuan Ranking")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:1rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Pairwise competition eigenvector method (Yuan & Yuan, 2023) for stable ranking convergence.
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 0.75rem 0;">', unsafe_allow_html=True)

        st.info("This method constructs a competition matrix based on pairwise comparisons across weighted metrics, then applies power iteration to converge on an eigenvector representing the relative strengths of alternatives.")
        st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 0.75rem 0;">', unsafe_allow_html=True)

        with st.expander("📚 Yuan & Yuan Methodology Walkthrough"):
            st.markdown("""
            <p style="font-size:0.85rem;color:#5A5A5A;margin-bottom:1rem;">
                How the pairwise competition matrix and eigenvector ranking was computed
                for this analysis.
            </p>
            """, unsafe_allow_html=True)
            yuan_steps = [
                ("Step 1: Pairwise Competition Matrix",
                 "An {n}×{n} matrix C is constructed where each cell c(i,j) represents how strongly "
                 "fund i outperforms fund j across all {m} weighted metrics simultaneously. "
                 "Values range from 0 (complete dominance by j) to 100 (complete dominance by i)."),
                ("Step 2: Power Iteration",
                 "The principal eigenvector of matrix C is extracted using power iteration. "
                 "The algorithm multiplies C by a random vector repeatedly until it converges. "
                 "The resulting vector represents each fund's relative competitive strength."),
                ("Step 3: Score Normalisation",
                 "The raw eigenvector values are normalised so they sum to 1.0. "
                 "These scores form a ratio scale: a fund with score 0.60 is "
                 "twice as dominant as a fund with score 0.30 in pairwise competition."),
                ("Step 4: Final Ranking",
                 "Funds are ranked by descending eigenvector score. "
                 "Where TOPSIS measures distance from an ideal point, Yuan & Yuan measures "
                 "relative competitive dominance across all pairwise comparisons. The two methods cross-validate each other."),
            ]
            try:
                n_funds = len(st.session_state.fund_names)
                mm = st.session_state.metrics_matrix
                for title, description in yuan_steps:
                    st.markdown(f"""
                    <div style="background:#FFFFFF;border-radius:12px;padding:1rem 1.25rem;
                                margin-bottom:0.75rem;border-left:3px solid #4ECDC4;
                                box-shadow:0 1px 4px rgba(0,0,0,0.04);">
                        <div style="font-size:0.78rem;font-weight:700;color:#154D57;
                                    margin-bottom:0.35rem;">{title}</div>
                        <div style="font-size:0.82rem;color:#5A5A5A;line-height:1.55;">
                            {description.format(n=n_funds, m=len(mm))}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
            except Exception as _e:
                logging.debug("Yuan methodology cards skipped (non-critical): %s", _e)

        if 'pillar_weights' not in st.session_state:
            st.warning("Please set pillar weights in the TOPSIS page first.")
        else:
            benefits = ['Ann. Return (%)', 'Alpha (ann. %)', 'R²', 'Sharpe Ratio', 'Treynor Ratio', 'Sortino Ratio', 'Information Ratio', 'Upside Capture (%)', 'Calmar Ratio', 'ESG Globe Rating', 'Max Drawdown (%)']
            costs = ['Ann. Volatility (%)', 'Beta', 'Tracking Error (%)', 'Max DD Duration (mths)', 'Downside Capture (%)', 'OCF', 'Carbon Risk Score']
            yuan_obj = YuanYuan(st.session_state.metrics_matrix, benefits, costs)
            _yuan_excluded_pillars = {_pn for _pn in ["Returns", "Risk-Adj", "Risk/DD", "Costs", "ESG"] if st.session_state.get(f"excl_{_pn}", False)}
            _yuan_pw = dict(st.session_state.pillar_weights)
            for _excl_p in _yuan_excluded_pillars:
                _yuan_pw[_excl_p] = 0
            _yuan_remaining = sum(_yuan_pw.values())
            if _yuan_remaining > 0:
                _yuan_scale = 100.0 / _yuan_remaining
                _yuan_pw = {k: v * _yuan_scale for k, v in _yuan_pw.items()}
            yuan_ranking, c_matrix, n = yuan_obj.run(_yuan_pw)
            st.session_state.yuan_obj = yuan_obj
            st.session_state.yuan_ranking = yuan_ranking
            st.info(f"Power iteration converged at step {n}")
            st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 1.5rem 0;">', unsafe_allow_html=True)
            
            colours = {name: _get_colour(name, i) for i, name in enumerate(st.session_state.fund_names)}
            fig_c = chart_cmatrix(c_matrix, PLOTLY_LAYOUT)
            st.plotly_chart(fig_c, use_container_width=True, config={"displayModeBar": False})
            st.markdown("<div style='height: 1.25rem;'></div>", unsafe_allow_html=True)

            fig_yuan_bar = chart_scores_bar(yuan_ranking, colours, PLOTLY_LAYOUT, "Yuan & Yuan Scores")
            st.plotly_chart(fig_yuan_bar, use_container_width=True)
            
            st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 1.5rem 0;">', unsafe_allow_html=True)
            
            combined_df = pd.DataFrame({
                'TOPSIS Score': st.session_state.topsis_ranking['Score'],
                'TOPSIS Rank': st.session_state.topsis_ranking['Rank'],
                'Yuan Score': yuan_ranking['Score'],
                'Yuan Rank': yuan_ranking['Rank']
            })
            combined_df = combined_df.reindex(st.session_state.fund_names)
            st.session_state.combined_df = combined_df
            
            # Add badges
            combined_df['Fund'] = combined_df.index
            combined_df['Badge'] = combined_df['Yuan Rank'].apply(lambda x: '1st' if x == 1 else '2nd' if x == 2 else '3rd' if x == 3 else '4th')
            combined_df = combined_df[['Fund', 'Badge', 'TOPSIS Score', 'TOPSIS Rank', 'Yuan Score', 'Yuan Rank']]
            st.dataframe(combined_df.set_index('Fund').round(3))
            
            topsis_order = st.session_state.topsis_ranking.index.tolist()
            yuan_order = yuan_ranking.index.tolist()
            if topsis_order == yuan_order:
                st.success("Both methods agree on the full ranking order.")
            else:
                st.warning("~ Methods differ. See sensitivity analysis for context.")

            # ── Feature 4: Recommendation Badges ──────────────────────────
            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
            st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">RECOMMENDATIONS</p>', unsafe_allow_html=True)

            def _get_rec_winner(metric, direction):
                _df = st.session_state.metrics_matrix
                _fcols = [c for c in _df.columns if c != "Benchmark"]
                if metric not in _df.index:
                    raise KeyError(metric)
                _row = _df.loc[metric, _fcols].astype(float)
                _w = _row.idxmax() if direction == "highest" else _row.idxmin()
                return _w, float(_row[_w])

            def _get_yuan_winner():
                _yr = st.session_state.yuan_ranking
                _best = _yr[_yr["Rank"] == 1].index
                if len(_best) == 0:
                    raise KeyError("Yuan rank 1")
                return _best[0], float(_yr.loc[_best[0], "Score"])

            RECOMMENDATIONS = [
                ("🏆 Best All-Rounder",          None,                    None,       "yuan"),
                ("📈 Best Return",                "Ann. Return (%)",       "highest",  "matrix"),
                ("🛡️ Best Downside Protection",   "Downside Capture (%)",  "lowest",   "matrix"),
                ("⚡ Best Risk-Adjusted",          "Sharpe Ratio",          "highest",  "matrix"),
                ("💰 Best Value (Cost)",           "OCF",                   "lowest",   "matrix"),
                ("🌱 Best ESG",                   "ESG Globe Rating",      "highest",  "matrix"),
                ("📉 Smallest Drawdown",           "Max Drawdown (%)",      "highest",  "matrix"),
            ]
            _badge_colours = ["#0D5C63","#1A8A94","#4ECDC4","#C8A96E","#8B5CF6","#F97316","#059669"]
            for _row_start in range(0, len(RECOMMENDATIONS), 4):
                _row_items = RECOMMENDATIONS[_row_start:_row_start+4]
                _rcols = st.columns(len(_row_items))
                for _rcol, (_rlabel, _rmet, _rdir, _rsrc), _rbg in zip(_rcols, _row_items, _badge_colours[_row_start:]):
                    with _rcol:
                        try:
                            if _rsrc == "yuan":
                                _rwin, _rval = _get_yuan_winner()
                                _rval_fmt = f"{_rval:.3f}"
                            else:
                                _rwin, _rval = _get_rec_winner(_rmet, _rdir)
                                _rval_fmt = f"{_rval:.2f}"
                            _rfcol = COLOURS.get(_rwin, _rbg)
                            st.markdown(
                                f'<div style="background:#FFFFFF;border-radius:14px;padding:1.1rem 1.2rem;'
                                f'box-shadow:0 1px 6px rgba(0,0,0,0.07);border-top:3px solid {_rfcol};height:100%;">'
                                f'<div style="font-size:0.6rem;font-weight:700;text-transform:uppercase;'
                                f'letter-spacing:0.1em;color:#ADADAD;margin-bottom:0.55rem;">{_rlabel}</div>'
                                f'<div style="font-size:0.88rem;font-weight:700;color:#111111;'
                                f'margin-bottom:0.5rem;line-height:1.25;">{_rwin}</div>'
                                f'<div style="display:inline-flex;align-items:center;background:{_rfcol}18;'
                                f'color:{_rfcol};border-radius:6px;padding:0.2rem 0.55rem;'
                                f'font-size:0.75rem;font-weight:700;letter-spacing:0.02em;">{_rval_fmt}</div>'
                                f'</div>',
                                unsafe_allow_html=True
                            )
                        except Exception as _e:
                            st.warning(f"Badge unavailable: {_e}")

            # ── Feature 5: Investor-Type Summary Cards ─────────────────────
            st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)
            st.markdown('<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">FOR DIFFERENT INVESTOR TYPES</p>', unsafe_allow_html=True)

            INVESTOR_TYPES = {
                "⚡ Return Seeker": {
                    "description": "Prioritises maximum long-term growth above all else.",
                    "metric": "Ann. Return (%)", "direction": "highest",
                    "rationale_template": "{fund} leads with {value:.1f}% annualised return over the sample period.",
                },
                "🛡️ Cautious Investor": {
                    "description": "Prioritises capital preservation and downside protection.",
                    "metric": "Downside Capture (%)", "direction": "lowest",
                    "rationale_template": "{fund} absorbs only {value:.1f}% of benchmark losses in down months.",
                },
                "🌱 ESG Investor": {
                    "description": "Prioritises responsible investment and sustainability ratings.",
                    "metric": "ESG Globe Rating", "direction": "highest",
                    "rationale_template": "{fund} achieves the highest ESG Globe Rating of {value:.0f}/5.",
                },
                "💰 Cost-Conscious Investor": {
                    "description": "Prioritises minimising ongoing charges.",
                    "metric": "OCF", "direction": "lowest",
                    "rationale_template": "{fund} has the lowest ongoing charge of {value:.2f}%.",
                },
            }
            _inv_cols = st.columns(4)
            _card_accents = ["#0D5C63", "#4ECDC4", "#C8A96E", "#8B5CF6"]
            for _icol, (_itype, _icfg), _iacc in zip(_inv_cols, INVESTOR_TYPES.items(), _card_accents):
                with _icol:
                    try:
                        _iwin, _ival = _get_rec_winner(_icfg["metric"], _icfg["direction"])
                        _iration = _icfg["rationale_template"].format(fund=_iwin, value=_ival)
                        _ifc = COLOURS.get(_iwin, _iacc)
                        st.markdown(
                            f'<div style="background:#FFFFFF;border-radius:18px;overflow:hidden;'
                            f'box-shadow:0 2px 14px rgba(0,0,0,0.08);">'
                            f'<div style="background:linear-gradient(135deg,{_iacc},{_iacc}CC);padding:1.1rem 1.25rem;">'
                            f'<div style="font-size:1rem;font-weight:700;color:#FFFFFF;letter-spacing:-0.01em;">{_itype}</div>'
                            f'<div style="font-size:0.7rem;color:rgba(255,255,255,0.75);margin-top:0.25rem;line-height:1.4;">{_icfg["description"]}</div>'
                            f'</div>'
                            f'<div style="padding:1.1rem 1.25rem;">'
                            f'<div style="font-size:0.6rem;font-weight:700;text-transform:uppercase;letter-spacing:0.12em;color:#ADADAD;margin-bottom:0.4rem;">Recommended Fund</div>'
                            f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.55rem;">'
                            f'<div style="width:8px;height:8px;border-radius:50%;background:{_ifc};flex-shrink:0;"></div>'
                            f'<div style="font-size:0.97rem;font-weight:700;color:#111111;">{_iwin}</div>'
                            f'</div>'
                            f'<div style="font-size:0.74rem;color:#6A6A6A;line-height:1.55;">{_iration}</div>'
                            f'</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                    except Exception as _e:
                        st.warning(f"Investor card unavailable: {_e}")

        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)

elif page == "Sensitivity & Report":
    if not st.session_state.analysis_run:
        st.markdown("# Sensitivity Analysis & Report")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Robustness testing across weight schemes and downloadable PDF report with methodology.
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center; padding: 4rem 2rem; background: #F5EEE7; border-radius: 20px; border: 2px dashed #D4C3B0;">
            <div style="margin-bottom: 1rem;"><svg width="48" height="48" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><rect x="3" y="3" width="7" height="7" rx="1" stroke="#B7A08B" stroke-width="1.5"/><rect x="14" y="3" width="7" height="7" rx="1" stroke="#B7A08B" stroke-width="1.5"/><rect x="3" y="14" width="7" height="7" rx="1" stroke="#B7A08B" stroke-width="1.5"/><rect x="14" y="14" width="7" height="7" rx="1" stroke="#B7A08B" stroke-width="1.5"/></svg></div>
            <div style="font-size: 1rem; font-weight: 600; color: #0A0A0A; margin-bottom: 0.5rem;">No data yet</div>
            <div style="color: #7A6F65; margin-bottom: 1.5rem;">Head to Setup and click Run Analysis to begin.</div>
        </div>
        """, unsafe_allow_html=True)
        # Footer
        st.markdown("""
        <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("# Sensitivity Analysis & Report")
        st.markdown("""
        <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
            <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
                Robustness testing across weight schemes and downloadable PDF report with methodology.
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 1.5rem 0;">', unsafe_allow_html=True)
        
        # Sensitivity
        st.markdown("## Sensitivity Analysis")
        schemes = {
            "Baseline (40/25/20/10/5)": (0.40, 0.25, 0.20, 0.10, 0.05),
            "Return-Heavy (60/20/10/5/5)": (0.60, 0.20, 0.10, 0.05, 0.05),
            "Risk-Heavy (25/35/30/5/5)": (0.25, 0.35, 0.30, 0.05, 0.05),
            "Equal Weights (20/20/20/20/20)": (0.20, 0.20, 0.20, 0.20, 0.20),
        }
        results = {}
        benefits = ['Ann. Return (%)', 'Alpha (ann. %)', 'R²', 'Sharpe Ratio', 'Treynor Ratio', 'Sortino Ratio', 'Information Ratio', 'Upside Capture (%)', 'Calmar Ratio', 'ESG Globe Rating', 'Max Drawdown (%)']
        costs = ['Ann. Volatility (%)', 'Beta', 'Tracking Error (%)', 'Max DD Duration (mths)', 'Downside Capture (%)', 'OCF', 'Carbon Risk Score']
        yuan_obj = YuanYuan(st.session_state.metrics_matrix, benefits, costs)
        _sens_excluded_pillars = {_pn for _pn in ["Returns", "Risk-Adj", "Risk/DD", "Costs", "ESG"] if st.session_state.get(f"excl_{_pn}", False)}
        for scheme_name, weights in schemes.items():
            pillar_w = dict(zip(['Returns', 'Risk-Adj', 'Risk/DD', 'Costs', 'ESG'], weights))
            for _excl_p in _sens_excluded_pillars:
                pillar_w[_excl_p] = 0
            _sens_remaining = sum(pillar_w.values())
            if _sens_remaining > 0:
                _sens_scale = 1.0 / _sens_remaining
                pillar_w = {k: v * _sens_scale for k, v in pillar_w.items()}
            ranking, _, _ = yuan_obj.run(pillar_w)
            results[scheme_name] = ranking['Score']
        score_df = pd.DataFrame(results).T
        text_df = pd.DataFrame(index=score_df.index, columns=score_df.columns)
        for scheme in score_df.index:
            ranks = score_df.loc[scheme].rank(ascending=False)
            for fund in score_df.columns:
                _r = int(ranks[fund])
                _sfx = "th" if 11 <= (_r % 100) <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(_r % 10, "th")
                text_df.loc[scheme, fund] = f"{score_df.loc[scheme, fund]:.1%} ({_r}{_sfx})"
        
        fig_sens = chart_sensitivity_heatmap(score_df, text_df, PLOTLY_LAYOUT)
        st.plotly_chart(fig_sens, use_container_width=True)
        
        st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 1.5rem 0;">', unsafe_allow_html=True)
        
        # Stability
        stable_positions = []
        for pos in range(1, len(st.session_state.fund_names) + 1):
            ranks_across = [score_df.loc[scheme].rank(ascending=False).sort_values().index[pos-1] for scheme in score_df.index]
            if all(r == ranks_across[0] for r in ranks_across):
                stable_positions.append(pos)
        
        st.markdown("## Ordinal Stability")
        col1, col2, col3, col4 = st.columns(4)
        for i, pos in enumerate(range(1, 5)):
            fund_holding = score_df.loc[list(schemes.keys())[0]].rank(ascending=False).sort_values().index[pos-1] if pos <= len(st.session_state.fund_names) else "N/A"
            is_stable = pos in stable_positions
            color = "#154D57" if is_stable else "#B7A08B"
            border_color = "#A3C9CE" if is_stable else "#D4C3B0"
            with [col1, col2, col3, col4][i]:
                st.markdown(f"""
                <div style="background:#FFFFFF; border:2px solid {border_color}; border-radius:16px; padding:1.5rem; text-align:center; margin-bottom:1rem;">
                    <div style="font-size:1.25rem; font-weight:700; color:{color}; margin-bottom:0.5rem;">{pos}</div>
                    <div style="font-size:0.875rem; color:#7A6F65;">{fund_holding}</div>
                    <div style="font-size:0.75rem; color:{color}; margin-top:0.5rem;">{"Stable" if is_stable else "Unstable"}</div>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 1.5rem 0;">', unsafe_allow_html=True)
        
        combined_df = pd.DataFrame({
            'TOPSIS Score': st.session_state.topsis_ranking['Score'],
            'TOPSIS Rank': st.session_state.topsis_ranking['Rank'],
            'Yuan Score': st.session_state.yuan_ranking['Score'],
            'Yuan Rank': st.session_state.yuan_ranking['Rank']
        }, index=st.session_state.fund_names)
        with st.expander("⚠️ Limitations & Assumptions"):
            limitations = [
                ("ETF Proxies for UK Funds",
                 "UK-domiciled OEICs are not directly available on Yahoo Finance. US-listed ETFs with equivalent investment mandates are used as proxies. This introduces tracking differences, currency effects, and potential divergence from actual fund NAVs."),
                ("Sample Period Dependency",
                 "All metrics are computed over the selected sample window. Rankings may differ materially over different periods, particularly for funds with regime-dependent strategies. The sensitivity analysis partially addresses this by testing alternative weighting schemes, but not alternative time windows."),
                ("Weight Sensitivity",
                 "TOPSIS and Yuan & Yuan rankings depend on the pillar weights assigned by the user. While the sensitivity analysis shows ordinal stability across four schemes, extreme weight configurations may produce different outcomes. No single weighting scheme should be treated as definitively correct."),
                ("Single-Factor CAPM",
                 "Alpha and beta are estimated from a single-factor OLS regression against the benchmark. This ignores known risk factors (size, value, momentum, quality) that may explain part of each fund's return. True alpha may be lower than reported once multi-factor exposures are accounted for."),
                ("Normal Distribution Assumption",
                 "Several metrics (Sharpe ratio, tracking error) implicitly assume normally distributed returns. Emerging market fund returns exhibit negative skewness and excess kurtosis. The Sortino ratio and max drawdown are more appropriate risk measures for this asset class."),
                ("Static Cost and ESG Data",
                 "OCF, ESG globe rating, and carbon risk scores are entered manually and assumed constant over the sample period. In practice these change over time. Historical ESG ratings in particular are not available in this implementation."),
                ("Small Fund Universe",
                 f"This analysis covers {len(st.session_state.get('fund_names', []))} funds. With a small number of alternatives, MCDA rankings are more sensitive to the inclusion or exclusion of individual funds. Adding or removing a fund can shift the ranking of others."),
            ]
            for title, text in limitations:
                st.markdown(f"""
                <div style="background:#FFFFFF;border-radius:12px;padding:1rem 1.25rem;
                            margin-bottom:0.6rem;border-left:3px solid #D97706;
                            box-shadow:0 1px 4px rgba(0,0,0,0.04);">
                    <div style="font-size:0.78rem;font-weight:700;color:#92400E;
                                margin-bottom:0.3rem;">⚠️ {title}</div>
                    <div style="font-size:0.8rem;color:#5A5A5A;line-height:1.55;">{text}</div>
                </div>
                """, unsafe_allow_html=True)

        with st.expander("📐 Model Assumptions"):
            assumptions = [
                ("Log Returns", "Monthly returns are computed as natural log differences: r_t = ln(P_t / P_{t-1}). Log returns are time-additive and approximately normally distributed for short horizons."),
                ("Monthly Frequency", "All metrics use monthly return observations. Intra-month volatility is not captured. The sample uses month-end closing prices adjusted for dividends and splits."),
                ("Risk-Free Rate", f"A constant risk-free rate of {st.session_state.get('rf_rate', 5.0):.1f}% per annum is used throughout. This is applied to Sharpe, Sortino, Treynor, and Jensen's alpha calculations."),
                ("CAPM Regression", "Alpha and beta are estimated via OLS on EXCESS returns (fund return minus Rf, regressed on benchmark return minus Rf). Using excess returns is the academically correct specification."),
                ("Drawdown Calculation", "The wealth index W_t = exp(∑r_t) is used rather than (1+r_1)(1+r_2)... The log-return wealth index is equivalent for continuously compounded returns and avoids compounding approximation errors."),
                ("Metric Normalisation in TOPSIS", "Min-max normalisation is applied within each metric column. For cost and risk metrics where lower is better (volatility, max drawdown, beta, tracking error, downside capture, OCF, carbon risk), the direction is reversed so that higher normalised scores always represent better performance."),
                ("Equal Metric Weighting Within Pillars", "Within each pillar, all active metrics receive equal weight. The pillar-level weight sliders control how much each group contributes to the final ranking, but individual metrics within a pillar are treated as equally informative."),
            ]
            for title, text in assumptions:
                st.markdown(f"""
                <div style="background:#FFFFFF;border-radius:12px;padding:1rem 1.25rem;
                            margin-bottom:0.6rem;border-left:3px solid #154D57;
                            box-shadow:0 1px 4px rgba(0,0,0,0.04);">
                    <div style="font-size:0.78rem;font-weight:700;color:#154D57;
                                margin-bottom:0.3rem;">📐 {title}</div>
                    <div style="font-size:0.8rem;color:#5A5A5A;line-height:1.55;">{text}</div>
                </div>
                """, unsafe_allow_html=True)

        with st.expander("📑 Academic References"):
            references = [
                ("Hwang, C.L. & Yoon, K. (1981)",
                 "Multiple Attribute Decision Making: Methods and Applications.",
                 "Springer-Verlag, Berlin.",
                 "Foundational paper introducing the TOPSIS method used in this analysis."),
                ("Yuan, Y. & Yuan, J. (2023)",
                 "A novel multi-criteria decision-making method based on pairwise comparisons and eigenvectors.",
                 "Expert Systems with Applications.",
                 "Source of the eigenvector ranking method implemented alongside TOPSIS."),
                ("Jensen, M.C. (1968)",
                 "The performance of mutual funds in the period 1945-1964.",
                 "Journal of Finance, 23(2), 389-416.",
                 "Introduced Jensen's alpha as a measure of risk-adjusted fund performance."),
                ("Sharpe, W.F. (1966)",
                 "Mutual fund performance.",
                 "Journal of Business, 39(1), 119-138.",
                 "Original paper introducing the reward-to-variability ratio, now called the Sharpe ratio."),
                ("Sortino, F.A. & van der Meer, R. (1991)",
                 "Downside risk.",
                 "Journal of Portfolio Management, 17(4), 27-31.",
                 "Introduced the Sortino ratio as a downside-risk-adjusted performance measure."),
                ("Treynor, J.L. (1965)",
                 "How to rate management of investment funds.",
                 "Harvard Business Review, 43(1), 63-75.",
                 "Introduced the Treynor ratio measuring return per unit of systematic risk."),
                ("Maginn, J.L. et al. (2007)",
                 "Managing Investment Portfolios: A Dynamic Process (3rd ed.).",
                 "CFA Institute Investment Series.",
                 "Reference for capture ratios, drawdown analysis, and institutional portfolio management frameworks."),
            ]
            for authors, title, journal, note in references:
                st.markdown(f"""
                <div style="padding:0.75rem 0;border-bottom:1px solid #F0EFEC;">
                    <div style="font-size:0.78rem;font-weight:700;color:#1A1A1A;">{authors}</div>
                    <div style="font-size:0.8rem;color:#5A5A5A;font-style:italic;margin:0.15rem 0;">
                        {title}
                    </div>
                    <div style="font-size:0.75rem;color:#9B9B9B;">{journal}</div>
                    <div style="font-size:0.75rem;color:#C8A96E;margin-top:0.2rem;">→ {note}</div>
                </div>
                """, unsafe_allow_html=True)

        # ── Download Reports ────────────────────────────────────────────────────
        st.markdown('<hr style="border:none; border-top: 1px solid #E8DDD3; margin: 1.5rem 0 1rem 0;">', unsafe_allow_html=True)
        st.markdown('<p style="font-size:1rem;font-weight:700;color:#1D1D1F;margin-bottom:0.75rem;">Download Reports</p>', unsafe_allow_html=True)

        def _export_link(file_bytes, file_name, mime_type, label):
            b64 = base64.b64encode(file_bytes).decode()
            return (
                f'<a class="export-dl-btn" '
                f'href="data:{mime_type};base64,{b64}" '
                f'download="{file_name}">{label}</a>'
            )

        _pdf_slot = '<p style="color:#FF3B30;font-size:0.8rem;margin-top:0.75rem;">PDF generation failed.</p>'
        try:
            sample_period = f"{st.session_state.start_date} to {st.session_state.end_date}"
            pdf_bytes = generate_pdf(
                st.session_state.fund_names, sample_period,
                st.session_state.core_metrics_df, st.session_state.downside_df,
                st.session_state.topsis_ranking, st.session_state.yuan_ranking,
                combined_df, text_df
            )
            _pdf_slot = _export_link(pdf_bytes, "fund_analysis_report.pdf", "application/pdf", "Download PDF Report")
        except Exception as _e:
            logging.debug("PDF generation failed: %s", _e)

        _excel_slot = '<p style="color:#FF3B30;font-size:0.8rem;margin-top:0.75rem;">Excel generation failed.</p>'
        try:
            excel_bytes = generate_excel(
                fund_names      = st.session_state.fund_names,
                sample_period   = f"{st.session_state.get('start_date', '2020-01-01')} to {st.session_state.get('end_date', '2025-10-31')}",
                core_metrics_df = st.session_state.core_metrics_df,
                downside_df     = st.session_state.downside_df,
                topsis_ranking  = st.session_state.topsis_ranking,
                yuan_ranking    = st.session_state.yuan_ranking,
                metrics_matrix  = st.session_state.metrics_matrix,
                naive_ranking   = st.session_state.get("naive_ranking"),
                borda_ranking   = st.session_state.get("borda_ranking"),
            )
            _excel_fname = f"fund_analysis_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
            _excel_slot = _export_link(
                excel_bytes, _excel_fname,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Download Excel Workbook"
            )
        except Exception as _e:
            logging.debug("Excel generation failed: %s", _e)

        st.markdown(f"""
        <style>
        .export-dl-btn {{
            display: block;
            background-color: #154D57;
            color: #FFFFFF !important;
            text-align: center;
            padding: 0.55rem 1rem;
            border-radius: 8px;
            font-weight: 600;
            font-size: 0.875rem;
            text-decoration: none !important;
            margin-top: 0.75rem;
            transition: background-color 0.15s ease;
        }}
        .export-dl-btn:hover {{
            background-color: #1A6070;
            color: #FFFFFF !important;
            text-decoration: none !important;
        }}
        </style>
        <div style="display:flex;gap:1.5rem;">
            <div style="flex:1;background:#F9F7F4;border:1px solid #E8DDD3;border-radius:12px;padding:1.1rem 1.25rem;">
                <div style="font-size:0.9rem;font-weight:700;color:#1D1D1F;margin-bottom:0.35rem;">PDF Report</div>
                <div style="font-size:0.82rem;color:#5A5A5A;line-height:1.55;margin-bottom:0.5rem;">
                    Professional report with methodology, tables, rankings, and key outputs.
                </div>
                <div style="font-size:0.75rem;color:#9B9B9B;">Best for submission or sharing.</div>
                {_pdf_slot}
            </div>
            <div style="flex:1;background:#F9F7F4;border:1px solid #E8DDD3;border-radius:12px;padding:1.1rem 1.25rem;">
                <div style="font-size:0.9rem;font-weight:700;color:#1D1D1F;margin-bottom:0.35rem;">Excel Workbook</div>
                <div style="font-size:0.82rem;color:#5A5A5A;line-height:1.55;margin-bottom:0.5rem;">
                    Structured workbook with summary rankings, metrics, matrices, and model outputs.
                </div>
                <div style="font-size:0.75rem;color:#9B9B9B;">Best for checking formulas and raw outputs.</div>
                {_excel_slot}
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Footer
        st.markdown("""
        <div style="margin-top:2rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
            <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
                Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
            </p>
            <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
                Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
            </p>
        </div>
        """, unsafe_allow_html=True)

elif page == "📋 Model Comparison":
    st.markdown("# Model Comparison")
    st.markdown("""
    <div style="background:#EAF2F3; border-left:4px solid #154D57; border-radius:0 8px 8px 0; padding:1rem 1.25rem; margin-bottom:2rem;">
        <p style="margin:0; color:#154D57; font-size:0.875rem; line-height:1.6;">
            Side-by-side comparison of all four ranking methods: TOPSIS, Yuan &amp; Yuan, Borda Count, and Naive.
        </p>
    </div>
    """, unsafe_allow_html=True)

    if not st.session_state.get("analysis_run"):
        st.info("Run the analysis first (Setup page) to use this section.")
    else:
        _t_r = st.session_state.topsis_ranking
        _y_r = st.session_state.yuan_ranking
        _b_r = st.session_state.borda_ranking
        _n_r = st.session_state.naive_ranking

        if not all(x is not None and not (hasattr(x, "empty") and x.empty) for x in [_t_r, _y_r, _b_r, _n_r]):
            st.warning("Navigate to the TOPSIS page and set weights to compute all four rankings.")
        else:
            # Filter TOPSIS to fund-only rows (excludes metric names from module index)
            _t_funds = [f for f in st.session_state.fund_names if f in _t_r.index]
            _t_disp = _t_r.loc[_t_funds].copy()
            _t_disp["Rank"] = range(1, len(_t_disp) + 1)

            # Build unified comparison table
            _cmp_rows = []
            for _fund in _t_funds:
                _cmp_rows.append({
                    "Fund": _fund,
                    "TOPSIS Score": f"{float(_t_disp.loc[_fund, 'Score'])*100:.1f}%",
                    "TOPSIS Rank":  int(_t_disp.loc[_fund, "Rank"]),
                    "Yuan Score":   f"{float(_y_r.loc[_fund, 'Score'])*100:.1f}%" if _fund in _y_r.index else "N/A",
                    "Yuan Rank":    int(_y_r.loc[_fund, "Rank"]) if _fund in _y_r.index else "N/A",
                    "Borda Score":  f"{float(_b_r.loc[_fund, 'Borda Score (%)']):.1f}%" if _fund in _b_r.index else "N/A",
                    "Borda Rank":   int(_b_r.loc[_fund, "Borda Rank"]) if _fund in _b_r.index else "N/A",
                    "Naive Score":  f"{float(_n_r.loc[_fund, 'Naive Score (%)']):.1f}%" if _fund in _n_r.index else "N/A",
                    "Naive Rank":   int(_n_r.loc[_fund, "Naive Rank"]) if _fund in _n_r.index else "N/A",
                })
            _cmp_df = pd.DataFrame(_cmp_rows).set_index("Fund")
            with st.container(border=True):
                st.dataframe(_cmp_df, use_container_width=True)

            # Grouped bar chart  -  all four methods
            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
            _fig_cmp = go.Figure()
            _methods = [
                ("TOPSIS",  _t_disp, "Score",          "#0D5C63",  lambda s: s * 100),
                ("Yuan",    _y_r,    "Score",           "#4ECDC4",  lambda s: s * 100),
                ("Borda",   _b_r,    "Borda Score (%)", "#C8A96E",  lambda s: s),
                ("Naive",   _n_r,    "Naive Score (%)", "#8B5CF6",  lambda s: s),
            ]
            for _mn, _mdf, _mcol, _mclr, _mscale in _methods:
                _mfunds = [f for f in _t_funds if f in _mdf.index]
                _mscores = [float(_mscale(_mdf.loc[f, _mcol])) for f in _mfunds]
                _fig_cmp.add_trace(go.Bar(
                    name=_mn, x=_mfunds, y=_mscores,
                    marker_color=_mclr,
                    text=[f"{v:.1f}%" for v in _mscores],
                    textposition="outside",
                ))
            _fig_cmp.update_layout(
                barmode="group",
                title=dict(text="Score Comparison: All Ranking Methods", font=dict(size=16, color="#1D1D1F")),
                **{k: v for k, v in PLOTLY_LAYOUT.items() if k != "title"}
            )
            st.plotly_chart(_fig_cmp, use_container_width=True)

    st.markdown("""
    <div style="margin-top:1rem; padding-top:0.75rem; border-top:1px solid #E8DDD3; text-align:center;">
        <p style="color:#7A6F65; font-size:0.75rem; margin:0;">
            Fund Analysis Engine · Capital Markets Analysis BUSI4519 · University of Nottingham · 2025
        </p>
        <p style="color:#9B8C7E; font-size:0.75rem; margin:0.25rem 0 0 0;">
            Data sourced via Yahoo Finance. ETF proxies used for UK-domiciled funds.
        </p>
    </div>
    """, unsafe_allow_html=True)

elif page == "🗃️ Raw Data":
    if not st.session_state.get("analysis_run"):
        st.info("Run the analysis first to view raw data.")
    else:
        st.markdown("## 🗃️ Raw Data")
        st.markdown('<p style="font-size:0.85rem;color:#9B9B9B;margin-bottom:2rem;">Download the underlying data used in this analysis as CSV files.</p>', unsafe_allow_html=True)

        def _csv_link(csv_str, file_name, label):
            _b64 = base64.b64encode(csv_str.encode("utf-8")).decode()
            return (
                f'<a class="export-dl-btn" href="data:text/csv;base64,{_b64}" '
                f'download="{file_name}">{label}</a>'
            )

        tab1, tab2, tab3, tab4 = st.tabs([
            "📈 Monthly Returns",
            "📊 Core Metrics",
            "📉 Downside Metrics",
            "🧮 Full Matrix"
        ])

        with tab1:
            st.markdown('<p style="font-size:0.78rem;color:#9B9B9B;margin-bottom:0.75rem;">Monthly log returns for all funds and the benchmark.</p>', unsafe_allow_html=True)
            log_returns = st.session_state.log_returns
            st.dataframe(log_returns.round(6), use_container_width=True)
            st.markdown(_csv_link(log_returns.round(6).to_csv(), "monthly_log_returns.csv", "Download Monthly Returns CSV"), unsafe_allow_html=True)

        with tab2:
            st.markdown('<p style="font-size:0.78rem;color:#9B9B9B;margin-bottom:0.75rem;">Core risk-adjusted performance metrics for each fund.</p>', unsafe_allow_html=True)
            core_df = st.session_state.core_metrics_df
            st.dataframe(core_df.round(4), use_container_width=True)
            st.markdown(_csv_link(core_df.round(4).to_csv(), "core_metrics.csv", "Download Core Metrics CSV"), unsafe_allow_html=True)

        with tab3:
            st.markdown('<p style="font-size:0.78rem;color:#9B9B9B;margin-bottom:0.75rem;">Downside risk and capture metrics for each fund.</p>', unsafe_allow_html=True)
            down_df = st.session_state.downside_df
            st.dataframe(down_df.round(4), use_container_width=True)
            st.markdown(_csv_link(down_df.round(4).to_csv(), "downside_metrics.csv", "Download Downside Metrics CSV"), unsafe_allow_html=True)

        with tab4:
            st.markdown('<p style="font-size:0.78rem;color:#9B9B9B;margin-bottom:0.75rem;">The full 19-metric evaluation matrix used as input to TOPSIS and Yuan & Yuan.</p>', unsafe_allow_html=True)
            matrix = st.session_state.metrics_matrix
            st.dataframe(matrix.round(4), use_container_width=True)
            st.markdown(_csv_link(matrix.round(4).to_csv(), "metrics_matrix.csv", "Download Full Matrix CSV"), unsafe_allow_html=True)

elif page == "⚖️ Comparison":
    st.markdown("## ⚖️ Head-to-Head Fund Comparison")
    if not st.session_state.get("analysis_run"):
        st.info("Run the analysis first (Setup page) to use the comparison tool.")
    else:
        _fn = st.session_state.fund_names
        _cc1, _cc2 = st.columns(2)
        with _cc1:
            _fund_a = st.selectbox("Fund A", options=_fn, index=0, key="cmp_a")
        with _cc2:
            _fund_b = st.selectbox("Fund B", options=_fn, index=min(1, len(_fn)-1), key="cmp_b")

        if _fund_a == _fund_b:
            st.markdown('<div style="background:#FEF3C7;border-radius:14px;padding:1rem 1.25rem;border-left:4px solid #D97706;"><span style="font-weight:600;color:#92400E;">Please select two different funds to compare.</span></div>', unsafe_allow_html=True)
        else:
            _core = st.session_state.core_metrics_df
            _down = st.session_state.downside_df

            # Build combined metrics  -  metrics as rows, fund columns
            if _fund_a in _core.index:
                _comb = pd.DataFrame({
                    _fund_a: pd.concat([_core.loc[_fund_a], _down.loc[_fund_a]]),
                    _fund_b: pd.concat([_core.loc[_fund_b], _down.loc[_fund_b]])
                })
            else:
                _comb = pd.DataFrame({
                    _fund_a: pd.concat([_core[_fund_a], _down[_fund_a]]),
                    _fund_b: pd.concat([_core[_fund_b], _down[_fund_b]])
                })

            _HIB = ["Ann. Return (%)", "Alpha (ann. %)", "Sharpe Ratio", "Treynor Ratio",
                    "Sortino Ratio", "Information Ratio", "R²", "Upside Capture (%)", "Calmar Ratio",
                    "ESG Globe Rating"]
            _LIB = ["Ann. Volatility (%)", "Tracking Error (%)", "Beta",
                    "Downside Capture (%)", "Max Drawdown (%)", "Max DD Duration (mths)",
                    "OCF", "Carbon Risk Score"]

            _winners = []
            for _met in _comb.index:
                try:
                    _va = float(_comb.loc[_met, _fund_a])
                    _vb = float(_comb.loc[_met, _fund_b])
                    if pd.isna(_va) or pd.isna(_vb):
                        _winners.append("N/A")
                    elif _met in _HIB:
                        _winners.append(_fund_a if _va > _vb else _fund_b)
                    elif _met in _LIB:
                        _winners.append(_fund_a if _va < _vb else _fund_b)
                    else:
                        _winners.append("N/A")
                except Exception as _e:
                    _winners.append("N/A")

            _wins_a = _winners.count(_fund_a)
            _wins_b = _winners.count(_fund_b)
            _total  = _wins_a + _wins_b

            _sc1, _sc2, _sc3 = st.columns(3)
            with _sc1:
                st.markdown(f'<div style="background:linear-gradient(135deg,#154D57,#1A6B77);border-radius:16px;padding:1.25rem;text-align:center;"><div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:rgba(255,255,255,0.7);">{_fund_a}</div><div style="font-size:2rem;font-weight:700;color:#FFFFFF;margin:0.25rem 0;">{_wins_a}</div><div style="font-size:0.78rem;color:rgba(255,255,255,0.7);">metrics won</div></div>', unsafe_allow_html=True)
            with _sc2:
                st.markdown(f'<div style="background:#F5F0EB;border-radius:16px;padding:1.25rem;text-align:center;"><div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#9B8C7E;">VS</div><div style="font-size:1.1rem;font-weight:700;color:#0A0A0A;margin:0.5rem 0;">{_total} metrics<br>compared</div></div>', unsafe_allow_html=True)
            with _sc3:
                st.markdown(f'<div style="background:linear-gradient(135deg,#B7A08B,#8B7668);border-radius:16px;padding:1.25rem;text-align:center;"><div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:rgba(255,255,255,0.7);">{_fund_b}</div><div style="font-size:2rem;font-weight:700;color:#FFFFFF;margin:0.25rem 0;">{_wins_b}</div><div style="font-size:0.78rem;color:rgba(255,255,255,0.7);">metrics won</div></div>', unsafe_allow_html=True)

            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)

            # Comparison table
            _disp = _comb.copy().round(4)
            _disp["Winner"] = _winners

            def _hl(row):
                s = [""] * len(row)
                if row["Winner"] == _fund_a:
                    s[0] = "background-color:#D1FAE5;color:#065F46;font-weight:600;"
                elif row["Winner"] == _fund_b:
                    s[1] = "background-color:#D1FAE5;color:#065F46;font-weight:600;"
                return s
            st.dataframe(_disp.style.apply(_hl, axis=1), use_container_width=True)

# ── 📅 Rolling Rankings ───────────────────────────────────────────────────────
elif page == "📅 Rolling Rankings":
    if not st.session_state.analysis_run:
        st.markdown("# Rolling Rankings")
        st.markdown("""
<div style="text-align:center; padding: 4rem 2rem; background: #F5EEE7; border-radius: 20px; border: 2px dashed #D4C3B0;">
    <div style="font-size: 1rem; font-weight: 600; color: #0A0A0A; margin-bottom: 0.5rem;">No data yet</div>
    <div style="color: #7A6F65; margin-bottom: 1.5rem;">Head to Setup and click Run Analysis to begin.</div>
</div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("# Rolling Rankings")
        st.markdown(
            '<div style="background:#EAF2F3;border-left:4px solid #154D57;border-radius:0 8px 8px 0;padding:1rem 1.25rem;margin-bottom:2rem;">'
            '<p style="margin:0;color:#154D57;font-size:0.875rem;line-height:1.6;">'
            'Shows how TOPSIS rankings evolve across rolling windows. '
            'Compare specific market periods, stress-test against different regimes, and examine bear market performance.'
            '</p></div>',
            unsafe_allow_html=True
        )

        _rr_pw = st.session_state.pillar_weights or {'Returns': 40, 'Risk-Adj': 25, 'Risk/DD': 20, 'Costs': 10, 'ESG': 5}
        _rr_fn = st.session_state.fund_names
        _rr_tk = st.session_state.tickers
        _rr_fr = st.session_state.fund_returns
        _rr_br = st.session_state.benchmark_returns
        _rr_rf = st.session_state.rf_annual
        _rr_costs = st.session_state.costs
        _rr_esg = st.session_state.esg_globe
        _rr_carbon = st.session_state.carbon_risk
        _rr_colours = {name: COLOURS.get(name, '#0071E3') for name in _rr_fn}

        _rr_tab1, _rr_tab2, _rr_tab3, _rr_tab4 = st.tabs(
            ["📊 Rolling Rankings", "⚖️ Period Comparison", "🔬 Stress Test", "🐻 Bear Market Filter"]
        )

        # ── Tab 1: Rolling Window Rankings ──────────────────────────────────
        with _rr_tab1:
            st.markdown(
                '<p style="color:#7A6F65;font-size:0.875rem;line-height:1.6;margin-bottom:1rem;">'
                'TOPSIS is re-run on each 36-month rolling window (stepped every 6 months) using the current '
                'pillar weights. The bump chart shows how fund rankings shift over time; the score chart '
                'shows the underlying TOPSIS score trajectory.</p>',
                unsafe_allow_html=True
            )

            if st.session_state.rolling_rankings is None:
                if st.button("▶ Run Rolling TOPSIS (36-month window, 6-month step)", key="run_rolling", use_container_width=True):
                    with st.spinner("Computing rolling TOPSIS across all windows…"):
                        _rr_result = compute_rolling_rankings(
                            _rr_fr, _rr_br, _rr_fn, _rr_tk,
                            _rr_costs, _rr_esg, _rr_carbon,
                            _rr_rf, _rr_pw
                        )
                        st.session_state.rolling_rankings = _rr_result
                    st.rerun()
            else:
                _rr_col_info, _rr_col_btn = st.columns([4, 1])
                with _rr_col_info:
                    st.markdown(
                        f'<p style="font-size:0.78rem;color:#154D57;font-weight:600;margin:0;">'
                        f'&#10003; {len(st.session_state.rolling_rankings)} windows computed '
                        f'· weights: Returns {_rr_pw.get("Returns",0)}% / Risk-Adj {_rr_pw.get("Risk-Adj",0)}% '
                        f'/ Risk/DD {_rr_pw.get("Risk/DD",0)}% / Costs {_rr_pw.get("Costs",0)}% '
                        f'/ ESG {_rr_pw.get("ESG",0)}%</p>',
                        unsafe_allow_html=True
                    )
                with _rr_col_btn:
                    if st.button("↺ Recompute", key="rerun_rolling", use_container_width=True):
                        st.session_state.rolling_rankings = None
                        st.rerun()

            _rolling = st.session_state.rolling_rankings
            if _rolling:
                _roll_dates = sorted(_rolling.keys())

                # Bump chart
                _fig_bump = go.Figure()
                for _rn in _rr_fn:
                    _bx = [d for d in _roll_dates if _rn in _rolling[d].index]
                    _by = [int(_rolling[d].loc[_rn, 'Rank']) for d in _bx]
                    _fig_bump.add_trace(go.Scatter(
                        x=_bx, y=_by,
                        mode='lines+markers',
                        name=_rn,
                        line=dict(color=_rr_colours.get(_rn, '#0071E3'), width=2.5),
                        marker=dict(size=9, color=_rr_colours.get(_rn, '#0071E3'),
                                    line=dict(color='#FFFFFF', width=1.5))
                    ))
                _fig_bump.update_layout(**{k: v for k, v in PLOTLY_LAYOUT.items() if k != 'title'})
                _fig_bump.update_layout(title=dict(text="Fund Rank Over Time (36-Month Rolling Windows)", font=dict(size=16, color="#1D1D1F")))
                _fig_bump.update_yaxes(
                    autorange='reversed',
                    tickvals=list(range(1, len(_rr_fn) + 1)),
                    ticktext=[f"#{i}" for i in range(1, len(_rr_fn) + 1)],
                    title="Rank (lower = better)"
                )
                _fig_bump.update_xaxes(title="Window End Date")
                st.plotly_chart(_fig_bump, use_container_width=True)

                # Score line chart
                _fig_score = go.Figure()
                for _rn in _rr_fn:
                    _sx = [d for d in _roll_dates if _rn in _rolling[d].index]
                    _sy = [float(_rolling[d].loc[_rn, 'TOPSIS Score (%)']) for d in _sx]
                    _fig_score.add_trace(go.Scatter(
                        x=_sx, y=_sy,
                        mode='lines',
                        name=_rn,
                        line=dict(color=_rr_colours.get(_rn, '#0071E3'), width=2)
                    ))
                _fig_score.update_layout(**{k: v for k, v in PLOTLY_LAYOUT.items() if k != 'title'})
                _fig_score.update_layout(title=dict(text="TOPSIS Score (%) Over Time", font=dict(size=16, color="#1D1D1F")))
                _fig_score.update_yaxes(title="TOPSIS Score (%)", tickformat=".1f")
                _fig_score.update_xaxes(title="Window End Date")
                st.plotly_chart(_fig_score, use_container_width=True)

        # ── Tab 2: Period Comparison ─────────────────────────────────────────
        with _rr_tab2:
            st.markdown(
                '<p style="color:#7A6F65;font-size:0.875rem;line-height:1.6;margin-bottom:1.25rem;">'
                'Select two sub-periods and compare how the TOPSIS ranking differs between them. '
                'Useful for assessing consistency of outperformance across market regimes.</p>',
                unsafe_allow_html=True
            )

            _dr_start = _rr_fr.index.min().to_pydatetime()
            _dr_end = _rr_fr.index.max().to_pydatetime()
            _dr_mid = _rr_fr.index[len(_rr_fr) // 2].to_pydatetime()

            _pca, _pcb = st.columns(2)
            with _pca:
                st.markdown(
                    '<div style="background:linear-gradient(135deg,#154D57,#1A6B77);border-radius:10px 10px 0 0;'
                    'padding:0.6rem 1rem;">'
                    '<span style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;'
                    'color:rgba(255,255,255,0.85);">Period A</span></div>',
                    unsafe_allow_html=True
                )
                _pa_s = st.date_input("Period A start", value=_dr_start, key="pa_start",
                                      min_value=_dr_start, max_value=_dr_end,
                                      label_visibility="collapsed")
                _pa_e = st.date_input("Period A end", value=_dr_mid, key="pa_end",
                                      min_value=_dr_start, max_value=_dr_end,
                                      label_visibility="collapsed")
            with _pcb:
                st.markdown(
                    '<div style="background:linear-gradient(135deg,#B7A08B,#8B7668);border-radius:10px 10px 0 0;'
                    'padding:0.6rem 1rem;">'
                    '<span style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;'
                    'color:rgba(255,255,255,0.85);">Period B</span></div>',
                    unsafe_allow_html=True
                )
                _pb_s = st.date_input("Period B start", value=_dr_mid, key="pb_start",
                                      min_value=_dr_start, max_value=_dr_end,
                                      label_visibility="collapsed")
                _pb_e = st.date_input("Period B end", value=_dr_end, key="pb_end",
                                      min_value=_dr_start, max_value=_dr_end,
                                      label_visibility="collapsed")

            st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
            if st.button("⚖️ Compare Periods", key="compare_periods", use_container_width=True):
                with st.spinner("Running TOPSIS on both periods…"):
                    _ra = run_period_topsis(_rr_fr, _rr_br, _rr_fn, _rr_tk,
                                            _rr_costs, _rr_esg, _rr_carbon,
                                            _rr_rf, _rr_pw, _pa_s, _pa_e)
                    _rb = run_period_topsis(_rr_fr, _rr_br, _rr_fn, _rr_tk,
                                            _rr_costs, _rr_esg, _rr_carbon,
                                            _rr_rf, _rr_pw, _pb_s, _pb_e)
                    st.session_state.period_a_ranking = _ra
                    st.session_state.period_b_ranking = _rb
                    st.session_state.period_a_label = f"{_pa_s} → {_pa_e}"
                    st.session_state.period_b_label = f"{_pb_s} → {_pb_e}"
                st.rerun()

            _p_ra = st.session_state.period_a_ranking
            _p_rb = st.session_state.period_b_ranking
            if _p_ra is not None and not _p_ra.empty:
                _pla = st.session_state.period_a_label
                _plb = st.session_state.period_b_label
                st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)

                _col_a, _col_b = st.columns(2)
                with _col_a:
                    st.markdown(
                        f'<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                        f'letter-spacing:0.1em;color:#154D57;margin-bottom:0.75rem;">Period A · {_pla}</p>',
                        unsafe_allow_html=True
                    )
                    for _pf in _p_ra.index:
                        _ps = float(_p_ra.loc[_pf, 'TOPSIS Score (%)'])
                        _pr = int(_p_ra.loc[_pf, 'Rank'])
                        _pc = _rr_colours.get(_pf, '#0071E3')
                        st.markdown(
                            f'<div style="background:#FFFFFF;border-radius:12px;padding:0.9rem 1.2rem;'
                            f'box-shadow:0 1px 5px rgba(0,0,0,0.06);border-left:4px solid {_pc};margin-bottom:0.5rem;">'
                            f'<div style="display:flex;align-items:center;justify-content:space-between;">'
                            f'<div><div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;'
                            f'letter-spacing:0.1em;color:#ADADAD;">#{_pr}</div>'
                            f'<div style="font-size:0.9rem;font-weight:700;color:#111111;">{_pf}</div></div>'
                            f'<div style="font-size:1.1rem;font-weight:700;color:{_pc};">{_ps:.1f}%</div>'
                            f'</div></div>',
                            unsafe_allow_html=True
                        )
                with _col_b:
                    st.markdown(
                        f'<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                        f'letter-spacing:0.1em;color:#8B7668;margin-bottom:0.75rem;">Period B · {_plb}</p>',
                        unsafe_allow_html=True
                    )
                    if _p_rb is not None and not _p_rb.empty:
                        for _pf in _p_rb.index:
                            _ps = float(_p_rb.loc[_pf, 'TOPSIS Score (%)'])
                            _pr = int(_p_rb.loc[_pf, 'Rank'])
                            _pc = _rr_colours.get(_pf, '#0071E3')
                            st.markdown(
                                f'<div style="background:#FFFFFF;border-radius:12px;padding:0.9rem 1.2rem;'
                                f'box-shadow:0 1px 5px rgba(0,0,0,0.06);border-left:4px solid {_pc};margin-bottom:0.5rem;">'
                                f'<div style="display:flex;align-items:center;justify-content:space-between;">'
                                f'<div><div style="font-size:0.62rem;font-weight:700;text-transform:uppercase;'
                                f'letter-spacing:0.1em;color:#ADADAD;">#{_pr}</div>'
                                f'<div style="font-size:0.9rem;font-weight:700;color:#111111;">{_pf}</div></div>'
                                f'<div style="font-size:1.1rem;font-weight:700;color:{_pc};">{_ps:.1f}%</div>'
                                f'</div></div>',
                                unsafe_allow_html=True
                            )
                    else:
                        st.info("Insufficient data for Period B.")

                if _p_rb is not None and not _p_rb.empty:
                    st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
                    st.markdown(
                        '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                        'letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.5rem;">Rank Change A → B</p>',
                        unsafe_allow_html=True
                    )
                    _chg_rows = []
                    for _pf in _rr_fn:
                        if _pf in _p_ra.index and _pf in _p_rb.index:
                            _ra_r = int(_p_ra.loc[_pf, 'Rank'])
                            _rb_r = int(_p_rb.loc[_pf, 'Rank'])
                            _delta = _ra_r - _rb_r
                            _arrow = "▲" if _delta > 0 else ("▼" if _delta < 0 else "")
                            _chg_rows.append({
                                'Fund': _pf,
                                f'Rank A': _ra_r,
                                f'Rank B': _rb_r,
                                'Change (A to B)': f"{_arrow} {abs(_delta)}" if _delta != 0 else "no change"
                            })
                    if _chg_rows:
                        st.dataframe(pd.DataFrame(_chg_rows).set_index('Fund'), use_container_width=True)

        # ── Tab 3: Stress Test Mode ──────────────────────────────────────────
        with _rr_tab3:
            st.markdown(
                '<p style="color:#7A6F65;font-size:0.875rem;line-height:1.6;margin-bottom:1.25rem;">'
                'Re-run TOPSIS under specific market regimes to test ranking robustness. '
                'Compare stress-period rankings against the full-sample baseline.</p>',
                unsafe_allow_html=True
            )

            _st_presets = {
                "🦠 Exclude COVID": {"ranges": [("2020-02-01", "2020-11-30")], "exclude": True,
                                     "desc": "Full sample excluding COVID crash (Feb–Nov 2020)"},
                "📉 EM Bear Market Only": {"ranges": [("2021-02-01", "2022-10-31")], "exclude": False,
                                           "desc": "EM bear market window only (Feb 2021–Oct 2022)"},
                "📈 Post-Recovery Only": {"ranges": [("2023-01-01", "2025-10-31")], "exclude": False,
                                          "desc": "Post-recovery period only (Jan 2023–Oct 2025)"},
            }

            _stc1, _stc2, _stc3, _stc4 = st.columns(4)
            for _st_col, (_st_name, _st_cfg) in zip([_stc1, _stc2, _stc3], _st_presets.items()):
                with _st_col:
                    if st.button(_st_name, key=f"stress_{_st_name}", use_container_width=True):
                        with st.spinner(f"Running {_st_name}…"):
                            _st_result = run_topsis_on_filtered_returns(
                                _rr_fr, _rr_br, _rr_fn, _rr_tk,
                                _rr_costs, _rr_esg, _rr_carbon,
                                _rr_rf, _rr_pw,
                                _st_cfg["ranges"], _st_cfg["exclude"]
                            )
                            st.session_state.stress_ranking = _st_result
                            st.session_state.active_stress_test = _st_name
                            st.session_state.stress_description = _st_cfg["desc"]
                        st.rerun()
            with _stc4:
                if st.button("✕ Clear", key="clear_stress", use_container_width=True):
                    st.session_state.stress_ranking = None
                    st.session_state.active_stress_test = None
                    st.session_state.stress_description = ""
                    st.rerun()

            _st_r = st.session_state.stress_ranking
            if _st_r is not None and not _st_r.empty:
                st.markdown('<div style="height:0.75rem;"></div>', unsafe_allow_html=True)
                st.markdown(
                    f'<div style="background:#EAF2F3;border-left:4px solid #154D57;border-radius:0 8px 8px 0;'
                    f'padding:0.65rem 1rem;margin-bottom:1rem;">'
                    f'<span style="font-size:0.78rem;font-weight:600;color:#154D57;">'
                    f'{st.session_state.active_stress_test}: {st.session_state.stress_description}</span></div>',
                    unsafe_allow_html=True
                )

                _st_col_r, _st_col_b = st.columns(2)
                with _st_col_r:
                    st.markdown(
                        '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                        'letter-spacing:0.1em;color:#154D57;margin-bottom:0.75rem;">Stress Test Ranking</p>',
                        unsafe_allow_html=True
                    )
                    for _sf in _st_r.index:
                        _ss = float(_st_r.loc[_sf, 'TOPSIS Score (%)'])
                        _sr = int(_st_r.loc[_sf, 'Rank'])
                        _sc_col = _rr_colours.get(_sf, '#0071E3')
                        _medals = ["🥇", "🥈", "🥉", "4️⃣"]
                        _medal = _medals[_sr - 1] if _sr <= len(_medals) else f"#{_sr}"
                        st.markdown(
                            f'<div style="background:#FFFFFF;border-radius:12px;padding:0.9rem 1.2rem;'
                            f'box-shadow:0 1px 5px rgba(0,0,0,0.06);border-left:4px solid {_sc_col};margin-bottom:0.5rem;">'
                            f'<div style="display:flex;align-items:center;gap:0.75rem;">'
                            f'<div style="font-size:1.35rem;">{_medal}</div>'
                            f'<div style="flex:1;"><div style="font-size:0.9rem;font-weight:700;color:#111111;">{_sf}</div></div>'
                            f'<div style="font-size:1.05rem;font-weight:700;color:{_sc_col};">{_ss:.1f}%</div>'
                            f'</div></div>',
                            unsafe_allow_html=True
                        )

                with _st_col_b:
                    _base_r = st.session_state.topsis_ranking
                    if _base_r is not None:
                        st.markdown(
                            '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                            'letter-spacing:0.1em;color:#9B8C7E;margin-bottom:0.75rem;">Full-Sample Baseline</p>',
                            unsafe_allow_html=True
                        )
                        _base_fn = [f for f in _rr_fn if f in _base_r.index]
                        _base_sorted = _base_r.loc[_base_fn].sort_values('Rank')
                        for _bf in _base_sorted.index:
                            _bs = float(_base_sorted.loc[_bf, 'Score']) * 100
                            _br2 = int(_base_sorted.loc[_bf, 'Rank'])
                            _bc = _rr_colours.get(_bf, '#0071E3')
                            _st_rank = int(_st_r.loc[_bf, 'Rank']) if _bf in _st_r.index else None
                            if _st_rank and _st_rank != _br2:
                                _chg = _br2 - _st_rank
                                _note = f'<span style="font-size:0.62rem;color:#D97706;font-weight:600;">&#8597; stress #{_st_rank}</span>'
                            elif _st_rank == _br2:
                                _note = f'<span style="font-size:0.62rem;color:#059669;font-weight:600;">&#10003; same rank</span>'
                            else:
                                _note = ""
                            st.markdown(
                                f'<div style="background:#FAFAFA;border-radius:12px;padding:0.9rem 1.2rem;'
                                f'box-shadow:0 1px 3px rgba(0,0,0,0.04);border-left:4px solid {_bc};margin-bottom:0.5rem;">'
                                f'<div style="display:flex;align-items:center;justify-content:space-between;">'
                                f'<div><div style="font-size:0.9rem;font-weight:700;color:#111111;">{_bf}</div>'
                                f'{_note}</div>'
                                f'<div style="text-align:right;"><div style="font-size:1.05rem;font-weight:700;color:{_bc};">{_bs:.1f}%</div>'
                                f'<div style="font-size:0.62rem;color:#ADADAD;">rank #{_br2}</div></div>'
                                f'</div></div>',
                                unsafe_allow_html=True
                            )
                    else:
                        st.info("Run TOPSIS on the TOPSIS page first to see baseline comparison.")

        # ── Tab 4: Bear Market Filter ────────────────────────────────────────
        with _rr_tab4:
            st.markdown(
                '<p style="color:#7A6F65;font-size:0.875rem;line-height:1.6;margin-bottom:1.25rem;">'
                'Fixed window: <strong>Feb 2021 – Oct 2022</strong> (EM bear market). '
                'Four key metrics show each fund\'s resilience during peak-to-trough drawdown conditions.</p>',
                unsafe_allow_html=True
            )

            _bm_mask = ((_rr_fr.index >= pd.Timestamp("2021-02-01")) &
                        (_rr_fr.index <= pd.Timestamp("2022-10-31")))
            _bm_fund = _rr_fr.loc[_bm_mask]
            _bm_bench = _rr_br.loc[_bm_mask]

            if len(_bm_fund) < 3:
                st.warning("Insufficient data in the 2021-02 to 2022-10 window.")
            else:
                _bm_data = {}
                for _bi, _bname in enumerate(_rr_fn):
                    _bticker = _rr_tk[_bi]
                    if _bticker not in _bm_fund.columns:
                        continue
                    _bret = _bm_fund[_bticker].dropna()
                    if len(_bret) < 3:
                        continue
                    _bcumul = (np.exp(_bret.sum()) - 1) * 100
                    _bvol = _bret.std() * np.sqrt(12) * 100
                    _bwealth = np.exp(_bret.cumsum())
                    _bcummax = _bwealth.cummax()
                    _bmaxdd = ((_bwealth - _bcummax) / _bcummax).min() * 100
                    _bdown_m = _bm_bench < 0
                    if _bdown_m.sum() > 0:
                        _bdc = (_bret[_bdown_m].mean() / _bm_bench[_bdown_m].mean()) * 100
                    else:
                        _bdc = np.nan
                    _bm_data[_bname] = {
                        'Cumul. Return (%)': round(_bcumul, 2),
                        'Max Drawdown (%)': round(_bmaxdd, 2),
                        'Downside Capture (%)': round(_bdc, 2) if not np.isnan(_bdc) else np.nan,
                        'Ann. Volatility (%)': round(_bvol, 2),
                    }

                if _bm_data:
                    _bm_df = pd.DataFrame(_bm_data).T
                    _bm_metrics = ['Cumul. Return (%)', 'Max Drawdown (%)', 'Downside Capture (%)', 'Ann. Volatility (%)']

                    st.markdown(
                        '<div style="background:linear-gradient(135deg,#991B1B,#DC2626);border-radius:10px 10px 0 0;'
                        'padding:0.65rem 1rem;">'
                        '<span style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;'
                        'color:rgba(255,255,255,0.9);">Bear Market Performance · Feb 2021 – Oct 2022</span></div>',
                        unsafe_allow_html=True
                    )

                    def _bm_style(row):
                        metric = row.name
                        higher_better = metric in ['Cumul. Return (%)']
                        best = row.max() if higher_better else row.min()
                        return [
                            'background-color:#D1FAE5;color:#065F46;font-weight:700;' if v == best else ''
                            for v in row
                        ]

                    st.dataframe(
                        _bm_df[_bm_metrics].style.apply(_bm_style, axis=1),
                        use_container_width=True
                    )

                    st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)

                    # Grouped bar chart
                    _fig_bear = go.Figure()
                    for _bname in _rr_fn:
                        if _bname in _bm_df.index:
                            _fig_bear.add_trace(go.Bar(
                                name=_bname,
                                x=_bm_metrics,
                                y=[_bm_df.loc[_bname, m] for m in _bm_metrics],
                                marker_color=_rr_colours.get(_bname, '#0071E3'),
                                text=[f"{_bm_df.loc[_bname, m]:.1f}" for m in _bm_metrics],
                                textposition='auto'
                            ))
                    _fig_bear.update_layout(**{k: v for k, v in PLOTLY_LAYOUT.items() if k != 'title'})
                    _fig_bear.update_layout(
                        barmode='group',
                        title=dict(text="Bear Market Metrics by Fund (Feb 2021 – Oct 2022)", font=dict(size=16, color="#1D1D1F"))
                    )
                    st.plotly_chart(_fig_bear, use_container_width=True)

                    # Winner card
                    _dc_col = 'Downside Capture (%)'
                    if _dc_col in _bm_df.columns and not _bm_df[_dc_col].isna().all():
                        _bm_winner = _bm_df[_dc_col].idxmin()
                        _bm_winner_dc = _bm_df.loc[_bm_winner, _dc_col]
                        _bm_winner_ret = _bm_df.loc[_bm_winner, 'Cumul. Return (%)']
                        _bm_winner_dd = _bm_df.loc[_bm_winner, 'Max Drawdown (%)']
                        _bm_wc = _rr_colours.get(_bm_winner, '#DC2626')
                        st.markdown(
                            f'<div style="background:linear-gradient(135deg,#991B1B,#DC2626);border-radius:16px;'
                            f'padding:1.5rem 1.75rem;margin-top:0.5rem;">'
                            f'<div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                            f'letter-spacing:0.1em;color:rgba(255,255,255,0.65);margin-bottom:0.4rem;">'
                            f'Best Downside Protection · Feb 2021 – Oct 2022</div>'
                            f'<div style="display:flex;align-items:center;gap:1rem;">'
                            f'<div style="font-size:2rem;">🛡️</div>'
                            f'<div><div style="font-size:1.35rem;font-weight:700;color:#FFFFFF;">{_bm_winner}</div>'
                            f'<div style="font-size:0.78rem;color:rgba(255,255,255,0.75);margin-top:0.2rem;">'
                            f'Downside Capture {_bm_winner_dc:.1f}% &nbsp;·&nbsp; '
                            f'Cumul. Return {_bm_winner_ret:.1f}% &nbsp;·&nbsp; '
                            f'Max DD {_bm_winner_dd:.1f}%</div></div>'
                            f'</div></div>',
                            unsafe_allow_html=True
                        )
                else:
                    st.warning("Could not compute bear market metrics. Check that fund data covers Feb 2021 – Oct 2022.")

# ── 🎨 Visualisations ─────────────────────────────────────────────────────────
elif page == "🎨 Visualisations":
    section_header("🎨", "CHARTS", "Advanced Visualisations", "visualisations")
    st.markdown(
        '<div style="background:#EFF6FF;border-left:4px solid #0D5C63;border-radius:0 8px 8px 0;'
        'padding:1rem 1.25rem;margin-bottom:2rem;">'
        '<p style="margin:0;color:#1E40AF;font-size:0.85rem;line-height:1.6;">'
        'Seven charts covering fund behaviour, risk distribution, '
        'return patterns, and correlation structure beyond the summary metrics.'
        '</p></div>',
        unsafe_allow_html=True
    )

    if st.session_state.get("analysis_run", False):
        _vl_lr = st.session_state.log_returns
        _vl_fr = st.session_state.fund_returns
        _vl_br = st.session_state.benchmark_returns
        _vl_mm = st.session_state.metrics_matrix  # metrics as rows, funds+benchmark as cols

        # Rename fund_returns columns from tickers to fund names
        _vl_fr_named = _vl_fr[[t for t in st.session_state.tickers if t in _vl_fr.columns]].copy()
        _vl_fr_named = _vl_fr_named.rename(columns=dict(zip(st.session_state.tickers, st.session_state.fund_names)))

        # Rename log_returns columns to fund names + benchmark name
        _vl_rename = dict(zip(st.session_state.tickers, st.session_state.fund_names))
        _vl_rename[st.session_state.benchmark_ticker] = st.session_state.benchmark_name
        _vl_lr_named = _vl_lr.rename(columns=_vl_rename)

        tab1, tab2, tab3 = st.tabs(["📊 Risk & Return", "📅 Return Patterns", "🔗 Relationships"])

        with tab1:
            fig_scatter = chart_risk_return_scatter(_vl_mm, COLOURS, PLOTLY_LAYOUT)
            chart_card(
                "Risk-Return Profile",
                "Each bubble represents a fund. X = volatility, Y = return, size = Sharpe ratio. "
                "Funds in the upper-left quadrant offer the best risk-adjusted profile.",
                fig_scatter
            )

            fig_attr = chart_factor_attribution(_vl_mm, _vl_br, _vl_fr_named, COLOURS, PLOTLY_LAYOUT)
            chart_card(
                "Return Attribution",
                "How much of each fund's return comes from market exposure (beta) vs manager skill (alpha). "
                "A positive alpha bar means the manager added value beyond what the market provided.",
                fig_attr
            )

            fig_dd_timeline = chart_drawdown_recovery(_vl_fr_named, _vl_br, COLOURS, PLOTLY_LAYOUT)
            chart_card(
                "Drawdown Recovery Timeline",
                "Annotated drawdown series showing when each fund reached its worst point and how quickly "
                "it recovered. Labels mark each fund's maximum drawdown.",
                fig_dd_timeline
            )

        with tab2:
            fig_monthly = chart_monthly_heatmap(_vl_fr_named, COLOURS, PLOTLY_LAYOUT)
            chart_card(
                "Monthly Returns Heatmap",
                "Each cell shows the fund's log return for that month. Red = negative, green = positive. "
                "Patterns across rows reveal seasonal or cyclical behaviour.",
                fig_monthly
            )

            fig_dist = chart_return_distribution(_vl_fr_named, COLOURS, PLOTLY_LAYOUT)
            chart_card(
                "Monthly Return Distribution",
                "Violin plots showing the full distribution of monthly returns for each fund. "
                "The box shows the interquartile range, the white dot is the median, and the width shows return frequency.",
                fig_dist
            )

        with tab3:
            try:
                fig_corr = chart_correlation_heatmap(_vl_lr_named, PLOTLY_LAYOUT)
                chart_card(
                    "Return Correlation Matrix",
                    "Pairwise correlations between all fund return series and benchmark. "
                    "Values close to 1.0 indicate funds move together. Lower values suggest diversification potential.",
                    fig_corr
                )
            except Exception as _e:
                st.error(f"Correlation matrix error: {_e}")

            _vl_funds_only = [c for c in _vl_lr_named.columns if c != st.session_state.benchmark_name]
            _vl_corr_vals = _vl_lr_named[_vl_funds_only].corr().values.copy()
            np.fill_diagonal(_vl_corr_vals, np.nan)
            _vl_mean_corr = float(np.nanmean(_vl_corr_vals))
            _vl_corr_note = (
                "High correlation: funds move closely together, which limits diversification benefit."
                if _vl_mean_corr > 0.85
                else "Moderate correlation: some diversification benefit exists across these funds."
            )
            st.markdown(
                f'<div style="background:#F5F3F0;border-radius:12px;padding:1rem 1.25rem;'
                f'margin-top:-0.5rem;margin-bottom:1.5rem;">'
                f'<p style="margin:0;font-size:0.82rem;color:#5A5A5A;">'
                f'Average inter-fund correlation: <strong>{_vl_mean_corr:.3f}</strong>. {_vl_corr_note}'
                f'</p></div>',
                unsafe_allow_html=True
            )

            try:
                fig_rolling_corr = chart_rolling_correlation(_vl_fr_named, _vl_br, COLOURS, PLOTLY_LAYOUT)
                chart_card(
                    "Rolling Correlation with Benchmark",
                    "12-month rolling correlation between each fund and the benchmark. "
                    "A fund consistently near 1.0 tracks the benchmark closely. Dips suggest periods of divergence.",
                    fig_rolling_corr
                )
            except Exception as _e:
                st.error(f"Rolling correlation error: {_e}")

    else:
        st.markdown(
            '<div style="text-align:center;padding:4rem 2rem;background:#F5F3F0;'
            'border-radius:20px;border:2px dashed #D4D2CF;">'
            '<div style="font-size:3rem;margin-bottom:1rem;">🎨</div>'
            '<div style="font-size:1.1rem;font-weight:600;color:#1A1A1A;margin-bottom:0.5rem;">No data yet</div>'
            '<div style="color:#9B9B9B;">Run Analysis in Setup to unlock all visualisations.</div>'
            '</div>',
            unsafe_allow_html=True
        )

# ── 📦 Portfolio ──────────────────────────────────────────────────────────────
elif page == "📦 Portfolio":
    section_header("📦", "CONSTRUCTION", "Portfolio Builder", "portfolio")

    if not st.session_state.get("analysis_run", False):
        st.markdown(
            '<div style="text-align:center;padding:4rem 2rem;background:#F5F3F0;'
            'border-radius:20px;border:2px dashed #D4D2CF;">'
            '<div style="font-size:3rem;margin-bottom:1rem;">📦</div>'
            '<div style="font-size:1.1rem;font-weight:600;color:#1A1A1A;margin-bottom:0.5rem;">No data yet</div>'
            '<div style="color:#9B9B9B;">Run Analysis in Setup to unlock Portfolio Builder.</div>'
            '</div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div style="background:#EFF6FF;border-left:4px solid #154D57;border-radius:0 8px 8px 0;'
            'padding:1rem 1.25rem;margin-bottom:2rem;">'
            '<p style="margin:0;color:#1E40AF;font-size:0.85rem;line-height:1.6;">'
            'Build a blended portfolio from the top-ranked funds and compare its '
            'performance against the benchmark. Adjust allocation weights below.'
            '</p></div>',
            unsafe_allow_html=True
        )

        fund_names = st.session_state.fund_names
        _pt_fr = st.session_state.fund_returns          # tickers as columns
        _pt_tk = st.session_state.tickers
        benchmark_returns = st.session_state.benchmark_returns

        # Map fund_returns columns (tickers) → fund names for arithmetic
        _pt_fr_named = _pt_fr[[t for t in _pt_tk if t in _pt_fr.columns]].copy()
        _pt_fr_named = _pt_fr_named.rename(columns=dict(zip(_pt_tk, fund_names)))

        # Apply pending preset allocations before widgets are instantiated (Bug 1)
        if st.session_state.get("_apply_alloc"):
            for _pf in fund_names:
                _pk = f"_pending_alloc_{_pf.replace(' ', '_')}"
                if _pk in st.session_state:
                    st.session_state[f"alloc_{_pf.replace(' ', '_')}"] = st.session_state[_pk]
                    del st.session_state[_pk]
            del st.session_state["_apply_alloc"]

        # ── Feature 1: Allocation sliders ────────────────────────────────────
        st.markdown(
            '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">PORTFOLIO ALLOCATION</p>',
            unsafe_allow_html=True
        )

        default_equal = round(100 / len(fund_names))
        alloc_cols = st.columns(len(fund_names))
        allocations = {}
        for _ac, fund in zip(alloc_cols, fund_names):
            with _ac:
                fund_colour = COLOURS.get(fund, "#154D57")
                st.markdown(
                    f'<div style="width:100%;height:4px;background:{fund_colour};'
                    f'border-radius:2px;margin-bottom:0.5rem;"></div>',
                    unsafe_allow_html=True
                )
                _skey = f"alloc_{fund.replace(' ', '_')}"
                allocations[fund] = st.slider(
                    fund,
                    min_value=0, max_value=100,
                    value=st.session_state.get(_skey, default_equal),
                    step=5,
                    key=_skey
                )

        total_alloc = sum(allocations.values())

        # Clear active preset if sliders no longer match the preset's expected values
        if st.session_state.get("_active_preset"):
            _ap = st.session_state["_active_preset"]
            _n_funds = len(fund_names)
            _eq_val = 100 // _n_funds
            _sf = []
            _tr = st.session_state.get("topsis_ranking")
            if _tr is not None:
                _sf = [f for f in _tr.sort_values("Rank").index.tolist() if f in fund_names]
            else:
                _mm = st.session_state.get("metrics_matrix")
                if _mm is not None and "Sharpe Ratio" in _mm.index:
                    _sh = _mm.loc["Sharpe Ratio", [f for f in fund_names if f in _mm.columns]]
                    _sf = _sh.sort_values(ascending=False).index.tolist()
            if not _sf:
                _sf = list(fund_names)

            _preset_vals = {}
            if _ap == "equal":
                _rem = 100 - _eq_val * _n_funds
                for _i, _f in enumerate(fund_names):
                    _preset_vals[_f] = _eq_val + (1 if _i < _rem else 0)
            elif _ap == "top2":
                for _f in fund_names:
                    _preset_vals[_f] = 0
                if len(_sf) >= 2:
                    _preset_vals[_sf[0]] = 60
                    _preset_vals[_sf[1]] = 40
            elif _ap == "top1":
                for _f in fund_names:
                    _preset_vals[_f] = 0
                if _sf:
                    _preset_vals[_sf[0]] = 100

            if any(allocations.get(_f, 0) != _preset_vals.get(_f, 0) for _f in fund_names):
                st.session_state["_active_preset"] = None

        bar_segments = "".join([
            f'<div style="width:{allocations[f]}%;background:{COLOURS.get(f,"#154D57")};'
            f'height:100%;display:inline-block;"></div>'
            for f in fund_names if allocations[f] > 0
        ])
        _ta_colour = "#059669" if total_alloc == 100 else "#DC2626"
        _ta_note = "✅ Weights sum to 100%" if total_alloc == 100 else f"⚠️ {100 - total_alloc}% remaining"
        st.markdown(
            f'<div style="margin:1rem 0 0.5rem 0;">'
            f'<div style="display:flex;height:10px;border-radius:5px;overflow:hidden;background:#F0EFEC;">'
            f'{bar_segments}</div>'
            f'<div style="display:flex;justify-content:space-between;margin-top:0.4rem;">'
            f'<span style="font-size:0.75rem;color:#9B9B9B;">Total allocated: '
            f'<strong style="color:{_ta_colour};">{total_alloc}%</strong></span>'
            f'<span style="font-size:0.75rem;color:#9B9B9B;">{_ta_note}</span>'
            f'</div></div>',
            unsafe_allow_html=True
        )

        # Quick presets
        st.markdown(
            '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:0.1em;color:#9B9B9B;margin:1rem 0 0.5rem 0;">QUICK ALLOCATIONS</p>',
            unsafe_allow_html=True
        )
        # Helper: rank funds by TOPSIS if available, else fall back to Sharpe ratio
        def _ranked_funds():
            _tr = st.session_state.get("topsis_ranking")
            if _tr is not None:
                return [f for f in _tr.sort_values("Rank").index.tolist() if f in fund_names]
            _mm = st.session_state.get("metrics_matrix")
            if _mm is not None and "Sharpe Ratio" in _mm.index:
                _sh = _mm.loc["Sharpe Ratio", [f for f in fund_names if f in _mm.columns]]
                return _sh.sort_values(ascending=False).index.tolist()
            return list(fund_names)

        _active_preset = st.session_state.get("_active_preset")
        _q1, _q2, _q3 = st.columns(3)
        _n = len(fund_names)

        with _q1:
            _lbl_eq = "✓ Equal Weight" if _active_preset == "equal" else "Equal Weight"
            if st.button(_lbl_eq, key="alloc_equal", use_container_width=True):
                _eq = 100 // _n
                _rem = 100 - _eq * _n
                for _i, _f in enumerate(fund_names):
                    st.session_state[f"_pending_alloc_{_f.replace(' ', '_')}"] = _eq + (1 if _i == 0 else 0) * _rem
                st.session_state["_apply_alloc"] = True
                st.session_state["_active_preset"] = "equal"
                st.rerun()
        with _q2:
            _lbl_t2 = "✓ Top 2 Focus (60/40)" if _active_preset == "top2" else "Top 2 Focus (60/40)"
            if st.button(_lbl_t2, key="alloc_top2", use_container_width=True):
                _sf = _ranked_funds()
                for _f in fund_names:
                    st.session_state[f"_pending_alloc_{_f.replace(' ', '_')}"] = 0
                if len(_sf) >= 2:
                    st.session_state[f"_pending_alloc_{_sf[0].replace(' ', '_')}"] = 60
                    st.session_state[f"_pending_alloc_{_sf[1].replace(' ', '_')}"] = 40
                elif len(_sf) == 1:
                    st.session_state[f"_pending_alloc_{_sf[0].replace(' ', '_')}"] = 100
                st.session_state["_apply_alloc"] = True
                st.session_state["_active_preset"] = "top2"
                st.rerun()
        with _q3:
            _lbl_t1 = "✓ Top-Ranked Only (100%)" if _active_preset == "top1" else "Top-Ranked Only (100%)"
            if st.button(_lbl_t1, key="alloc_top1", use_container_width=True):
                _sf = _ranked_funds()
                for _f in fund_names:
                    st.session_state[f"_pending_alloc_{_f.replace(' ', '_')}"] = 0
                if _sf:
                    st.session_state[f"_pending_alloc_{_sf[0].replace(' ', '_')}"] = 100
                st.session_state["_apply_alloc"] = True
                st.session_state["_active_preset"] = "top1"
                st.rerun()

        # ── Feature 2 & 3: Portfolio metrics + charts (only when weights sum to 100) ──
        if total_alloc == 100:
            weights_decimal = {f: allocations[f] / 100.0 for f in fund_names}
            portfolio_returns = sum(
                _pt_fr_named[f] * weights_decimal[f] for f in fund_names
            )

            ann_return = (np.exp(portfolio_returns.mean() * 12) - 1) * 100
            ann_vol    = portfolio_returns.std() * np.sqrt(12) * 100
            rf         = st.session_state.get("rf_annual", 0.05)
            sharpe     = (ann_return / 100 - rf) / (ann_vol / 100) if ann_vol > 0 else 0
            wealth     = np.exp(portfolio_returns.cumsum())
            peak       = wealth.cummax()
            max_dd     = ((wealth - peak) / peak * 100).min()
            cumret     = (np.exp(portfolio_returns.sum()) - 1) * 100

            bm_ann_ret = (np.exp(benchmark_returns.mean() * 12) - 1) * 100
            bm_ann_vol = benchmark_returns.std() * np.sqrt(12) * 100
            bm_sharpe  = (bm_ann_ret / 100 - rf) / (bm_ann_vol / 100) if bm_ann_vol > 0 else 0
            bm_wealth  = np.exp(benchmark_returns.cumsum())
            bm_peak    = bm_wealth.cummax()
            bm_max_dd  = ((bm_wealth - bm_peak) / bm_peak * 100).min()
            bm_cumret  = (np.exp(benchmark_returns.sum()) - 1) * 100

            st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)
            st.markdown(
                '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                'letter-spacing:0.1em;color:#9B9B9B;margin-bottom:1rem;">PORTFOLIO vs BENCHMARK</p>',
                unsafe_allow_html=True
            )

            _met_cols = st.columns(5)
            _metrics = [
                ("Cumulative Return", f"{cumret:.1f}%",    f"{bm_cumret:.1f}%",  cumret > bm_cumret),
                ("Ann. Return",       f"{ann_return:.2f}%", f"{bm_ann_ret:.2f}%", ann_return > bm_ann_ret),
                ("Ann. Volatility",   f"{ann_vol:.2f}%",   f"{bm_ann_vol:.2f}%", ann_vol < bm_ann_vol),
                ("Sharpe Ratio",      f"{sharpe:.3f}",      f"{bm_sharpe:.3f}",   sharpe > bm_sharpe),
                ("Max Drawdown",      f"{max_dd:.2f}%",     f"{bm_max_dd:.2f}%",  max_dd > bm_max_dd),
            ]
            for _mc, (_label, _pv, _bv, _wins) in zip(_met_cols, _metrics):
                with _mc:
                    _acc = "#059669" if _wins else "#DC2626"
                    _arr = "↑" if _wins else "↓"
                    st.markdown(
                        f'<div style="background:#FFFFFF;border-radius:16px;padding:1.1rem 1.25rem;'
                        f'box-shadow:0 2px 10px rgba(0,0,0,0.06);border-top:3px solid {_acc};">'
                        f'<div style="font-size:0.65rem;font-weight:700;text-transform:uppercase;'
                        f'letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.5rem;">{_label}</div>'
                        f'<div style="font-size:1.4rem;font-weight:700;color:#1A1A1A;'
                        f'letter-spacing:-0.02em;">{_pv}</div>'
                        f'<div style="font-size:0.72rem;color:{_acc};margin-top:0.3rem;font-weight:600;">'
                        f'{_arr} vs {_bv} benchmark</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

            # Chart A  -  Cumulative return
            st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)
            _fig_port = go.Figure()
            for _f in fund_names:
                if allocations[_f] > 0:
                    _cum = (np.exp(_pt_fr_named[_f].cumsum()) - 1) * 100
                    _fig_port.add_trace(go.Scatter(
                        x=_cum.index, y=_cum.values,
                        name=f"{_f} ({allocations[_f]}%)",
                        mode="lines",
                        line=dict(color=COLOURS.get(_f, "#9B9B9B"), width=1.5, dash="dot"),
                        opacity=0.6,
                        hovertemplate=f"<b>{_f}</b><br>%{{x}}<br>Return: %{{y:.1f}}%<extra></extra>"
                    ))
            _bm_cum = (np.exp(benchmark_returns.cumsum()) - 1) * 100
            _fig_port.add_trace(go.Scatter(
                x=_bm_cum.index, y=_bm_cum.values,
                name="Benchmark",
                mode="lines",
                line=dict(color="#9B9B9B", width=2, dash="dash"),
                hovertemplate="<b>Benchmark</b><br>%{x}<br>Return: %{y:.1f}%<extra></extra>"
            ))
            _port_cum = (np.exp(portfolio_returns.cumsum()) - 1) * 100
            _fig_port.add_trace(go.Scatter(
                x=_port_cum.index, y=_port_cum.values,
                name="Blended Portfolio",
                mode="lines",
                line=dict(color="#154D57", width=3.5),
                hovertemplate="<b>Blended Portfolio</b><br>%{x}<br>Return: %{y:.1f}%<extra></extra>"
            ))
            _fig_port.add_hline(y=0, line_dash="dot", line_color="#9B9B9B", opacity=0.4)
            _fig_port.update_layout(**{k: v for k, v in PLOTLY_LAYOUT.items() if k != "title"})
            _fig_port.update_layout(
                xaxis=dict(title="Date", gridcolor="#F0EFEC"),
                yaxis=dict(title="Cumulative Return (%)", gridcolor="#F0EFEC"),
                height=460,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            chart_card(
                "Cumulative Return Comparison",
                "The bold teal line shows the blended portfolio. Dotted lines show individual fund contributions. Dashed grey = benchmark.",
                _fig_port
            )

            # Chart B  -  Drawdown comparison
            _fig_dd_port = go.Figure()
            for _f in fund_names:
                if allocations[_f] > 0:
                    _w = np.exp(_pt_fr_named[_f].cumsum())
                    _pk = _w.cummax()
                    _dd = (_w - _pk) / _pk * 100
                    _fig_dd_port.add_trace(go.Scatter(
                        x=_dd.index, y=_dd.values,
                        name=_f,
                        mode="lines",
                        line=dict(color=COLOURS.get(_f, "#9B9B9B"), width=1.5, dash="dot"),
                        opacity=0.5,
                        hovertemplate=f"<b>{_f}</b><br>%{{x}}<br>DD: %{{y:.2f}}%<extra></extra>"
                    ))
            _bm_w = np.exp(benchmark_returns.cumsum())
            _bm_pk = _bm_w.cummax()
            _bm_dd = (_bm_w - _bm_pk) / _bm_pk * 100
            _fig_dd_port.add_trace(go.Scatter(
                x=_bm_dd.index, y=_bm_dd.values,
                name="Benchmark",
                mode="lines",
                line=dict(color="#9B9B9B", width=2, dash="dash"),
                hovertemplate="<b>Benchmark</b><br>%{x}<br>DD: %{y:.2f}%<extra></extra>"
            ))
            _p_w = np.exp(portfolio_returns.cumsum())
            _p_pk = _p_w.cummax()
            _p_dd = (_p_w - _p_pk) / _p_pk * 100
            _fig_dd_port.add_trace(go.Scatter(
                x=_p_dd.index, y=_p_dd.values,
                name="Blended Portfolio",
                mode="lines",
                line=dict(color="#154D57", width=3.5),
                fill="tozeroy",
                fillcolor="rgba(21,77,87,0.08)",
                hovertemplate="<b>Blended Portfolio</b><br>%{x}<br>DD: %{y:.2f}%<extra></extra>"
            ))
            _fig_dd_port.add_hline(y=0, line_dash="dash", line_color="#9B9B9B", opacity=0.3)
            _fig_dd_port.update_layout(**{k: v for k, v in PLOTLY_LAYOUT.items() if k != "title"})
            _fig_dd_port.update_layout(
                xaxis=dict(title="Date", gridcolor="#F0EFEC"),
                yaxis=dict(title="Drawdown (%)", gridcolor="#F0EFEC"),
                height=400,
            )
            chart_card(
                "Portfolio Drawdown vs Benchmark",
                "How the blended portfolio's peak-to-trough decline compares to individual funds and the benchmark.",
                _fig_dd_port
            )

            # ── Feature 4: Monte Carlo ────────────────────────────────────────
            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
            st.markdown(
                '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                'letter-spacing:0.1em;color:#9B9B9B;margin-bottom:0.5rem;">MONTE CARLO SIMULATION</p>',
                unsafe_allow_html=True
            )
            st.markdown(
                '<p style="font-size:0.85rem;color:#5A5A5A;margin-bottom:1rem;">'
                'Simulates 1,000 random weight combinations to show which fund wins most often '
                'under different allocations.</p>',
                unsafe_allow_html=True
            )

            if st.button("▶ Run Monte Carlo (1,000 simulations)", key="run_monte_carlo", use_container_width=True):
                import random
                _n_sims  = 1000
                _n_funds = len(fund_names)
                _win_counts = {f: 0 for f in fund_names}
                _sim_results = []
                with st.spinner("Running 1,000 simulations…"):
                    for _ in range(_n_sims):
                        _raw   = [random.random() for _ in range(_n_funds)]
                        _tot   = sum(_raw)
                        _w     = [r / _tot for r in _raw]
                        _port_r = sum(_pt_fr_named[f] * _w[i] for i, f in enumerate(fund_names))
                        _sh = (
                            (np.exp(_port_r.mean() * 12) - 1) - rf
                        ) / (_port_r.std() * np.sqrt(12))
                        _sim_results.append({
                            "weights": dict(zip(fund_names, _w)),
                            "sharpe":  float(_sh),
                            "return":  float((np.exp(_port_r.mean() * 12) - 1) * 100),
                            "vol":     float(_port_r.std() * np.sqrt(12) * 100),
                        })
                    _best_sim = max(_sim_results, key=lambda x: x["sharpe"])
                    _top_sims = sorted(_sim_results, key=lambda x: x["sharpe"], reverse=True)[:100]
                    for _s in _top_sims:
                        _top_f = max(_s["weights"], key=_s["weights"].get)
                        _win_counts[_top_f] += 1
                st.session_state.mc_results    = _sim_results
                st.session_state.mc_best       = _best_sim
                st.session_state.mc_win_counts = _win_counts
                st.rerun()

            if st.session_state.get("mc_results"):
                _sim_results = st.session_state.mc_results
                _best_sim    = st.session_state.mc_best
                _win_counts  = st.session_state.mc_win_counts

                _rets_l   = [s["return"] for s in _sim_results]
                _vols_l   = [s["vol"]    for s in _sim_results]
                _sharps_l = [s["sharpe"] for s in _sim_results]

                _fig_mc = go.Figure()
                _fig_mc.add_trace(go.Scatter(
                    x=_vols_l, y=_rets_l,
                    mode="markers",
                    marker=dict(
                        size=4,
                        color=_sharps_l,
                        colorscale="Teal",
                        showscale=True,
                        colorbar=dict(title="Sharpe", thickness=12, len=0.6),
                        opacity=0.6,
                    ),
                    hovertemplate="Return: %{y:.1f}%<br>Vol: %{x:.1f}%<extra></extra>",
                    name="Simulations",
                    showlegend=False,
                ))
                _fig_mc.add_trace(go.Scatter(
                    x=[_best_sim["vol"]], y=[_best_sim["return"]],
                    mode="markers+text",
                    marker=dict(size=16, color="#154D57", symbol="star",
                                line=dict(width=2, color="#FFFFFF")),
                    text=["Best Sharpe"],
                    textposition="top right",
                    textfont=dict(size=11, color="#154D57"),
                    name="Best Sharpe",
                ))
                _fig_mc.add_trace(go.Scatter(
                    x=[ann_vol], y=[ann_return],
                    mode="markers+text",
                    marker=dict(size=14, color="#C8A96E", symbol="diamond",
                                line=dict(width=2, color="#FFFFFF")),
                    text=["Your Portfolio"],
                    textposition="top right",
                    textfont=dict(size=11, color="#C8A96E"),
                    name="Your Portfolio",
                ))
                _fig_mc.update_layout(**{k: v for k, v in PLOTLY_LAYOUT.items() if k != "title"})
                _fig_mc.update_layout(
                    xaxis=dict(title="Annualised Volatility (%)", gridcolor="#F0EFEC"),
                    yaxis=dict(title="Annualised Return (%)", gridcolor="#F0EFEC"),
                    height=500,
                )
                chart_card(
                    "Efficient Frontier Simulation",
                    "Each dot is a random portfolio. Colour = Sharpe ratio (darker = better). ★ = highest Sharpe found. ◆ = your current allocation.",
                    _fig_mc
                )

                st.markdown(
                    '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                    'letter-spacing:0.1em;color:#9B9B9B;margin:1rem 0 0.5rem 0;">'
                    'DOMINANCE IN TOP 10% SIMULATIONS</p>',
                    unsafe_allow_html=True
                )
                _wc_cols = st.columns(len(fund_names))
                for _wc, _f in zip(_wc_cols, fund_names):
                    with _wc:
                        _wins = _win_counts.get(_f, 0)
                        _fc = COLOURS.get(_f, "#154D57")
                        st.markdown(
                            f'<div style="background:#FFFFFF;border-radius:14px;padding:1rem;'
                            f'box-shadow:0 1px 6px rgba(0,0,0,0.05);text-align:center;'
                            f'border-top:3px solid {_fc};">'
                            f'<div style="font-size:0.65rem;font-weight:700;text-transform:uppercase;'
                            f'letter-spacing:0.08em;color:#9B9B9B;margin-bottom:0.4rem;">{_f}</div>'
                            f'<div style="font-size:1.6rem;font-weight:700;color:#1A1A1A;">{_wins}</div>'
                            f'<div style="font-size:0.72rem;color:#9B9B9B;">times dominant</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                st.markdown(
                    '<p style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
                    'letter-spacing:0.1em;color:#9B9B9B;margin:1.5rem 0 0.5rem 0;">'
                    'OPTIMAL WEIGHTS (HIGHEST SHARPE SIMULATION)</p>',
                    unsafe_allow_html=True
                )
                _bw = _best_sim["weights"]
                _ow_cols = st.columns(len(fund_names))
                for _ow, _f in zip(_ow_cols, fund_names):
                    with _ow:
                        _wp = _bw.get(_f, 0) * 100
                        _fc = COLOURS.get(_f, "#154D57")
                        st.markdown(
                            f'<div style="background:#FFFFFF;border-radius:14px;padding:1rem;'
                            f'box-shadow:0 1px 6px rgba(0,0,0,0.05);text-align:center;'
                            f'border-left:4px solid {_fc};">'
                            f'<div style="font-size:0.65rem;font-weight:700;text-transform:uppercase;'
                            f'letter-spacing:0.08em;color:#9B9B9B;margin-bottom:0.4rem;">{_f}</div>'
                            f'<div style="font-size:1.4rem;font-weight:700;color:#1A1A1A;">{_wp:.1f}%</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

        else:
            st.markdown(
                '<div style="text-align:center;padding:3rem 2rem;background:#F5F3F0;'
                'border-radius:20px;border:2px dashed #D4D2CF;margin-top:1rem;">'
                '<div style="font-size:2.5rem;margin-bottom:1rem;">⚖️</div>'
                '<div style="font-size:1rem;font-weight:600;color:#1A1A1A;margin-bottom:0.5rem;">'
                'Weights don\'t sum to 100%</div>'
                '<div style="color:#9B9B9B;font-size:0.88rem;">'
                'Adjust the sliders above until total allocation reaches exactly 100%.'
                '</div></div>',
                unsafe_allow_html=True
            )